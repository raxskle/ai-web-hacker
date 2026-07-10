#!/usr/bin/env python3
"""批量分析关键词指标（SEM keywords.GetInfo + SIM suggest）并输出标准词表。

Usage:
  python3 analyze-words/_internal/scripts/analyze_words.py run --keyword "image to text"
  python3 analyze-words/_internal/scripts/analyze_words.py run --keyword-file /tmp/keywords.txt
  python3 analyze-words/_internal/scripts/analyze_words.py run --keywords-csv "image to text,ocr online"
  python3 analyze-words/_internal/scripts/analyze_words.py run --input-table standard-word-analysis/sample/standard-word-table.v1.sample.json
  python3 analyze-words/_internal/scripts/analyze_words.py rebuild-reports
  python3 analyze-words/_internal/scripts/analyze_words.py validate-report
"""

from __future__ import annotations

import argparse
import json
import re
import time
from datetime import datetime
from pathlib import Path
from shutil import copyfile
from typing import Dict, List, Optional, Sequence, Tuple
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

PROJECT_DIR = Path(__file__).resolve().parents[2]
REPO_ROOT = PROJECT_DIR.parent
LOCAL_SERVICE_TOKEN_PATH = PROJECT_DIR / "local-service" / "bridge_token.txt"
LOCAL_SERVICE_GMITM_PATH = PROJECT_DIR / "local-service" / "__gmitm.txt"
LEGACY_LOCAL_SERVICE_GMITM_PATH = PROJECT_DIR / "local-service" / "gmitm.txt"

DEFAULT_API_BASE = "http://127.0.0.1:17311"
DEFAULT_ENDPOINT = "/sem/kwogw/v2/webapi/keywords.GetInfo"
DEFAULT_SIM_ENDPOINT = "/sim/api/KeywordGenerator/google/suggest"

SIM_ROWS_PER_PAGE = 5
SIM_TYPE = "Related"
SIM_SORT_FIELD = "score"
SIM_ASC = False
SIM_COUNTRY = "999"
SIM_LATEST = "28d"
SIM_WEB_SOURCE = "Total"
SIM_IS_WINDOW = True
SIM_PAGE = 1

STANDARD_WORD_TABLE_VERSION = "v1"
STANDARD_WORD_TABLE_SPEC_PATH = REPO_ROOT / "standard-word-analysis" / "spec" / f"standard-word-table.{STANDARD_WORD_TABLE_VERSION}.json"
SCORE_FORMULA = "simWindowVolume * simCpc / simKd"

MAX_RETRIES = 2
RETRY_BACKOFF_SECONDS = [0.8, 1.6]
REQUEST_TIMEOUT_SECONDS = 180

TOP_PREVIEW_ROWS = 30
SNAPSHOT_RE = re.compile(r"^snapshot-(\d{8}-\d{6})\.json$")
SPACE_RE = re.compile(r"\s+")
BULLET_PREFIX_RE = re.compile(r"^[\-\*•·]+\s*")
KEYWORD_SPLIT_RE = re.compile(r"[,，;；、|\n\r\t]+")

DEFAULT_COLUMN_PADDING = 4
DEFAULT_COLUMN_MIN_WIDTH = 12
DEFAULT_COLUMN_MAX_WIDTH = 72

REQUIRED_SECTIONS = [
    "## 摘要",
    "## 抓取概览",
    "## 聚合结果概览",
    "## 关键词结果预览",
    "## 标准词表预览",
    "## 失败关键词",
    "## 产物路径",
    "## 备注",
]

REQUIRED_REMARKS = [
    "globalVolume = sum(volume)",
    "globalCpcAvg / globalDifficultyAvg 仅统计非 null 项",
    "标准词表 score 仅由 SIM 字段计算：simWindowVolume * simCpc / simKd",
    "SIM Suggest 使用 rowsPerPage=5、type=Related、sort=score、asc=false（其它参数走默认口径）",
    "SIM 请求失败或未命中对应关键词时：继续执行并将 SIM 字段置空",
    "任一关键词 SEM 抓取失败时默认整次失败，不落任何新产物",
]

SOURCE_PRESENCE_LABEL_TO_ENUM = {
    "both": "both",
    "both（sim+sem）": "both",
    "sim_only": "sim_only",
    "sim_only（仅 sim）": "sim_only",
    "sem_only": "sem_only",
    "sem_only（仅 sem）": "sem_only",
}

SOURCE_PRESENCE_ENUM_TO_LABEL = {
    "both": "both（SIM+SEM）",
    "sim_only": "sim_only（仅 SIM）",
    "sem_only": "sem_only（仅 SEM）",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="批量调用 keywords.GetInfo + SIM Suggest 并聚合关键词指标")
    subparsers = parser.add_subparsers(dest="command", required=True)

    run_parser = subparsers.add_parser("run", help="抓取 + 聚合 + 快照 + 报告 + 标准词表")
    run_parser.add_argument("--keyword", action="append", default=[], help="关键词，可重复传入；支持以逗号/中文逗号/分号分隔多个词")
    run_parser.add_argument(
        "--keyword-file",
        action="append",
        default=[],
        help="关键词文件（UTF-8，每行一个关键词，支持 # 注释）",
    )
    run_parser.add_argument(
        "--keywords-csv",
        action="append",
        default=[],
        help="逗号分隔关键词列表，可重复传入",
    )
    run_parser.add_argument(
        "--input-table",
        action="append",
        default=[],
        help="标准词表输入（.json/.xlsx），可重复传入",
    )

    run_parser.add_argument("--api-base", default=DEFAULT_API_BASE)
    run_parser.add_argument("--endpoint", default=DEFAULT_ENDPOINT)
    run_parser.add_argument("--sim-endpoint", default=DEFAULT_SIM_ENDPOINT)
    run_parser.add_argument("--token-path", default=str(LOCAL_SERVICE_TOKEN_PATH))
    run_parser.add_argument("--gmitm-path", default=str(LOCAL_SERVICE_GMITM_PATH))

    run_parser.add_argument("--device", type=int, default=0)
    run_parser.add_argument("--currency", default="USD")
    run_parser.add_argument("--database", default="us")
    run_parser.add_argument("--location", type=int, default=0)
    run_parser.add_argument("--date", default="")
    run_parser.add_argument("--timeout-ms", type=int, default=45000)
    run_parser.add_argument("--wait-timeout-ms", type=int, default=120000)

    run_parser.add_argument("--data-dir", default=str(PROJECT_DIR / "data"))
    run_parser.add_argument("--snapshot-dir", default=str(PROJECT_DIR / "_internal" / "snapshots"))
    run_parser.add_argument("--report-dir", default=str(PROJECT_DIR / "report" / "history"))
    run_parser.add_argument("--latest-report-path", default=str(PROJECT_DIR / "report" / "latest.md"))
    run_parser.add_argument("--latest-xlsx-path", default=str(PROJECT_DIR / "report" / "latest.xlsx"))
    run_parser.add_argument("--latest-table-json-path", default=str(PROJECT_DIR / "report" / "latest.json"))

    rebuild_parser = subparsers.add_parser("rebuild-reports", help="根据快照重建历史报告（Markdown + 标准词表）")
    rebuild_parser.add_argument("--snapshot-dir", default=str(PROJECT_DIR / "_internal" / "snapshots"))
    rebuild_parser.add_argument("--report-dir", default=str(PROJECT_DIR / "report" / "history"))
    rebuild_parser.add_argument("--latest-report-path", default=str(PROJECT_DIR / "report" / "latest.md"))
    rebuild_parser.add_argument("--latest-xlsx-path", default=str(PROJECT_DIR / "report" / "latest.xlsx"))
    rebuild_parser.add_argument("--latest-table-json-path", default=str(PROJECT_DIR / "report" / "latest.json"))

    validate_parser = subparsers.add_parser("validate-report", help="校验 Markdown 报告结构与标准词表产物")
    validate_parser.add_argument("--report", default=str(PROJECT_DIR / "report" / "latest.md"))
    validate_parser.add_argument("--xlsx", default=str(PROJECT_DIR / "report" / "latest.xlsx"))
    validate_parser.add_argument("--table-json", default=str(PROJECT_DIR / "report" / "latest.json"))

    return parser.parse_args()


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d-%H%M%S")


def safe_float(value) -> float:
    try:
        if value is None or value == "":
            return 0.0
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def to_display_number(value, digits: int = 2) -> str:
    if value is None:
        return "-"
    number = safe_float(value)
    if abs(number - round(number)) < 1e-9:
        return str(int(round(number)))
    return f"{number:.{digits}f}"


def load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def dump_json(path: Path, payload: dict) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _normalize_export_column(raw: dict) -> dict:
    if not isinstance(raw, dict):
        raise RuntimeError("标准词表列配置项必须是对象")

    header = str(raw.get("header") or "").strip()
    field = str(raw.get("field") or "").strip()
    if not header or not field:
        raise RuntimeError("标准词表列配置缺少 header/field")

    column = {
        "header": header,
        "field": field,
    }

    transform = raw.get("transform")
    if transform not in (None, ""):
        column["transform"] = str(transform)

    number_format = raw.get("number_format")
    if number_format not in (None, ""):
        column["number_format"] = str(number_format)

    width_profile = raw.get("width_profile")
    if isinstance(width_profile, dict):
        normalized_width: dict = {}
        if width_profile.get("fixed") is not None:
            normalized_width["fixed"] = width_profile.get("fixed")
        if width_profile.get("min") is not None:
            normalized_width["min"] = width_profile.get("min")
        if width_profile.get("max") is not None:
            normalized_width["max"] = width_profile.get("max")
        if normalized_width:
            column["width_profile"] = normalized_width

    return column


def load_standard_word_table_spec() -> dict:
    if not STANDARD_WORD_TABLE_SPEC_PATH.exists():
        raise RuntimeError(
            "标准词表规范缺失: "
            f"{STANDARD_WORD_TABLE_SPEC_PATH}。"
            "请先补齐 standard-word-analysis/spec/standard-word-table.v1.json。"
        )

    spec = load_json(STANDARD_WORD_TABLE_SPEC_PATH)
    if not isinstance(spec, dict):
        raise RuntimeError("标准词表规范格式错误：顶层必须是对象")

    version = str(spec.get("version") or "").strip()
    if version != STANDARD_WORD_TABLE_VERSION:
        raise RuntimeError(
            f"标准词表版本不匹配：expected={STANDARD_WORD_TABLE_VERSION} actual={version or '-'}"
        )

    raw_columns = spec.get("excelColumns")
    if not isinstance(raw_columns, list) or not raw_columns:
        raise RuntimeError("标准词表规范缺少 excelColumns 数组")

    columns = [_normalize_export_column(item) for item in raw_columns]
    fields = [str(column.get("field") or "") for column in columns]
    if len(fields) != len(set(fields)):
        raise RuntimeError("标准词表规范字段重复：excelColumns.field 必须唯一")

    score_header = str(spec.get("scoreColumnHeader") or "").strip()
    if not score_header:
        raise RuntimeError("标准词表规范缺少 scoreColumnHeader")

    score_field = str(spec.get("scoreField") or "score").strip() or "score"
    score_candidates = [column for column in columns if column.get("field") == score_field]
    if len(score_candidates) != 1:
        raise RuntimeError("标准词表规范中 scoreField 对应列必须且仅能有一列")

    if score_candidates[0].get("header") != score_header:
        raise RuntimeError("标准词表规范 scoreColumnHeader 与 scoreField 对应列表头不一致")

    score_formula = str(spec.get("scoreFormula") or "").strip() or SCORE_FORMULA
    if score_formula != SCORE_FORMULA:
        raise RuntimeError(
            "标准词表规范 scoreFormula 与 analyze-words 当前公式不一致："
            f"expected={SCORE_FORMULA} actual={score_formula}"
        )

    return {
        "version": version,
        "columns": columns,
        "scoreColumnHeader": score_header,
        "scoreField": score_field,
        "scoreFormula": score_formula,
    }


def get_standard_word_table_spec() -> dict:
    return STANDARD_WORD_TABLE_SPEC


def get_keywords_export_columns() -> List[dict]:
    return list(get_standard_word_table_spec()["columns"])


def get_keywords_export_headers() -> List[str]:
    return [column["header"] for column in get_keywords_export_columns()]


def get_header_for_field(field_name: str) -> str:
    for column in get_keywords_export_columns():
        if column.get("field") == field_name:
            return str(column.get("header") or "")
    raise RuntimeError(f"标准词表缺少字段: {field_name}")


def get_score_column_index_from_headers(headers: Sequence[str]) -> int:
    score_header = get_standard_word_table_spec()["scoreColumnHeader"]
    try:
        return list(headers).index(score_header)
    except ValueError as exc:
        raise RuntimeError(f"标准词表缺少 score 列: {score_header}") from exc


def _require_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "缺少依赖 openpyxl。请先执行 `pip3 install openpyxl` 再运行。"
        ) from exc
    return Workbook, Font, Alignment, get_column_letter, load_workbook


MULTILINE_EXPORT_FIELDS = {
    "group",
    "correspondingDomain",
}


def _format_multiline_export_text(field: str, value):
    if value is None:
        return value
    text = str(value).strip()
    if not text:
        return text
    if field not in MULTILINE_EXPORT_FIELDS:
        return text
    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    normalized = re.sub(r"\s*\|\s*", "\n", normalized)
    normalized = normalized.replace(" / ", "\n")
    lines = [line.strip() for line in normalized.split("\n") if line.strip()]
    return "\n".join(lines) if lines else ""


def normalize_keyword_text(text: str) -> str:
    return SPACE_RE.sub(" ", str(text or "").strip().lower())


def _clean_keyword_piece(text: str) -> str:
    value = str(text or "").strip()
    value = BULLET_PREFIX_RE.sub("", value)
    value = value.strip("\"'“”‘’` ")
    return value


def _split_keywords_from_text(text: str) -> List[str]:
    raw = str(text or "").strip()
    if not raw:
        return []

    normalized = raw.replace("　", " ").strip()
    if "：" in normalized:
        left, right = normalized.split("：", 1)
        if "词" in left or "keyword" in left.lower():
            normalized = right.strip() or normalized
    elif ":" in normalized:
        left, right = normalized.split(":", 1)
        if "词" in left or "keyword" in left.lower():
            normalized = right.strip() or normalized

    parts = [_clean_keyword_piece(part) for part in KEYWORD_SPLIT_RE.split(normalized)]
    result = [part for part in parts if part]
    if result:
        return result

    cleaned = _clean_keyword_piece(normalized)
    return [cleaned] if cleaned else []


def read_text_required(path: Path, label: str) -> str:
    if not path.exists():
        raise RuntimeError(f"{label} 不存在: {path}")
    value = path.read_text(encoding="utf-8").strip()
    if not value:
        raise RuntimeError(f"{label} 为空: {path}")
    return value


def read_gmitm(path: Path) -> str:
    if path.exists():
        return read_text_required(path, "gmitm 文件")
    if path.name == "gmitm.txt" and LEGACY_LOCAL_SERVICE_GMITM_PATH.exists():
        return read_text_required(LEGACY_LOCAL_SERVICE_GMITM_PATH, "gmitm 文件")
    raise RuntimeError(
        f"gmitm 文件不存在: {path}。请在 analyze-words/local-service/__gmitm.txt 写入当前会话 __gmitm 值。"
    )


def _read_keywords_from_file(path: Path) -> List[str]:
    if not path.exists():
        raise RuntimeError(f"关键词文件不存在: {path}")
    lines = path.read_text(encoding="utf-8").splitlines()
    keywords: List[str] = []
    for raw in lines:
        text = str(raw).strip()
        if not text:
            continue
        if text.startswith("#"):
            continue
        keywords.extend(_split_keywords_from_text(text))
    return keywords


def _coerce_optional_number(value) -> Optional[float]:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _normalize_source_presence(value) -> Optional[str]:
    if value in (None, ""):
        return None
    key = str(value).strip().lower()
    if not key:
        return None
    return SOURCE_PRESENCE_LABEL_TO_ENUM.get(key)


def _normalize_seed_row(raw: dict, *, source_path: Path) -> Optional[dict]:
    keyword = str(raw.get("keyword") or "").strip()
    if not keyword:
        return None

    row = {
        "keyword": keyword,
        "correspondingDomain": str(raw.get("correspondingDomain") or "").strip() or None,
        "group": str(raw.get("group") or "").strip() or None,
        "sourcePresence": _normalize_source_presence(raw.get("sourcePresence")),
        "score": _coerce_optional_number(raw.get("score")),
        "simWindowVolume": _coerce_optional_number(raw.get("simWindowVolume")),
        "simKd": _coerce_optional_number(raw.get("simKd")),
        "simCpc": _coerce_optional_number(raw.get("simCpc")),
        "semVolume": _coerce_optional_number(raw.get("semVolume")),
        "semKd": _coerce_optional_number(raw.get("semKd")),
        "semCpc": _coerce_optional_number(raw.get("semCpc")),
        "gefeiKD": _coerce_optional_number(raw.get("gefeiKD")),
        "_sourcePath": str(source_path),
    }
    return row


def read_standard_table_json(path: Path) -> List[dict]:
    payload = load_json(path)
    version = str(payload.get("version") or "").strip()
    if version != STANDARD_WORD_TABLE_VERSION:
        raise RuntimeError(
            f"标准词表 JSON 版本不匹配: expected={STANDARD_WORD_TABLE_VERSION} actual={version or '-'} path={path}"
        )

    rows = payload.get("rows")
    if not isinstance(rows, list):
        raise RuntimeError(f"标准词表 JSON rows 非数组: {path}")

    normalized_rows: List[dict] = []
    for item in rows:
        if not isinstance(item, dict):
            continue
        normalized = _normalize_seed_row(item, source_path=path)
        if normalized is not None:
            normalized_rows.append(normalized)
    return normalized_rows


def read_standard_table_xlsx(path: Path) -> List[dict]:
    _Workbook, _Font, _Alignment, _get_column_letter, load_workbook = _require_openpyxl()
    workbook = load_workbook(path, data_only=True)
    try:
        if "keywords" not in workbook.sheetnames:
            raise RuntimeError(f"标准词表 Excel 缺少 keywords sheet: {path}")
        sheet = workbook["keywords"]

        expected_headers = get_keywords_export_headers()
        actual_headers = [sheet.cell(row=1, column=index).value for index in range(1, sheet.max_column + 1)]
        if actual_headers != expected_headers:
            details: List[str] = []
            max_len = max(len(expected_headers), len(actual_headers))
            for idx in range(1, max_len + 1):
                expected = expected_headers[idx - 1] if idx <= len(expected_headers) else None
                actual = actual_headers[idx - 1] if idx <= len(actual_headers) else None
                if expected != actual:
                    details.append(f"col#{idx} expected={expected!r} actual={actual!r}")
            raise RuntimeError("标准词表 Excel 表头不匹配: " + "; ".join(details))

        columns = get_keywords_export_columns()
        result: List[dict] = []
        for row_index in range(2, sheet.max_row + 1):
            row_payload: dict = {}
            is_empty = True
            for col_index, column in enumerate(columns, start=1):
                field = str(column.get("field") or "")
                cell_value = sheet.cell(row=row_index, column=col_index).value
                if cell_value not in (None, ""):
                    is_empty = False
                row_payload[field] = cell_value

            if is_empty:
                continue
            normalized = _normalize_seed_row(row_payload, source_path=path)
            if normalized is not None:
                result.append(normalized)
        return result
    finally:
        workbook.close()


def _load_table_rows(path: Path) -> List[dict]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        return read_standard_table_json(path)
    if suffix == ".xlsx":
        return read_standard_table_xlsx(path)
    raise RuntimeError(f"不支持的标准词表格式: {path}")


def _merge_seed_row(base_row: dict, incoming_row: dict) -> dict:
    merged = dict(base_row)
    for key, value in incoming_row.items():
        if key.startswith("_"):
            continue
        if value in (None, ""):
            continue
        merged[key] = value
    return merged


def collect_keywords(args: argparse.Namespace) -> Tuple[List[dict], dict]:
    table_rows: List[dict] = []
    for table_arg in args.input_table or []:
        table_path = Path(table_arg).resolve()
        table_rows.extend(_load_table_rows(table_path))

    raw_keywords: List[str] = []
    for text in args.keyword or []:
        raw_keywords.extend(_split_keywords_from_text(text))

    for file_arg in args.keyword_file or []:
        file_path = Path(file_arg).resolve()
        raw_keywords.extend(_read_keywords_from_file(file_path))

    for csv_arg in args.keywords_csv or []:
        text = str(csv_arg or "").strip()
        if not text:
            continue
        raw_keywords.extend(_split_keywords_from_text(text))

    deduped: List[dict] = []
    seen: Dict[str, int] = {}

    def upsert_keyword(keyword: str, *, input_source: str, seed_row: Optional[dict] = None) -> None:
        display_keyword = str(keyword or "").strip()
        normalized = normalize_keyword_text(display_keyword)
        if not normalized:
            return

        if normalized not in seen:
            base_seed = {
                "keyword": display_keyword,
                "correspondingDomain": None,
                "group": display_keyword,
                "sourcePresence": None,
                "score": None,
                "simWindowVolume": None,
                "simKd": None,
                "simCpc": None,
                "semVolume": None,
                "semKd": None,
                "semCpc": None,
                "gefeiKD": None,
            }
            if seed_row is not None:
                base_seed = _merge_seed_row(base_seed, seed_row)

            seen[normalized] = len(deduped)
            deduped.append(
                {
                    "keyword": display_keyword,
                    "keywordNormalized": normalized,
                    "seedRow": base_seed,
                    "inputSources": [input_source],
                }
            )
            return

        index = seen[normalized]
        item = deduped[index]
        if input_source not in item["inputSources"]:
            item["inputSources"].append(input_source)
        if seed_row is not None:
            item["seedRow"] = _merge_seed_row(item["seedRow"], seed_row)

    for row in table_rows:
        upsert_keyword(str(row.get("keyword") or ""), input_source="input-table", seed_row=row)

    for keyword in raw_keywords:
        upsert_keyword(keyword, input_source="keyword")

    if not deduped:
        raise SystemExit("未获取到有效关键词。请使用 --keyword / --keyword-file / --keywords-csv / --input-table 提供输入。")

    input_overview = {
        "rawKeywordCount": len(raw_keywords),
        "tableRowCount": len(table_rows),
        "dedupedKeywordCount": len(deduped),
        "inputSources": sorted({source for item in deduped for source in item.get("inputSources", [])}),
    }
    return deduped, input_overview


def list_snapshot_files(snapshot_dir: Path) -> List[Path]:
    if not snapshot_dir.exists():
        return []
    files = []
    for path in snapshot_dir.iterdir():
        if path.is_file() and SNAPSHOT_RE.match(path.name):
            files.append(path)
    files.sort(key=lambda p: p.name)
    return files


def _post_json(url: str, headers: dict, payload: dict, timeout_seconds: int) -> Tuple[int, str]:
    req = Request(
        url,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers=headers,
        method="POST",
    )
    try:
        with urlopen(req, timeout=timeout_seconds) as resp:
            body = resp.read().decode("utf-8", errors="replace")
            return int(getattr(resp, "status", 0) or 0), body
    except HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace") if hasattr(exc, "read") else ""
        return int(getattr(exc, "code", 0) or 0), body


def post_local_service(*, api_url: str, headers: dict, payload: dict, expect_jsonrpc: bool = False) -> dict:
    try:
        status_code, raw_body = _post_json(
            api_url,
            headers=headers,
            payload=payload,
            timeout_seconds=REQUEST_TIMEOUT_SECONDS,
        )
    except URLError as exc:
        raise RuntimeError(f"请求本地服务失败: {exc}") from exc

    try:
        wrapper = json.loads(raw_body)
    except json.JSONDecodeError as exc:
        raise RuntimeError("本地服务返回非 JSON") from exc

    if not isinstance(wrapper, dict):
        raise RuntimeError("本地服务返回非 JSON 对象")
    if not wrapper.get("ok"):
        err = wrapper.get("error") or {}
        raise RuntimeError(f"本地服务错误: {err.get('code', 'UNKNOWN')} {err.get('message', '')}".strip())

    data = wrapper.get("data") or {}
    upstream_status = int(data.get("status", status_code) or 0)
    if upstream_status != 200:
        raise RuntimeError(f"上游状态码非 200: {upstream_status}")

    upstream_body_raw = data.get("body")
    if not isinstance(upstream_body_raw, str):
        raise RuntimeError("上游 body 缺失或类型错误")

    try:
        upstream = json.loads(upstream_body_raw)
    except json.JSONDecodeError as exc:
        raise RuntimeError("上游 body 非合法 JSON") from exc

    if expect_jsonrpc and not isinstance(upstream, dict):
        raise RuntimeError("上游 JSON-RPC 响应不是对象")

    return {
        "wrapper": wrapper,
        "upstream": upstream,
        "status": upstream_status,
    }


def build_headers(token: str) -> dict:
    return {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }


def build_request_body(*, keyword: str, index: int, args: argparse.Namespace) -> dict:
    return {
        "id": 32 + index,
        "jsonrpc": "2.0",
        "method": "keywords.GetInfo",
        "params": {
            "phrase": keyword,
            "device": args.device,
            "currency": args.currency,
            "database": args.database,
            "locati0n": args.location,
            "date": args.date,
        },
    }


def build_payload(*, keyword: str, index: int, gmitm: str, args: argparse.Namespace) -> dict:
    return {
        "__gmitm": gmitm,
        "requestBody": build_request_body(keyword=keyword, index=index, args=args),
        "timeoutMs": args.timeout_ms,
        "waitTimeoutMs": args.wait_timeout_ms,
    }


def build_sim_payload(*, keyword: str, args: argparse.Namespace) -> dict:
    return {
        "keyword": keyword,
        "country": SIM_COUNTRY,
        "latest": SIM_LATEST,
        "isWindow": SIM_IS_WINDOW,
        "webSource": SIM_WEB_SOURCE,
        "rowsPerPage": SIM_ROWS_PER_PAGE,
        "asc": SIM_ASC,
        "sort": SIM_SORT_FIELD,
        "page": SIM_PAGE,
        "type": SIM_TYPE,
        "timeoutMs": args.timeout_ms,
        "waitTimeoutMs": args.wait_timeout_ms,
    }


def match_sim_record_for_keyword(records: Sequence[dict], keyword_normalized: str) -> Optional[dict]:
    target = str(keyword_normalized or "").strip()
    if not target:
        return None

    for item in list(records)[:SIM_ROWS_PER_PAGE]:
        if not isinstance(item, dict):
            continue
        candidate = normalize_keyword_text(item.get("keyword") or "")
        if candidate == target:
            return item
    return None


def fetch_sim_keyword_data(*, api_url: str, headers: dict, keyword: str, keyword_normalized: str, args: argparse.Namespace) -> dict:
    payload = build_sim_payload(keyword=keyword, args=args)
    try:
        result = post_local_service(
            api_url=api_url,
            headers=headers,
            payload=payload,
            expect_jsonrpc=False,
        )
        upstream = result.get("upstream")
        if not isinstance(upstream, dict):
            raise RuntimeError("SIM 上游响应不是对象")

        records = upstream.get("records") or []
        if not isinstance(records, list):
            raise RuntimeError("SIM 上游 records 非数组")

        matched_record = match_sim_record_for_keyword(records, keyword_normalized)
        if matched_record is None:
            return {
                "ok": True,
                "status": "no_match",
                "keyword": keyword,
                "keywordNormalized": keyword_normalized,
                "requestPayload": payload,
                "wrapper": result.get("wrapper"),
                "upstream": upstream,
                "recordsCount": len(records),
                "simWindowVolume": None,
                "simCpc": None,
                "simKd": None,
                "matchedKeyword": None,
            }

        return {
            "ok": True,
            "status": "matched",
            "keyword": keyword,
            "keywordNormalized": keyword_normalized,
            "requestPayload": payload,
            "wrapper": result.get("wrapper"),
            "upstream": upstream,
            "recordsCount": len(records),
            "simWindowVolume": _coerce_optional_number(matched_record.get("windowVolume")),
            "simCpc": _coerce_optional_number(matched_record.get("cpc")),
            "simKd": _coerce_optional_number(matched_record.get("difficulty")),
            "matchedKeyword": str(matched_record.get("keyword") or "").strip() or None,
        }
    except (RuntimeError, URLError, ValueError, json.JSONDecodeError) as exc:
        return {
            "ok": False,
            "status": "request_error",
            "keyword": keyword,
            "keywordNormalized": keyword_normalized,
            "requestPayload": payload,
            "error": str(exc),
            "simWindowVolume": None,
            "simCpc": None,
            "simKd": None,
            "matchedKeyword": None,
        }


def fetch_keyword_with_retry(
    *,
    api_url: str,
    headers: dict,
    gmitm: str,
    keyword: str,
    keyword_normalized: str,
    index: int,
    args: argparse.Namespace,
) -> dict:
    attempts: List[dict] = []

    for attempt in range(MAX_RETRIES + 1):
        payload = build_payload(keyword=keyword, index=index, gmitm=gmitm, args=args)
        try:
            result = post_local_service(
                api_url=api_url,
                headers=headers,
                payload=payload,
                expect_jsonrpc=True,
            )
            upstream = result["upstream"]
            if not isinstance(upstream, dict):
                raise RuntimeError("上游 JSON-RPC 响应不是对象")

            upstream_error = upstream.get("error")
            if isinstance(upstream_error, dict):
                code = upstream_error.get("code")
                message = upstream_error.get("message")
                raise RuntimeError(f"上游 JSON-RPC 错误: code={code} message={message}")

            keyword_rows = ((upstream.get("result") or {}).get("keywords") or [])
            if not isinstance(keyword_rows, list):
                raise RuntimeError("上游 result.keywords 非数组")

            attempts.append(
                {
                    "attempt": attempt + 1,
                    "status": "ok",
                    "requestPayload": payload,
                    "wrapper": result["wrapper"],
                    "upstream": upstream,
                    "rowsCount": len(keyword_rows),
                }
            )
            return {
                "ok": True,
                "keyword": keyword,
                "keywordNormalized": keyword_normalized,
                "attempts": attempts,
                "keywordsRows": keyword_rows,
            }
        except (RuntimeError, URLError, ValueError, json.JSONDecodeError) as exc:
            attempts.append(
                {
                    "attempt": attempt + 1,
                    "status": "error",
                    "requestPayload": payload,
                    "error": str(exc),
                }
            )
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_BACKOFF_SECONDS[min(attempt, len(RETRY_BACKOFF_SECONDS) - 1)])
                continue
            return {
                "ok": False,
                "keyword": keyword,
                "keywordNormalized": keyword_normalized,
                "attempts": attempts,
                "error": str(exc),
            }

    return {
        "ok": False,
        "keyword": keyword,
        "keywordNormalized": keyword_normalized,
        "attempts": attempts,
        "error": "未知错误",
    }


def aggregate_keyword_metrics(keyword_rows: List[dict]) -> dict:
    volumes: List[float] = []
    cpc_values: List[float] = []
    kd_values: List[float] = []
    databases: set[str] = set()

    for item in keyword_rows:
        if not isinstance(item, dict):
            continue

        volume = safe_float(item.get("volume"))
        volumes.append(volume)

        cpc = item.get("cpc")
        if isinstance(cpc, (int, float)):
            cpc_values.append(float(cpc))

        difficulty = item.get("difficulty")
        if isinstance(difficulty, (int, float)):
            kd_values.append(float(difficulty))

        database = str(item.get("database") or "").strip().lower()
        if database:
            databases.add(database)

    global_volume = sum(volumes)
    global_cpc_avg = (sum(cpc_values) / len(cpc_values)) if cpc_values else None
    global_difficulty_avg = (sum(kd_values) / len(kd_values)) if kd_values else None

    return {
        "globalVolume": int(round(global_volume)) if abs(global_volume - round(global_volume)) < 1e-9 else round(global_volume, 4),
        "globalCpcAvg": round(global_cpc_avg, 6) if global_cpc_avg is not None else None,
        "globalDifficultyAvg": round(global_difficulty_avg, 6) if global_difficulty_avg is not None else None,
        "databaseCount": len(databases),
        "cpcSampleCount": len(cpc_values),
        "difficultySampleCount": len(kd_values),
        "rowsCount": len(keyword_rows),
    }


def summarize_success_rows(
    rows: List[dict],
    input_count: int,
    failure_count: int,
    *,
    sim_matched_count: int = 0,
    sim_no_match_count: int = 0,
    sim_failure_count: int = 0,
) -> dict:
    total_global_volume = sum(safe_float(row.get("globalVolume")) for row in rows)

    cpc_values = [safe_float(row.get("globalCpcAvg")) for row in rows if row.get("globalCpcAvg") is not None]
    kd_values = [safe_float(row.get("globalDifficultyAvg")) for row in rows if row.get("globalDifficultyAvg") is not None]

    return {
        "inputCount": input_count,
        "successCount": len(rows),
        "failureCount": failure_count,
        "simMatchedCount": sim_matched_count,
        "simNoMatchCount": sim_no_match_count,
        "simFailureCount": sim_failure_count,
        "totalGlobalVolume": int(round(total_global_volume)) if abs(total_global_volume - round(total_global_volume)) < 1e-9 else round(total_global_volume, 4),
        "avgGlobalCpc": round(sum(cpc_values) / len(cpc_values), 6) if cpc_values else None,
        "avgGlobalDifficulty": round(sum(kd_values) / len(kd_values), 6) if kd_values else None,
    }



def sort_rows(rows: List[dict]) -> List[dict]:
    return sorted(
        rows,
        key=lambda row: (
            -safe_float(row.get("globalVolume")),
            -(safe_float(row.get("globalCpcAvg")) if row.get("globalCpcAvg") is not None else -1),
            str(row.get("keyword") or ""),
        ),
    )


def compute_sim_score(row: dict) -> Optional[float]:
    sim_window_volume = row.get("simWindowVolume")
    sim_cpc = row.get("simCpc")
    sim_kd = row.get("simKd")
    if sim_window_volume is None or sim_cpc is None or sim_kd is None:
        return None

    window_volume = safe_float(sim_window_volume)
    cpc = safe_float(sim_cpc)
    kd = safe_float(sim_kd)
    if window_volume <= 0 or cpc <= 0 or kd <= 0:
        return None
    return round(window_volume * cpc / kd, 6)


def _has_sim_source(row: dict) -> bool:
    return any(row.get(field) is not None for field in ("simWindowVolume", "simKd", "simCpc"))


def _has_sem_source(row: dict) -> bool:
    return any(row.get(field) is not None for field in ("semVolume", "semKd", "semCpc"))


def _coerce_int_if_whole(value: Optional[float]) -> Optional[float | int]:
    if value is None:
        return None
    if abs(value - round(value)) < 1e-9:
        return int(round(value))
    return round(value, 6)


def _build_standard_word_rows(input_keywords: List[dict], success_rows: List[dict]) -> List[dict]:
    success_by_normalized = {
        str(row.get("keywordNormalized") or ""): row
        for row in success_rows
        if str(row.get("keywordNormalized") or "")
    }

    table_rows: List[dict] = []
    for item in input_keywords:
        keyword = str(item.get("keyword") or "").strip()
        normalized = str(item.get("keywordNormalized") or "")
        seed = dict(item.get("seedRow") or {})

        sem_metrics = success_by_normalized.get(normalized)
        if sem_metrics is not None:
            seed["semVolume"] = _coerce_int_if_whole(_coerce_optional_number(sem_metrics.get("globalVolume")))
            seed["semCpc"] = _coerce_optional_number(sem_metrics.get("globalCpcAvg"))
            seed["semKd"] = _coerce_optional_number(sem_metrics.get("globalDifficultyAvg"))
            sim_window_volume = _coerce_optional_number(sem_metrics.get("simWindowVolume"))
            sim_cpc = _coerce_optional_number(sem_metrics.get("simCpc"))
            sim_kd = _coerce_optional_number(sem_metrics.get("simKd"))
            if sim_window_volume is not None:
                seed["simWindowVolume"] = _coerce_int_if_whole(sim_window_volume)
            if sim_cpc is not None:
                seed["simCpc"] = sim_cpc
            if sim_kd is not None:
                seed["simKd"] = sim_kd

        row = {
            "keyword": keyword,
            "correspondingDomain": seed.get("correspondingDomain"),
            "group": seed.get("group") or keyword,
            "sourcePresence": None,
            "score": None,
            "simWindowVolume": _coerce_optional_number(seed.get("simWindowVolume")),
            "simKd": _coerce_optional_number(seed.get("simKd")),
            "simCpc": _coerce_optional_number(seed.get("simCpc")),
            "semVolume": _coerce_optional_number(seed.get("semVolume")),
            "semKd": _coerce_optional_number(seed.get("semKd")),
            "semCpc": _coerce_optional_number(seed.get("semCpc")),
            "gefeiKD": _coerce_optional_number(seed.get("gefeiKD")),
        }

        has_sim = _has_sim_source(row)
        has_sem = _has_sem_source(row)
        if has_sim and has_sem:
            row["sourcePresence"] = "both"
        elif has_sim:
            row["sourcePresence"] = "sim_only"
        elif has_sem:
            row["sourcePresence"] = "sem_only"

        row["score"] = compute_sim_score(row)
        table_rows.append(row)

    def table_sort_key(row: dict):
        score = row.get("score")
        return (
            score is None,
            -(safe_float(score) if score is not None else -1),
            -safe_float(row.get("semVolume")),
            -safe_float(row.get("simWindowVolume")),
            str(row.get("keyword") or ""),
        )

    return sorted(table_rows, key=table_sort_key)


def _fallback_standard_word_rows_from_snapshot(snapshot: dict) -> List[dict]:
    rows = snapshot.get("rows") or []
    result: List[dict] = []
    for row in rows:
        keyword = str(row.get("keyword") or "").strip()
        if not keyword:
            continue
        result.append(
            {
                "keyword": keyword,
                "correspondingDomain": None,
                "group": keyword,
                "sourcePresence": "sem_only",
                "score": None,
                "simWindowVolume": None,
                "simKd": None,
                "simCpc": None,
                "semVolume": _coerce_optional_number(row.get("globalVolume")),
                "semKd": _coerce_optional_number(row.get("globalDifficultyAvg")),
                "semCpc": _coerce_optional_number(row.get("globalCpcAvg")),
                "gefeiKD": None,
            }
        )
    return result


def _extract_standard_word_rows(snapshot: dict) -> List[dict]:
    rows = snapshot.get("standardWordRows")
    if isinstance(rows, list) and rows:
        return rows
    return _fallback_standard_word_rows_from_snapshot(snapshot)


def _fmt_md_cell(value) -> str:
    if value is None:
        text = "-"
    else:
        text = str(value).strip()
    if not text:
        text = "-"

    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    normalized = re.sub(r"\s*\|\s*", "\n", normalized)
    normalized = normalized.replace(" / ", "\n")
    lines = [line.strip() for line in normalized.split("\n") if line.strip()]
    if not lines:
        lines = ["-"]

    escaped: List[str] = []
    for line in lines:
        line = re.sub(r"([/?&#=_-])", r"\1<wbr>", line)
        escaped.append(line.replace("|", "\\|"))
    return "<br>".join(escaped)


def _md_table(headers: Sequence[str], rows: Sequence[Sequence[str]]) -> List[str]:
    if not rows:
        return ["（无）"]
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(_fmt_md_cell(cell) for cell in row) + " |")
    return lines


def build_snapshot(
    *,
    stamp: str,
    args: argparse.Namespace,
    keywords: List[dict],
    rows: List[dict],
    standard_word_rows: List[dict],
    failures: List[dict],
    summary: dict,
    input_overview: dict,
    report_history_path: Path,
    excel_history_path: Path,
    table_json_history_path: Path,
) -> dict:
    scored_count = sum(1 for row in standard_word_rows if row.get("score") is not None)
    return {
        "meta": {
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "stamp": stamp,
            "target": {
                "keywords": [item["keyword"] for item in keywords],
                "keywordCount": len(keywords),
                "inputSources": input_overview.get("inputSources") or [],
            },
            "request": {
                "apiBase": args.api_base,
                "endpoint": args.endpoint,
                "simEndpoint": args.sim_endpoint,
                "simRowsPerPage": SIM_ROWS_PER_PAGE,
                "simType": SIM_TYPE,
                "simSort": SIM_SORT_FIELD,
                "simAsc": SIM_ASC,
                "simCountry": SIM_COUNTRY,
                "simLatest": SIM_LATEST,
                "simWebSource": SIM_WEB_SOURCE,
                "simIsWindow": SIM_IS_WINDOW,
                "simPage": SIM_PAGE,
                "device": args.device,
                "currency": args.currency,
                "database": args.database,
                "locati0n": args.location,
                "date": args.date,
                "timeoutMs": args.timeout_ms,
                "waitTimeoutMs": args.wait_timeout_ms,
                "maxRetries": MAX_RETRIES,
                "standardWordTableVersion": STANDARD_WORD_TABLE_VERSION,
                "scoreFormula": SCORE_FORMULA,
            },
            "inputOverview": input_overview,
            "summary": {
                **summary,
                "standardWordRowCount": len(standard_word_rows),
                "scoredCount": scored_count,
            },
            "output": {
                "reportHistoryPath": str(report_history_path),
                "excelHistoryPath": str(excel_history_path),
                "tableJsonHistoryPath": str(table_json_history_path),
            },
        },
        "rows": rows,
        "standardWordRows": standard_word_rows,
        "failures": failures,
    }


def render_report(snapshot: dict) -> str:
    meta = snapshot.get("meta") or {}
    request = meta.get("request") or {}
    summary = meta.get("summary") or {}
    output = meta.get("output") or {}

    rows = snapshot.get("rows") or []
    standard_word_rows = _extract_standard_word_rows(snapshot)
    failures = snapshot.get("failures") or []

    preview_rows: List[List[str]] = []
    for row in rows[:TOP_PREVIEW_ROWS]:
        preview_rows.append(
            [
                str(row.get("keyword") or "-"),
                to_display_number(row.get("globalVolume"), digits=0),
                to_display_number(row.get("globalCpcAvg")),
                to_display_number(row.get("globalDifficultyAvg")),
                to_display_number(row.get("databaseCount"), digits=0),
            ]
        )

    table_preview_rows: List[List[str]] = []
    for row in standard_word_rows[:TOP_PREVIEW_ROWS]:
        table_preview_rows.append(
            [
                str(row.get("keyword") or "-"),
                str(row.get("sourcePresence") or "-"),
                to_display_number(row.get("score"), digits=6),
                to_display_number(row.get("simWindowVolume"), digits=0),
                to_display_number(row.get("semVolume"), digits=0),
            ]
        )

    failure_rows: List[List[str]] = []
    for row in failures:
        failure_rows.append(
            [
                str(row.get("keyword") or "-"),
                str(row.get("attempts") or "-"),
                str(row.get("error") or "-"),
            ]
        )

    lines: List[str] = []
    lines.append(f"# 关键词批量分析报告（{meta.get('stamp', '-')}）")
    lines.append("")

    lines.append("## 摘要")
    lines.append("")
    lines.append(f"- 输入关键词数：{summary.get('inputCount', 0)}")
    lines.append(f"- 成功关键词数：{summary.get('successCount', 0)}")
    lines.append(f"- 失败关键词数：{summary.get('failureCount', 0)}")
    lines.append(f"- SIM 匹配成功数：{summary.get('simMatchedCount', 0)}")
    lines.append(f"- SIM 未命中数：{summary.get('simNoMatchCount', 0)}")
    lines.append(f"- SIM 请求失败数：{summary.get('simFailureCount', 0)}")
    lines.append(f"- 标准词表行数：{summary.get('standardWordRowCount', len(standard_word_rows))}")
    lines.append(f"- 有效 score 行数：{summary.get('scoredCount', 0)}")
    lines.append("")

    lines.append("## 抓取概览")
    lines.append("")
    lines.append(f"- API：{request.get('apiBase', '-')}{request.get('endpoint', '-')}")
    lines.append(f"- SIM API：{request.get('apiBase', '-')}{request.get('simEndpoint', '-')}")
    lines.append(
        f"- SIM 固定参数：rowsPerPage={request.get('simRowsPerPage', '-')}, type={request.get('simType', '-')}, "
        f"sort={request.get('simSort', '-')}, asc={request.get('simAsc', '-')}"
    )
    lines.append(
        f"- SIM 默认口径：country={request.get('simCountry', '-')}, latest={request.get('simLatest', '-')}, "
        f"webSource={request.get('simWebSource', '-')}, isWindow={request.get('simIsWindow', '-')}, page={request.get('simPage', '-')}"
    )
    lines.append(f"- 默认参数：device={request.get('device', '-')} currency={request.get('currency', '-')} database={request.get('database', '-')} locati0n={request.get('locati0n', '-')} date={request.get('date', '')!r}")
    lines.append(f"- timeoutMs={request.get('timeoutMs', '-')} / waitTimeoutMs={request.get('waitTimeoutMs', '-')} / maxRetries={request.get('maxRetries', '-')}")
    lines.append("")

    lines.append("## 聚合结果概览")
    lines.append("")
    lines.append(f"- totalGlobalVolume（所有关键词汇总）：{to_display_number(summary.get('totalGlobalVolume'), digits=0)}")
    lines.append(f"- avgGlobalCpc（关键词级均值）：{to_display_number(summary.get('avgGlobalCpc'))}")
    lines.append(f"- avgGlobalDifficulty（关键词级均值）：{to_display_number(summary.get('avgGlobalDifficulty'))}")
    lines.append("")

    lines.append("## 关键词结果预览")
    lines.append("")
    lines.extend(
        _md_table(
            ["keyword", "globalVolume", "globalCpcAvg", "globalDifficultyAvg", "databaseCount"],
            preview_rows,
        )
    )
    lines.append("")

    lines.append("## 标准词表预览")
    lines.append("")
    lines.extend(
        _md_table(
            ["keyword", "sourcePresence", "score(sim)", "volume(sim)", "volume(sem)"],
            table_preview_rows,
        )
    )
    lines.append("")

    lines.append("## 失败关键词")
    lines.append("")
    lines.extend(_md_table(["keyword", "attempts", "error"], failure_rows))
    lines.append("")

    lines.append("## 产物路径")
    lines.append("")
    lines.append(f"- Markdown：{output.get('reportHistoryPath', '-')}")
    lines.append(f"- Excel：{output.get('excelHistoryPath', '-')}")
    lines.append(f"- JSON：{output.get('tableJsonHistoryPath', '-')}")
    lines.append("")

    lines.append("## 备注")
    lines.append("")
    for remark in REQUIRED_REMARKS:
        lines.append(f"- {remark}")
    lines.append("")

    return "\n".join(lines)


def _build_standard_table_json_payload(*, snapshot: dict) -> dict:
    meta = snapshot.get("meta") or {}
    request = meta.get("request") or {}
    standard_word_rows = _extract_standard_word_rows(snapshot)
    return {
        "version": STANDARD_WORD_TABLE_VERSION,
        "meta": {
            "generatedAt": meta.get("generatedAt") or datetime.now().isoformat(timespec="seconds"),
            "standardWordTableVersion": request.get("standardWordTableVersion") or STANDARD_WORD_TABLE_VERSION,
            "scoreFormula": request.get("scoreFormula") or SCORE_FORMULA,
            "stamp": meta.get("stamp"),
        },
        "rows": standard_word_rows,
    }


def _transform_export_value(column_def: dict, value, row: Optional[dict] = None):
    transform = column_def.get("transform")
    if transform == "source_presence_label":
        transformed = SOURCE_PRESENCE_ENUM_TO_LABEL.get(str(value or ""), value)
    elif transform == "sim_score":
        transformed = compute_sim_score(row or {})
    else:
        transformed = value
    field = str(column_def.get("field") or "")
    return _format_multiline_export_text(field, transformed)


def _compute_column_width(values: Sequence[str], width_profile: Optional[dict] = None) -> float:
    profile = width_profile or {}
    fixed = profile.get("fixed")
    if fixed is not None:
        return float(fixed)

    max_length = 0
    for raw in values:
        text = "" if raw is None else str(raw)
        parts = text.splitlines() or [text]
        longest_line = max((len(part) for part in parts), default=0)
        if longest_line > max_length:
            max_length = longest_line

    width = max_length + DEFAULT_COLUMN_PADDING
    min_width = int(profile.get("min", DEFAULT_COLUMN_MIN_WIDTH))
    max_width = int(profile.get("max", DEFAULT_COLUMN_MAX_WIDTH))
    return float(max(min_width, min(width, max_width)))


def write_excel(snapshot: dict, output_path: Path) -> None:
    Workbook, Font, Alignment, get_column_letter, _ = _require_openpyxl()

    meta = snapshot.get("meta") or {}
    summary = meta.get("summary") or {}
    request = meta.get("request") or {}
    standard_word_rows = _extract_standard_word_rows(snapshot)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    keywords_sheet = workbook.create_sheet("keywords")

    summary_rows = [
        ("generatedAt", meta.get("generatedAt", "")),
        ("stamp", meta.get("stamp", "")),
        ("inputCount", summary.get("inputCount", 0)),
        ("successCount", summary.get("successCount", 0)),
        ("failureCount", summary.get("failureCount", 0)),
        ("simMatchedCount", summary.get("simMatchedCount", 0)),
        ("simNoMatchCount", summary.get("simNoMatchCount", 0)),
        ("simFailureCount", summary.get("simFailureCount", 0)),
        ("standardWordRows", summary.get("standardWordRowCount", len(standard_word_rows))),
        ("scoredCount", summary.get("scoredCount", 0)),
        ("scoreFormula", request.get("scoreFormula", SCORE_FORMULA)),
        ("standardWordTableVersion", request.get("standardWordTableVersion", STANDARD_WORD_TABLE_VERSION)),
    ]
    for idx, (name, value) in enumerate(summary_rows, start=1):
        summary_sheet.cell(row=idx, column=1, value=name)
        summary_sheet.cell(row=idx, column=2, value=value)
    summary_sheet.freeze_panes = "A2"

    export_columns = get_keywords_export_columns()
    headers = get_keywords_export_headers()
    keywords_sheet.append(headers)
    for cell in keywords_sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in standard_word_rows:
        export_row = []
        for column in export_columns:
            raw_value = row.get(column["field"])
            export_row.append(_transform_export_value(column, raw_value, row))
        keywords_sheet.append(export_row)

    for row_cells in keywords_sheet.iter_rows(min_row=2, max_row=keywords_sheet.max_row, min_col=1, max_col=len(headers)):
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    keywords_sheet.freeze_panes = "A2"
    keywords_sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(keywords_sheet.max_row, 1)}"

    for index, column in enumerate(export_columns, start=1):
        number_format = column.get("number_format")
        if not number_format:
            continue
        col_letter = get_column_letter(index)
        for cell in keywords_sheet[col_letter][1:]:
            cell.number_format = number_format

    for column_index, column in enumerate(export_columns, start=1):
        column_letter = get_column_letter(column_index)
        values = []
        for row_index in range(1, keywords_sheet.max_row + 1):
            value = keywords_sheet.cell(row=row_index, column=column_index).value
            values.append("" if value is None else str(value))
        keywords_sheet.column_dimensions[column_letter].width = _compute_column_width(values, column.get("width_profile"))

    for column_cells in summary_sheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        summary_sheet.column_dimensions[column_cells[0].column_letter].width = _compute_column_width(
            values,
            {"min": 14, "max": 48},
        )

    ensure_dir(output_path.parent)
    workbook.save(output_path)


def write_artifacts(
    *,
    stamp: str,
    snapshot: dict,
    fetch_archive: dict,
    data_dir: Path,
    snapshot_dir: Path,
    report_dir: Path,
    latest_report_path: Path,
    latest_xlsx_path: Path,
    latest_table_json_path: Path,
) -> dict:
    ensure_dir(data_dir)
    ensure_dir(snapshot_dir)
    ensure_dir(report_dir)
    ensure_dir(latest_report_path.parent)
    ensure_dir(latest_xlsx_path.parent)
    ensure_dir(latest_table_json_path.parent)

    fetch_archive_path = data_dir / f"fetch-{stamp}.json"
    snapshot_path = snapshot_dir / f"snapshot-{stamp}.json"
    report_history_path = report_dir / f"report-{stamp}.md"
    excel_history_path = report_dir / f"keyword-table-{stamp}.xlsx"
    table_json_history_path = report_dir / f"keyword-table-{stamp}.json"

    snapshot_meta = snapshot.setdefault("meta", {})
    output_meta = snapshot_meta.setdefault("output", {})
    output_meta["reportHistoryPath"] = str(report_history_path)
    output_meta["excelHistoryPath"] = str(excel_history_path)
    output_meta["tableJsonHistoryPath"] = str(table_json_history_path)
    snapshot_meta["latestReportPath"] = str(latest_report_path)
    snapshot_meta["latestExcelPath"] = str(latest_xlsx_path)
    snapshot_meta["latestTableJsonPath"] = str(latest_table_json_path)

    report_text = render_report(snapshot)
    table_json_payload = _build_standard_table_json_payload(snapshot=snapshot)

    dump_json(fetch_archive_path, fetch_archive)
    dump_json(snapshot_path, snapshot)
    report_history_path.write_text(report_text, encoding="utf-8")
    latest_report_path.write_text(report_text, encoding="utf-8")

    write_excel(snapshot, excel_history_path)
    copyfile(excel_history_path, latest_xlsx_path)

    dump_json(table_json_history_path, table_json_payload)
    copyfile(table_json_history_path, latest_table_json_path)

    return {
        "fetchArchivePath": fetch_archive_path,
        "snapshotPath": snapshot_path,
        "reportHistoryPath": report_history_path,
        "excelHistoryPath": excel_history_path,
        "tableJsonHistoryPath": table_json_history_path,
    }


def rebuild_history_artifacts(
    snapshot: dict,
    report_dir: Path,
    latest_report_path: Path,
    latest_xlsx_path: Path,
    latest_table_json_path: Path,
) -> Tuple[Path, Path, Path, str]:
    ensure_dir(report_dir)

    meta = snapshot.get("meta") or {}
    stamp = str(meta.get("stamp") or "")
    if not stamp:
        raise RuntimeError("snapshot 缺少 meta.stamp")

    report_history_path = report_dir / f"report-{stamp}.md"
    excel_history_path = report_dir / f"keyword-table-{stamp}.xlsx"
    table_json_history_path = report_dir / f"keyword-table-{stamp}.json"

    snapshot_meta = snapshot.setdefault("meta", {})
    output_meta = snapshot_meta.setdefault("output", {})
    output_meta["reportHistoryPath"] = str(report_history_path)
    output_meta["excelHistoryPath"] = str(excel_history_path)
    output_meta["tableJsonHistoryPath"] = str(table_json_history_path)
    snapshot_meta["latestReportPath"] = str(latest_report_path)
    snapshot_meta["latestExcelPath"] = str(latest_xlsx_path)
    snapshot_meta["latestTableJsonPath"] = str(latest_table_json_path)

    request = snapshot_meta.setdefault("request", {})
    request["standardWordTableVersion"] = STANDARD_WORD_TABLE_VERSION
    request["scoreFormula"] = SCORE_FORMULA

    if not snapshot.get("standardWordRows"):
        snapshot["standardWordRows"] = _fallback_standard_word_rows_from_snapshot(snapshot)

    report_text = render_report(snapshot)
    table_json_payload = _build_standard_table_json_payload(snapshot=snapshot)

    report_history_path.write_text(report_text, encoding="utf-8")
    write_excel(snapshot, excel_history_path)
    dump_json(table_json_history_path, table_json_payload)

    return report_history_path, excel_history_path, table_json_history_path, report_text


def run_pipeline(args: argparse.Namespace) -> int:
    data_dir = Path(args.data_dir).resolve()
    snapshot_dir = Path(args.snapshot_dir).resolve()
    report_dir = Path(args.report_dir).resolve()
    latest_report_path = Path(args.latest_report_path).resolve()
    latest_xlsx_path = Path(args.latest_xlsx_path).resolve()
    latest_table_json_path = Path(args.latest_table_json_path).resolve()
    token_path = Path(args.token_path).resolve()
    gmitm_path = Path(args.gmitm_path).resolve()

    token = read_text_required(token_path, "token 文件")
    gmitm = read_gmitm(gmitm_path)
    headers = build_headers(token)

    keywords, input_overview = collect_keywords(args)
    sem_api_url = args.api_base.rstrip("/") + args.endpoint
    sim_api_url = args.api_base.rstrip("/") + args.sim_endpoint

    success_rows: List[dict] = []
    failures: List[dict] = []
    raw_results: List[dict] = []

    sim_matched_count = 0
    sim_no_match_count = 0
    sim_failure_count = 0

    for index, item in enumerate(keywords, start=1):
        keyword = item["keyword"]
        keyword_normalized = item["keywordNormalized"]

        sem_fetch_result = fetch_keyword_with_retry(
            api_url=sem_api_url,
            headers=headers,
            gmitm=gmitm,
            keyword=keyword,
            keyword_normalized=keyword_normalized,
            index=index,
            args=args,
        )

        if not sem_fetch_result.get("ok"):
            attempts = sem_fetch_result.get("attempts") or []
            failures.append(
                {
                    "keyword": keyword,
                    "keywordNormalized": keyword_normalized,
                    "attempts": len(attempts),
                    "error": str(sem_fetch_result.get("error") or "未知错误"),
                }
            )
            raw_results.append(
                {
                    "keyword": keyword,
                    "keywordNormalized": keyword_normalized,
                    "sem": sem_fetch_result,
                    "sim": {
                        "ok": True,
                        "status": "skipped_due_to_sem_failure",
                        "keyword": keyword,
                        "keywordNormalized": keyword_normalized,
                        "simWindowVolume": None,
                        "simCpc": None,
                        "simKd": None,
                    },
                }
            )
            continue

        sim_fetch_result = fetch_sim_keyword_data(
            api_url=sim_api_url,
            headers=headers,
            keyword=keyword,
            keyword_normalized=keyword_normalized,
            args=args,
        )

        sim_status = str(sim_fetch_result.get("status") or "")
        if sim_status == "matched":
            sim_matched_count += 1
        elif sim_status == "no_match":
            sim_no_match_count += 1
        else:
            sim_failure_count += 1

        raw_results.append(
            {
                "keyword": keyword,
                "keywordNormalized": keyword_normalized,
                "sem": sem_fetch_result,
                "sim": sim_fetch_result,
            }
        )

        keyword_rows = sem_fetch_result.get("keywordsRows") or []
        metrics = aggregate_keyword_metrics(keyword_rows)
        success_rows.append(
            {
                "keyword": keyword,
                "keywordNormalized": keyword_normalized,
                **metrics,
                "simWindowVolume": _coerce_optional_number(sim_fetch_result.get("simWindowVolume")),
                "simCpc": _coerce_optional_number(sim_fetch_result.get("simCpc")),
                "simKd": _coerce_optional_number(sim_fetch_result.get("simKd")),
                "simStatus": sim_status or "request_error",
            }
        )

    sorted_rows = sort_rows(success_rows)
    summary = summarize_success_rows(
        sorted_rows,
        input_count=len(keywords),
        failure_count=len(failures),
        sim_matched_count=sim_matched_count,
        sim_no_match_count=sim_no_match_count,
        sim_failure_count=sim_failure_count,
    )

    if failures:
        print("以下关键词抓取失败：")
        for item in failures:
            print(f"- {item['keyword']} (attempts={item['attempts']}): {item['error']}")
        raise SystemExit("存在失败关键词，按默认策略终止且不落任何新产物。")

    standard_word_rows = _build_standard_word_rows(keywords, sorted_rows)

    stamp = now_stamp()
    placeholder_report_history_path = report_dir / f"report-{stamp}.md"
    placeholder_excel_history_path = report_dir / f"keyword-table-{stamp}.xlsx"
    placeholder_table_json_history_path = report_dir / f"keyword-table-{stamp}.json"

    snapshot = build_snapshot(
        stamp=stamp,
        args=args,
        keywords=keywords,
        rows=sorted_rows,
        standard_word_rows=standard_word_rows,
        failures=failures,
        summary=summary,
        input_overview=input_overview,
        report_history_path=placeholder_report_history_path,
        excel_history_path=placeholder_excel_history_path,
        table_json_history_path=placeholder_table_json_history_path,
    )

    fetch_archive = {
        "meta": {
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "stamp": stamp,
            "summary": summary,
            "request": {
                "semApiUrl": sem_api_url,
                "simApiUrl": sim_api_url,
                "simRowsPerPage": SIM_ROWS_PER_PAGE,
                "simType": SIM_TYPE,
                "simSort": SIM_SORT_FIELD,
                "simAsc": SIM_ASC,
                "simCountry": SIM_COUNTRY,
                "simLatest": SIM_LATEST,
                "simWebSource": SIM_WEB_SOURCE,
                "simIsWindow": SIM_IS_WINDOW,
                "simPage": SIM_PAGE,
                "maxRetries": MAX_RETRIES,
            },
            "inputOverview": input_overview,
        },
        "requests": raw_results,
    }

    artifact_paths = write_artifacts(
        stamp=stamp,
        snapshot=snapshot,
        fetch_archive=fetch_archive,
        data_dir=data_dir,
        snapshot_dir=snapshot_dir,
        report_dir=report_dir,
        latest_report_path=latest_report_path,
        latest_xlsx_path=latest_xlsx_path,
        latest_table_json_path=latest_table_json_path,
    )

    print(f"[done] fetch archive : {artifact_paths['fetchArchivePath']}")
    print(f"[done] snapshot      : {artifact_paths['snapshotPath']}")
    print(f"[done] report history: {artifact_paths['reportHistoryPath']}")
    print(f"[done] report latest : {latest_report_path}")
    print(f"[done] excel history : {artifact_paths['excelHistoryPath']}")
    print(f"[done] excel latest  : {latest_xlsx_path}")
    print(f"[done] table history : {artifact_paths['tableJsonHistoryPath']}")
    print(f"[done] table latest  : {latest_table_json_path}")
    print(f"[done] input keywords: {summary['inputCount']}")
    print(f"[done] success count : {summary['successCount']}")
    print(f"[done] failure count : {summary['failureCount']}")
    return 0


def rebuild_reports(args: argparse.Namespace) -> int:
    snapshot_dir = Path(args.snapshot_dir).resolve()
    report_dir = Path(args.report_dir).resolve()
    latest_report_path = Path(args.latest_report_path).resolve()
    latest_xlsx_path = Path(args.latest_xlsx_path).resolve()
    latest_table_json_path = Path(args.latest_table_json_path).resolve()

    ensure_dir(snapshot_dir)
    ensure_dir(report_dir)
    ensure_dir(latest_report_path.parent)
    ensure_dir(latest_xlsx_path.parent)
    ensure_dir(latest_table_json_path.parent)

    files = list_snapshot_files(snapshot_dir)
    if not files:
        raise SystemExit(f"未找到快照文件: {snapshot_dir}")

    latest_report_text = ""
    latest_excel_path: Optional[Path] = None
    latest_table_json_history_path: Optional[Path] = None

    for path in files:
        snapshot = load_json(path)
        report_history_path, excel_history_path, table_json_history_path, report_text = rebuild_history_artifacts(
            snapshot,
            report_dir,
            latest_report_path,
            latest_xlsx_path,
            latest_table_json_path,
        )
        latest_report_text = report_text
        latest_excel_path = excel_history_path
        latest_table_json_history_path = table_json_history_path
        print(f"[rebuild] report: {report_history_path}")
        print(f"[rebuild] excel : {excel_history_path}")
        print(f"[rebuild] json  : {table_json_history_path}")

    latest_report_path.write_text(latest_report_text, encoding="utf-8")
    if latest_excel_path is not None:
        copyfile(latest_excel_path, latest_xlsx_path)
    if latest_table_json_history_path is not None:
        copyfile(latest_table_json_history_path, latest_table_json_path)

    print(f"[done] latest report: {latest_report_path}")
    print(f"[done] latest excel : {latest_xlsx_path}")
    print(f"[done] latest json  : {latest_table_json_path}")
    return 0


def validate_report(args: argparse.Namespace) -> int:
    report_path = Path(args.report).resolve()
    xlsx_path = Path(args.xlsx).resolve()
    table_json_path = Path(args.table_json).resolve()

    if not report_path.exists():
        raise SystemExit(f"报告不存在: {report_path}")

    text = report_path.read_text(encoding="utf-8")
    missing_sections = [section for section in REQUIRED_SECTIONS if section not in text]
    missing_remarks = [remark for remark in REQUIRED_REMARKS if remark not in text]

    if missing_sections or missing_remarks:
        if missing_sections:
            print("缺少 section:")
            for item in missing_sections:
                print(f"- {item}")
        if missing_remarks:
            print("缺少 remark:")
            for item in missing_remarks:
                print(f"- {item}")
        raise SystemExit(1)

    if not xlsx_path.exists():
        raise SystemExit(f"Excel 不存在: {xlsx_path}")

    _Workbook, _Font, _Alignment, _get_column_letter, load_workbook = _require_openpyxl()
    workbook = load_workbook(xlsx_path, data_only=True)
    try:
        if "keywords" not in workbook.sheetnames:
            raise SystemExit("Excel 缺少 keywords sheet")
        sheet = workbook["keywords"]

        expected_headers = get_keywords_export_headers()
        actual_headers = [sheet.cell(row=1, column=index).value for index in range(1, sheet.max_column + 1)]
        if actual_headers != expected_headers:
            print("Excel 表头不匹配：")
            max_len = max(len(expected_headers), len(actual_headers))
            for index in range(1, max_len + 1):
                expected = expected_headers[index - 1] if index <= len(expected_headers) else None
                actual = actual_headers[index - 1] if index <= len(actual_headers) else None
                if expected != actual:
                    print(f"- 第 {index} 列 expected={expected!r} actual={actual!r}")
            raise SystemExit(1)

        keyword_col = expected_headers.index(get_header_for_field("keyword")) + 1
        score_col = get_score_column_index_from_headers(expected_headers) + 1
        numeric_fields = [
            "simWindowVolume",
            "simKd",
            "simCpc",
            "semVolume",
            "semKd",
            "semCpc",
            "gefeiKD",
        ]
        numeric_indexes = [expected_headers.index(get_header_for_field(field)) + 1 for field in numeric_fields]

        for row_index in range(2, sheet.max_row + 1):
            keyword = sheet.cell(row=row_index, column=keyword_col).value
            if keyword in (None, ""):
                raise SystemExit(f"keyword 列为空（row={row_index}）")

            score_value = sheet.cell(row=row_index, column=score_col).value
            if score_value not in (None, ""):
                try:
                    float(score_value)
                except (TypeError, ValueError) as exc:
                    raise SystemExit(f"score 列不是数字（row={row_index}）") from exc

            for col_index in numeric_indexes:
                value = sheet.cell(row=row_index, column=col_index).value
                if value in (None, ""):
                    continue
                try:
                    float(value)
                except (TypeError, ValueError) as exc:
                    raise SystemExit(f"数字列不是数字（row={row_index}, col={col_index}）") from exc
    finally:
        workbook.close()

    if not table_json_path.exists():
        raise SystemExit(f"标准词表 JSON 不存在: {table_json_path}")

    table_payload = load_json(table_json_path)
    version = str(table_payload.get("version") or "").strip()
    if version != STANDARD_WORD_TABLE_VERSION:
        raise SystemExit(
            f"标准词表 JSON 版本不匹配: expected={STANDARD_WORD_TABLE_VERSION} actual={version or '-'}"
        )

    rows = table_payload.get("rows")
    if not isinstance(rows, list):
        raise SystemExit("标准词表 JSON rows 非数组")
    for index, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            raise SystemExit(f"标准词表 JSON row 非对象（index={index}）")
        if str(row.get("keyword") or "").strip() == "":
            raise SystemExit(f"标准词表 JSON keyword 为空（index={index}）")

    print(f"[ok] report validated: {report_path}")
    print(f"[ok] xlsx validated  : {xlsx_path}")
    print(f"[ok] json validated  : {table_json_path}")
    return 0


def main() -> int:
    args = parse_args()
    if args.command == "run":
        return run_pipeline(args)
    if args.command == "rebuild-reports":
        return rebuild_reports(args)
    if args.command == "validate-report":
        return validate_report(args)
    raise SystemExit(f"未知命令: {args.command}")


STANDARD_WORD_TABLE_SPEC = load_standard_word_table_spec()


if __name__ == "__main__":
    raise SystemExit(main())
