#!/usr/bin/env python3
"""监控子域名落地页样本并输出对比报告（MVP）。

Usage:
  python3 word-monitor-sub-domain/_internal/scripts/word_monitor_subdomain.py run
  python3 word-monitor-sub-domain/_internal/scripts/word_monitor_subdomain.py rebuild-reports
  python3 word-monitor-sub-domain/_internal/scripts/word_monitor_subdomain.py validate-report
"""

from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path
from shutil import copyfile
from typing import Dict, List, Optional, Sequence, Tuple
from urllib.error import HTTPError, URLError
from urllib.parse import urlsplit
from urllib.request import Request, urlopen

PROJECT_DIR = Path(__file__).resolve().parents[2]
REPO_ROOT = PROJECT_DIR.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from shared_gefei_kd import fetch_gefei_kd_rows

LOCAL_SERVICE_TOKEN_PATH = PROJECT_DIR / "local-service" / "bridge_token.txt"

STANDARD_WORD_TABLE_VERSION = "v1"
STANDARD_WORD_TABLE_SPEC_PATH = REPO_ROOT / "standard-word-analysis" / "spec" / f"standard-word-table.{STANDARD_WORD_TABLE_VERSION}.json"
ANALYZE_WORDS_SCRIPT_PATH = REPO_ROOT / "analyze-words" / "_internal" / "scripts" / "analyze_words.py"
CHECK_GEFEI_KD_SCRIPT_PATH = REPO_ROOT / "check-gefei-kd" / "_internal" / "scripts" / "check_gefei_kd.py"

DEFAULT_API_BASE = "http://127.0.0.1:17311"
DEFAULT_ENDPOINT = "/sim/api/websiteOrganicLandingPagesV2"

MAX_RETRIES = 2
RETRY_BACKOFF_SECONDS = [0.8, 1.6]
REQUEST_TIMEOUT_SECONDS = 180

PAGE_NEW_MIN_CLICKS = 100.0
PAGE_RISING_MIN_CLICKS = 100.0
PAGE_RISING_MIN_DELTA = 30.0
PAGE_RISING_MIN_GROWTH = 0.05

SUBDOMAIN_NEW_MIN_CLICKS = 150.0
SUBDOMAIN_RISING_MIN_CLICKS = 150.0
SUBDOMAIN_RISING_MIN_DELTA = 50.0
SUBDOMAIN_RISING_MIN_GROWTH = 0.05

SNAPSHOT_RE = re.compile(r"^snapshot-(\d{8}-\d{6})\.json$")
SPACE_RE = re.compile(r"\s+")

DEFAULT_COLUMN_PADDING = 4
DEFAULT_COLUMN_MIN_WIDTH = 12
DEFAULT_COLUMN_MAX_WIDTH = 72

SIM_COLUMN_FILL_COLOR = "FFEAF4FF"
SEM_COLUMN_FILL_COLOR = "FFF4EAFF"

REQUIRED_REMARKS = [
    "当前监控仅覆盖 ClicksShare 排序下前8页样本",
    "“新进入样本”不等于全站首次出现",
    "子域名流量是样本内观测值，不代表全站完整总量",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="监控子域名落地页样本并生成日报")
    subparsers = parser.add_subparsers(dest="command", required=True)

    run_parser = subparsers.add_parser("run", help="抓取 + 快照 + 对比 + 报告")
    run_parser.add_argument("--api-base", default=DEFAULT_API_BASE)
    run_parser.add_argument("--endpoint", default=DEFAULT_ENDPOINT)
    run_parser.add_argument("--token-path", default=str(LOCAL_SERVICE_TOKEN_PATH))

    run_parser.add_argument("--key", default="vercel.app")
    run_parser.add_argument("--country", default="999")
    run_parser.add_argument("--latest", default="28d")
    run_parser.add_argument("--web-source", default="Total")
    run_parser.add_argument("--source-type", default="organic")
    run_parser.add_argument("--sort", default="ClicksShare")
    run_parser.add_argument("--asc", action="store_true", default=False)
    run_parser.add_argument("--include-subdomains", action="store_true", default=True)
    run_parser.add_argument("--is-window", action="store_true", default=True)
    run_parser.add_argument("--search-type", default="domain")
    run_parser.add_argument("--start-page", type=int, default=1)
    run_parser.add_argument("--end-page", type=int, default=8)

    run_parser.add_argument("--data-dir", default=str(PROJECT_DIR / "data"))
    run_parser.add_argument("--snapshot-dir", default=str(PROJECT_DIR / "_internal" / "snapshots"))
    run_parser.add_argument("--report-dir", default=str(PROJECT_DIR / "report" / "history"))
    run_parser.add_argument("--latest-report-path", default=str(PROJECT_DIR / "report" / "latest.md"))
    run_parser.add_argument("--latest-xlsx-path", default=str(PROJECT_DIR / "report" / "latest.xlsx"))
    run_parser.add_argument("--words-dir", default=str(REPO_ROOT / "words"))
    run_parser.add_argument("--chain-work-dir", default=str(PROJECT_DIR / "_internal" / "chained"))

    rebuild_parser = subparsers.add_parser("rebuild-reports", help="根据快照重建历史报告")
    rebuild_parser.add_argument("--snapshot-dir", default=str(PROJECT_DIR / "_internal" / "snapshots"))
    rebuild_parser.add_argument("--report-dir", default=str(PROJECT_DIR / "report" / "history"))
    rebuild_parser.add_argument("--latest-report-path", default=str(PROJECT_DIR / "report" / "latest.md"))
    rebuild_parser.add_argument("--latest-xlsx-path", default=str(PROJECT_DIR / "report" / "latest.xlsx"))

    validate_parser = subparsers.add_parser("validate-report", help="校验 Markdown 报告结构与 latest.xlsx")
    validate_parser.add_argument("--report", default=str(PROJECT_DIR / "report" / "latest.md"))
    validate_parser.add_argument("--xlsx", default=str(PROJECT_DIR / "report" / "latest.xlsx"))

    return parser.parse_args()


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def safe_float(value) -> float:
    try:
        if value is None or value == "":
            return 0.0
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def to_int_if_possible(v: float) -> str:
    if abs(v - round(v)) < 1e-9:
        return str(int(round(v)))
    return f"{v:.2f}"


def to_pct(ratio: float) -> str:
    return f"{ratio * 100:.1f}%"


def read_token(path: Path) -> str:
    token = path.read_text(encoding="utf-8").strip()
    if not token:
        raise RuntimeError(f"token 文件为空: {path}")
    return token


def now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d-%H%M%S")


def load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def dump_json(path: Path, payload: dict) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def normalize_keyword_text(text: str) -> str:
    if not isinstance(text, str):
        return ""
    return SPACE_RE.sub(" ", text.strip().lower())


def compute_sim_score(row: dict) -> Optional[float]:
    sim_window_volume = safe_float(row.get("simWindowVolume"))
    sim_cpc = safe_float(row.get("simCpc"))
    sim_kd = safe_float(row.get("simKd"))
    if sim_window_volume <= 0 or sim_cpc <= 0 or sim_kd <= 0:
        return None
    return sim_window_volume * sim_cpc / sim_kd


def _normalize_export_column(raw: dict) -> dict:
    if not isinstance(raw, dict):
        raise RuntimeError("标准词表列配置项必须是对象")

    header = str(raw.get("header") or "").strip()
    field = str(raw.get("field") or "").strip()
    if not header or not field:
        raise RuntimeError("标准词表列配置缺少 header/field")

    column = {
        "header": header,
        "field": field,
    }

    transform = raw.get("transform")
    if transform not in (None, ""):
        column["transform"] = str(transform)

    number_format = raw.get("number_format")
    if number_format not in (None, ""):
        column["number_format"] = str(number_format)

    width_profile = raw.get("width_profile")
    if isinstance(width_profile, dict):
        normalized_width: dict = {}
        if width_profile.get("fixed") is not None:
            normalized_width["fixed"] = width_profile.get("fixed")
        if width_profile.get("min") is not None:
            normalized_width["min"] = width_profile.get("min")
        if width_profile.get("max") is not None:
            normalized_width["max"] = width_profile.get("max")
        if normalized_width:
            column["width_profile"] = normalized_width

    return column


def load_standard_word_table_spec() -> dict:
    if not STANDARD_WORD_TABLE_SPEC_PATH.exists():
        raise RuntimeError(
            "标准词表规范缺失: "
            f"{STANDARD_WORD_TABLE_SPEC_PATH}。"
            "请先补齐 standard-word-analysis/spec/standard-word-table.v1.json。"
        )

    spec = load_json(STANDARD_WORD_TABLE_SPEC_PATH)
    if not isinstance(spec, dict):
        raise RuntimeError("标准词表规范格式错误：顶层必须是对象")

    version = str(spec.get("version") or "").strip()
    if version != STANDARD_WORD_TABLE_VERSION:
        raise RuntimeError(
            f"标准词表版本不匹配：expected={STANDARD_WORD_TABLE_VERSION} actual={version or '-'}"
        )

    raw_columns = spec.get("excelColumns")
    if not isinstance(raw_columns, list) or not raw_columns:
        raise RuntimeError("标准词表规范缺少 excelColumns 数组")

    columns = [_normalize_export_column(item) for item in raw_columns]
    fields = [str(column.get("field") or "") for column in columns]
    if len(fields) != len(set(fields)):
        raise RuntimeError("标准词表规范字段重复：excelColumns.field 必须唯一")

    return {
        "version": version,
        "columns": columns,
    }


def get_standard_word_table_spec() -> dict:
    return STANDARD_WORD_TABLE_SPEC


def get_keywords_export_columns() -> List[dict]:
    return list(get_standard_word_table_spec()["columns"])


def get_keywords_export_headers() -> List[str]:
    return [column["header"] for column in get_keywords_export_columns()]




def get_header_for_field(field_name: str) -> str:
    for column in get_keywords_export_columns():
        if column.get("field") == field_name:
            return str(column.get("header") or "")
    raise RuntimeError(f"标准词表缺少字段: {field_name}")


def enrich_standard_word_rows_with_gefei_kd(rows: List[dict]) -> dict:
    keywords = [str(row.get("keyword") or "").strip() for row in rows if str(row.get("keyword") or "").strip()]
    result = fetch_gefei_kd_rows(keywords=keywords)
    score_by_keyword = result.get("scoreByKeyword") or {}
    for row in rows:
        keyword = normalize_keyword_text(str(row.get("keyword") or ""))
        row["gefeiKD"] = score_by_keyword.get(keyword)
    return result


def get_score_column_index_from_headers(headers: Sequence[str]) -> int:
    score_header = str(get_standard_word_table_spec().get("scoreColumnHeader") or "").strip()
    if not score_header:
        score_header = "score(simWindowVolume*cpc/kd)"
    try:
        return list(headers).index(score_header)
    except ValueError as exc:
        raise RuntimeError(f"标准词表缺少 score 列: {score_header}") from exc


def normalize_context_item(*, hostname: str, landing_page_url: str) -> str:
    host = str(hostname or "").strip().lower()
    url = str(landing_page_url or "").strip()

    if not host and not url:
        return ""
    if not url:
        return host
    if not host:
        return url

    try:
        parsed = urlsplit(url)
    except ValueError:
        parsed = None

    if parsed is not None:
        parsed_host = (parsed.hostname or "").strip().lower().rstrip(".")
        if parsed_host and parsed_host == host:
            return url

    return f"{host} {url}"


def stable_unique_texts(values: Sequence[str]) -> List[str]:
    seen: set[str] = set()
    result: List[str] = []
    for value in values:
        item = str(value or "").strip()
        if not item:
            continue
        key = normalize_keyword_text(item)
        if key in seen:
            continue
        seen.add(key)
        result.append(item)
    return result


def _normalize_host_token(token: str) -> str:
    host = str(token or "").strip().lower().rstrip(".")
    if not host or " " in host:
        return ""
    if "://" in host or "/" in host:
        return ""
    if "." not in host:
        return ""
    return host


def _extract_host_from_url_candidate(candidate: str) -> str:
    text = str(candidate or "").strip()
    if not text:
        return ""
    try:
        parsed = urlsplit(text)
    except ValueError:
        return ""
    return (parsed.hostname or "").strip().lower().rstrip(".")


def _split_context_items_without_host_dup(values: Sequence[str]) -> List[str]:
    unique_items = stable_unique_texts(values)
    url_hosts: set[str] = set()

    for item in unique_items:
        parts = str(item or "").strip().split()
        if not parts:
            continue

        host_from_full_item = _extract_host_from_url_candidate(str(item))
        if host_from_full_item:
            url_hosts.add(host_from_full_item)
            continue

        host_from_last_part = _extract_host_from_url_candidate(parts[-1])
        if host_from_last_part:
            url_hosts.add(host_from_last_part)

    result: List[str] = []
    for item in unique_items:
        parts = str(item or "").strip().split()
        if not parts:
            continue

        if len(parts) == 1:
            bare_host = _normalize_host_token(parts[0])
            if bare_host and bare_host in url_hosts:
                continue

        result.append(item)

    return result


def aggregate_keyword_contexts(rows: List[dict]) -> str:
    values = [
        normalize_context_item(
            hostname=str(row.get("hostname") or ""),
            landing_page_url=str(row.get("landingPageUrl") or ""),
        )
        for row in rows
    ]
    deduped_values = _split_context_items_without_host_dup(values)
    return " | ".join(deduped_values)


STANDARD_WORD_TABLE_SPEC = load_standard_word_table_spec()


def normalize_landing_page_url(raw_url: str) -> Optional[Tuple[str, str]]:
    if not isinstance(raw_url, str):
        return None
    text = raw_url.strip()
    if not text:
        return None

    if "://" not in text:
        text = f"https://{text}"

    try:
        parsed = urlsplit(text)
    except ValueError:
        return None

    hostname = (parsed.hostname or "").strip().lower().rstrip(".")
    if not hostname:
        return None

    path = parsed.path or "/"
    if path != "/":
        path = "/" + path.lstrip("/")
        path = path.rstrip("/") or "/"

    query = parsed.query.strip()
    canonical = f"https://{hostname}{path}"
    if query:
        canonical += f"?{query}"

    return canonical, hostname


def extract_subdomain(hostname: str, key: str) -> Optional[str]:
    key = key.strip().lower()
    suffix = f".{key}"
    if hostname == key:
        return None
    if not hostname.endswith(suffix):
        return None
    subdomain = hostname[: -len(suffix)]
    if not subdomain:
        return None
    return subdomain


def _post_json(url: str, headers: dict, payload: dict, timeout_seconds: int) -> Tuple[int, str]:
    req = Request(
        url,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers=headers,
        method="POST",
    )
    try:
        with urlopen(req, timeout=timeout_seconds) as resp:
            body = resp.read().decode("utf-8", errors="replace")
            return int(getattr(resp, "status", 0) or 0), body
    except HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace") if hasattr(exc, "read") else ""
        return int(getattr(exc, "code", 0) or 0), body


def fetch_page_with_retry(
    *,
    api_url: str,
    headers: dict,
    base_payload: dict,
    page: int,
) -> dict:
    payload = dict(base_payload)
    payload["page"] = page

    last_error = None
    for attempt in range(MAX_RETRIES + 1):
        try:
            status_code, raw_body = _post_json(
                api_url,
                headers=headers,
                payload=payload,
                timeout_seconds=REQUEST_TIMEOUT_SECONDS,
            )

            wrapper = json.loads(raw_body)
            if not isinstance(wrapper, dict):
                raise RuntimeError("本地服务返回非 JSON 对象")
            if not wrapper.get("ok"):
                err = wrapper.get("error") or {}
                raise RuntimeError(f"本地服务错误: {err.get('code', 'UNKNOWN')} {err.get('message', '')}".strip())

            data = wrapper.get("data") or {}
            upstream_status = int(data.get("status", status_code) or 0)
            if upstream_status != 200:
                raise RuntimeError(f"上游状态码非 200: {upstream_status}")

            upstream_body_raw = data.get("body")
            if not isinstance(upstream_body_raw, str):
                raise RuntimeError("上游 body 缺失或类型错误")
            upstream = json.loads(upstream_body_raw)
            rows = upstream.get("Data")
            if not isinstance(rows, list):
                raise RuntimeError("上游响应 Data 非数组")

            return {
                "page": page,
                "attempt": attempt + 1,
                "requestPayload": payload,
                "wrapper": wrapper,
                "upstream": upstream,
                "rowsCount": len(rows),
            }
        except (RuntimeError, ValueError, json.JSONDecodeError, URLError) as exc:
            last_error = exc
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_BACKOFF_SECONDS[min(attempt, len(RETRY_BACKOFF_SECONDS) - 1)])
                continue
            break

    raise RuntimeError(f"page={page} 抓取失败（已重试 {MAX_RETRIES} 次）: {last_error}")


def fetch_pages(args: argparse.Namespace, token: str) -> Tuple[List[dict], dict]:
    api_url = args.api_base.rstrip("/") + args.endpoint
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }
    base_payload = {
        "key": args.key,
        "country": args.country,
        "latest": args.latest,
        "webSource": args.web_source,
        "sourceType": args.source_type,
        "sort": args.sort,
        "asc": bool(args.asc),
        "includeSubDomains": bool(args.include_subdomains),
        "isWindow": bool(args.is_window),
        "searchType": args.search_type,
    }

    page_results: List[dict] = []
    for page in range(args.start_page, args.end_page + 1):
        result = fetch_page_with_retry(
            api_url=api_url,
            headers=headers,
            base_payload=base_payload,
            page=page,
        )
        page_results.append(result)

    meta = {
        "apiUrl": api_url,
        "pagesRequested": list(range(args.start_page, args.end_page + 1)),
        "basePayload": base_payload,
    }
    return page_results, meta


def normalize_rows(page_results: List[dict], key: str) -> Tuple[List[dict], dict]:
    rows: List[dict] = []

    stats = {
        "rawRows": 0,
        "invalidUrlRows": 0,
        "nonTargetRows": 0,
    }

    for result in page_results:
        page_num = result["page"]
        upstream = result.get("upstream") or {}
        data_rows = upstream.get("Data") or []

        for idx, item in enumerate(data_rows, start=1):
            stats["rawRows"] += 1
            if not isinstance(item, dict):
                stats["invalidUrlRows"] += 1
                continue

            normalized = normalize_landing_page_url(item.get("Url", ""))
            if not normalized:
                stats["invalidUrlRows"] += 1
                continue

            landing_page_url, hostname = normalized
            subdomain = extract_subdomain(hostname, key)
            if not subdomain:
                stats["nonTargetRows"] += 1
                continue

            rows.append(
                {
                    "landingPageUrl": landing_page_url,
                    "hostname": hostname,
                    "subdomain": subdomain,
                    "clicks": safe_float(item.get("Clicks")),
                    "clicksShare": safe_float(item.get("ClicksShare")),
                    "clicksChangeApi": safe_float(item.get("ClicksChange")),
                    "topKeyword": (item.get("TopKeyword") or "") if isinstance(item.get("TopKeyword"), str) else "",
                    "page": page_num,
                    "rankInPage": idx,
                }
            )

    return rows, stats


def _is_better_sample(candidate: dict, existing: dict) -> bool:
    return (candidate["page"], candidate["rankInPage"]) < (existing["page"], existing["rankInPage"])


def dedupe_rows(rows: List[dict]) -> List[dict]:
    deduped: Dict[str, dict] = {}
    for row in rows:
        key = row["landingPageUrl"]
        old = deduped.get(key)
        if old is None or _is_better_sample(row, old):
            deduped[key] = row

    result = list(deduped.values())
    result.sort(key=lambda r: (-safe_float(r.get("clicks")), r.get("landingPageUrl", "")))
    return result


def aggregate_subdomains(rows: List[dict]) -> List[dict]:
    grouped: Dict[str, dict] = {}
    for row in rows:
        subdomain = row.get("subdomain", "")
        item = grouped.setdefault(
            subdomain,
            {
                "subdomain": subdomain,
                "observedSubdomainClicks": 0.0,
                "landingPagesCount": 0,
            },
        )
        item["observedSubdomainClicks"] += safe_float(row.get("clicks"))
        item["landingPagesCount"] += 1

    result = list(grouped.values())
    result.sort(key=lambda x: (-safe_float(x.get("observedSubdomainClicks")), x.get("subdomain", "")))
    return result


def list_snapshot_files(snapshot_dir: Path) -> List[Path]:
    if not snapshot_dir.exists():
        return []
    files = []
    for p in snapshot_dir.iterdir():
        if p.is_file() and SNAPSHOT_RE.match(p.name):
            files.append(p)
    files.sort(key=lambda p: p.name)
    return files


def snapshot_matches_target(snapshot: dict, target: dict) -> bool:
    meta = snapshot.get("meta") or {}
    meta_target = meta.get("target") or {}
    meta_request = meta.get("request") or {}
    return (
        str(meta_target.get("key", "")) == str(target.get("key", ""))
        and str(meta_target.get("country", "")) == str(target.get("country", ""))
        and str(meta_target.get("latest", "")) == str(target.get("latest", ""))
        and str(meta_target.get("sourceType", "")) == str(target.get("sourceType", ""))
        and int(meta_request.get("startPage", 1) or 1) == int(target.get("startPage", 1) or 1)
        and int(meta_request.get("endPage", 8) or 8) == int(target.get("endPage", 8) or 8)
    )


def find_recent_baselines(
    snapshot_dir: Path,
    target: dict,
    current_stamp: str,
    *,
    limit: int = 2,
) -> List[Tuple[dict, Path]]:
    if limit <= 0:
        return []

    candidates = list_snapshot_files(snapshot_dir)
    baselines: List[Tuple[dict, Path]] = []
    for path in reversed(candidates):
        m = SNAPSHOT_RE.match(path.name)
        if not m:
            continue
        stamp = m.group(1)
        if stamp >= current_stamp:
            continue
        snapshot = load_json(path)
        if snapshot_matches_target(snapshot, target):
            baselines.append((snapshot, path))
            if len(baselines) >= limit:
                break

    return baselines


def find_latest_baseline(snapshot_dir: Path, target: dict, current_stamp: str) -> Tuple[Optional[dict], Optional[Path]]:
    baselines = find_recent_baselines(
        snapshot_dir,
        target,
        current_stamp,
        limit=1,
    )
    if baselines:
        return baselines[0]
    return None, None



def build_page_comparison(today_rows: List[dict], baseline_rows: List[dict]) -> Tuple[List[dict], List[dict]]:
    baseline_map = {row["landingPageUrl"]: row for row in baseline_rows}
    newly: List[dict] = []
    rising: List[dict] = []

    for row in today_rows:
        url = row["landingPageUrl"]
        today_clicks = safe_float(row.get("clicks"))
        old = baseline_map.get(url)

        if old is None:
            if today_clicks >= PAGE_NEW_MIN_CLICKS:
                newly.append(row)
            continue

        old_clicks = safe_float(old.get("clicks"))
        delta = today_clicks - old_clicks
        if old_clicks <= 0:
            continue
        growth = delta / old_clicks

        if (
            today_clicks >= PAGE_RISING_MIN_CLICKS
            and delta >= PAGE_RISING_MIN_DELTA
            and growth > PAGE_RISING_MIN_GROWTH
        ):
            item = dict(row)
            item.update(
                {
                    "baseClicks": old_clicks,
                    "deltaClicks": delta,
                    "growthRate": growth,
                }
            )
            rising.append(item)

    newly.sort(key=lambda x: (-safe_float(x.get("clicks")), x.get("landingPageUrl", "")))
    rising.sort(key=lambda x: (-safe_float(x.get("deltaClicks")), x.get("landingPageUrl", "")))
    return newly, rising


def _is_subdomain_rising_by_threshold(today_clicks: float, old_clicks: float) -> Tuple[bool, float, float]:
    delta = today_clicks - old_clicks
    if old_clicks <= 0:
        return False, delta, 0.0

    growth = delta / old_clicks
    is_rising = (
        today_clicks >= SUBDOMAIN_RISING_MIN_CLICKS
        and delta >= SUBDOMAIN_RISING_MIN_DELTA
        and growth > SUBDOMAIN_RISING_MIN_GROWTH
    )
    return is_rising, delta, growth


def build_subdomain_comparison(today_subdomains: List[dict], baseline_subdomains: List[dict]) -> Tuple[List[dict], List[dict]]:
    baseline_map = {row["subdomain"]: row for row in baseline_subdomains}
    newly: List[dict] = []
    rising: List[dict] = []

    for row in today_subdomains:
        subdomain = row["subdomain"]
        today_clicks = safe_float(row.get("observedSubdomainClicks"))
        old = baseline_map.get(subdomain)

        if old is None:
            if today_clicks >= SUBDOMAIN_NEW_MIN_CLICKS:
                newly.append(row)
            continue

        old_clicks = safe_float(old.get("observedSubdomainClicks"))
        is_rising, delta, growth = _is_subdomain_rising_by_threshold(today_clicks, old_clicks)
        if is_rising:
            item = dict(row)
            item.update(
                {
                    "baseObservedSubdomainClicks": old_clicks,
                    "deltaClicks": delta,
                    "growthRate": growth,
                }
            )
            rising.append(item)

    newly.sort(key=lambda x: (-safe_float(x.get("observedSubdomainClicks")), x.get("subdomain", "")))
    rising.sort(key=lambda x: (-safe_float(x.get("deltaClicks")), x.get("subdomain", "")))
    return newly, rising


def filter_consecutive_rising_subdomains(
    *,
    rising_subdomains_current: List[dict],
    baseline_subdomains: List[dict],
    prev_baseline_subdomains: Optional[List[dict]],
) -> Tuple[List[dict], dict]:
    available_comparisons = 2 if prev_baseline_subdomains else 1
    if not prev_baseline_subdomains:
        return [], {
            "requiredComparisons": 2,
            "availableComparisons": available_comparisons,
            "currentPassCount": len(rising_subdomains_current),
            "previousPassCount": 0,
            "qualifiedCount": 0,
        }

    _new_prev, rising_previous = build_subdomain_comparison(baseline_subdomains, prev_baseline_subdomains)
    previous_set = {str(row.get("subdomain") or "") for row in rising_previous if str(row.get("subdomain") or "")}

    qualified = [
        row for row in rising_subdomains_current if str(row.get("subdomain") or "") in previous_set
    ]
    qualified.sort(key=lambda x: (-safe_float(x.get("deltaClicks")), x.get("subdomain", "")))

    return qualified, {
        "requiredComparisons": 2,
        "availableComparisons": available_comparisons,
        "currentPassCount": len(rising_subdomains_current),
        "previousPassCount": len(rising_previous),
        "qualifiedCount": len(qualified),
    }



def _path_from_url(raw_url: str) -> str:
    if not isinstance(raw_url, str) or not raw_url:
        return "-"
    try:
        parsed = urlsplit(raw_url)
    except ValueError:
        return "-"

    path = parsed.path or "/"
    if path != "/":
        path = "/" + path.lstrip("/")
        path = path.rstrip("/") or "/"

    if parsed.query:
        return f"{path}?{parsed.query}"
    return path


def build_subdomain_keyword_entries(rows: List[dict], *, max_keywords: int = 3) -> Dict[str, List[dict]]:
    grouped: Dict[str, List[dict]] = {}
    for row in rows:
        subdomain = row.get("subdomain") or ""
        if not subdomain:
            continue
        grouped.setdefault(subdomain, []).append(row)

    result: Dict[str, List[dict]] = {}
    for subdomain, items in grouped.items():
        sorted_items = sorted(
            items,
            key=lambda item: (-safe_float(item.get("clicks")), item.get("landingPageUrl", "")),
        )
        keywords: List[dict] = []
        seen = set()
        for item in sorted_items:
            keyword = (item.get("topKeyword") or "").strip()
            if not keyword or keyword in seen:
                continue
            seen.add(keyword)
            keywords.append(
                {
                    "keyword": keyword,
                    "correspondingDomain": (item.get("hostname") or "").strip().lower(),
                }
            )
            if len(keywords) >= max_keywords:
                break
        result[subdomain] = keywords

    return result


def build_subdomain_keywords_map(subdomain_keyword_entries: Dict[str, List[dict]]) -> Dict[str, str]:
    result: Dict[str, str] = {}
    for subdomain, entries in subdomain_keyword_entries.items():
        keywords = [str(entry.get("keyword") or "").strip() for entry in entries if str(entry.get("keyword") or "").strip()]
        result[subdomain] = " / ".join(keywords) if keywords else "-"
    return result


def _page_report_row(row: dict, *, trend: str) -> dict:
    growth = safe_float(row.get("growthRate")) if trend == "上涨" else None
    trend_label = trend if growth is None else f"{trend}（+{to_pct(growth)}）"
    return {
        "entityType": "page",
        "trend": trend,
        "trendLabel": trend_label,
        "subdomain": row.get("subdomain", "-") or "-",
        "path": _path_from_url(row.get("landingPageUrl", "")),
        "clicks": safe_float(row.get("clicks")),
        "topKeywords": (row.get("topKeyword") or "-").strip() or "-",
    }


def _subdomain_report_row(row: dict, *, trend: str, subdomain_keywords: Dict[str, str]) -> dict:
    growth = safe_float(row.get("growthRate")) if trend == "上涨" else None
    trend_label = trend if growth is None else f"{trend}（+{to_pct(growth)}）"
    subdomain = row.get("subdomain", "-") or "-"
    return {
        "entityType": "subdomain",
        "trend": trend,
        "trendLabel": trend_label,
        "subdomain": subdomain,
        "path": "-",
        "clicks": safe_float(row.get("observedSubdomainClicks")),
        "topKeywords": subdomain_keywords.get(subdomain, "-"),
    }


def _make_standard_word_row(keyword: str, corresponding_domain: str) -> dict:
    return {
        "keyword": keyword,
        "correspondingDomain": corresponding_domain,
        "group": "",
        "sourcePresence": "",
        "score": None,
        "simWindowVolume": None,
        "simKd": None,
        "simCpc": None,
        "semVolume": None,
        "semKd": None,
        "semCpc": None,
    }


def build_standard_word_rows(
    *,
    new_page_rows: List[dict],
    new_subdomain_rows: List[dict],
    rising_subdomain_rows: List[dict],
    subdomain_keyword_entries: Dict[str, List[dict]],
    all_today_rows: List[dict],
) -> Tuple[List[dict], dict]:
    rows: List[dict] = []
    keyword_map: Dict[str, dict] = {}
    new_page_keyword_rows = 0
    new_subdomain_keyword_rows = 0
    rising_subdomain_keyword_rows = 0

    rows_by_keyword: Dict[str, List[dict]] = {}
    for item in all_today_rows:
        keyword = str(item.get("topKeyword") or "").strip()
        normalized_keyword = normalize_keyword_text(keyword)
        if not normalized_keyword:
            continue
        rows_by_keyword.setdefault(normalized_keyword, []).append(item)

    def add_keyword(keyword: str) -> bool:
        normalized_keyword = normalize_keyword_text(keyword)
        if not normalized_keyword:
            return False
        if normalized_keyword in keyword_map:
            return False
        source_rows = rows_by_keyword.get(normalized_keyword, [])
        keyword_map[normalized_keyword] = _make_standard_word_row(
            keyword.strip(),
            aggregate_keyword_contexts(source_rows),
        )
        rows.append(keyword_map[normalized_keyword])
        return True

    for row in new_subdomain_rows:
        subdomain = str(row.get("subdomain") or "").strip()
        for entry in subdomain_keyword_entries.get(subdomain, []):
            if add_keyword(str(entry.get("keyword") or "")):
                new_subdomain_keyword_rows += 1

    for row in rising_subdomain_rows:
        subdomain = str(row.get("subdomain") or "").strip()
        for entry in subdomain_keyword_entries.get(subdomain, []):
            if add_keyword(str(entry.get("keyword") or "")):
                rising_subdomain_keyword_rows += 1

    for row in new_page_rows:
        if add_keyword(str(row.get("topKeyword") or "")):
            new_page_keyword_rows += 1

    return rows, {
        "rowCount": len(rows),
        "newPageKeywordRows": new_page_keyword_rows,
        "newSubdomainKeywordRows": new_subdomain_keyword_rows,
        "risingSubdomainKeywordRows": rising_subdomain_keyword_rows,
        "qualifiedRisingSubdomainCount": len(rising_subdomain_rows),
    }


def build_comparison(
    *,
    today_rows: List[dict],
    today_subdomains: List[dict],
    baseline_rows: List[dict],
    baseline_subdomains: List[dict],
    baseline_stamp: Optional[str],
    prev_baseline_subdomains: Optional[List[dict]] = None,
    prev_baseline_stamp: Optional[str] = None,
    enrich_gefei_kd: bool = True,
) -> dict:
    new_page, rising_page = build_page_comparison(today_rows, baseline_rows)
    new_sub, rising_sub_current = build_subdomain_comparison(today_subdomains, baseline_subdomains)
    rising_sub, subdomain_rising_meta = filter_consecutive_rising_subdomains(
        rising_subdomains_current=rising_sub_current,
        baseline_subdomains=baseline_subdomains,
        prev_baseline_subdomains=prev_baseline_subdomains,
    )
    subdomain_rising_meta["baselineStamp"] = baseline_stamp
    subdomain_rising_meta["previousBaselineStamp"] = prev_baseline_stamp

    subdomain_keyword_entries = build_subdomain_keyword_entries(today_rows)
    subdomain_keywords = build_subdomain_keywords_map(subdomain_keyword_entries)
    standard_word_rows, standard_word_summary = build_standard_word_rows(
        new_page_rows=new_page,
        new_subdomain_rows=new_sub,
        rising_subdomain_rows=rising_sub,
        subdomain_keyword_entries=subdomain_keyword_entries,
        all_today_rows=today_rows,
    )
    gefei_kd_fetch = enrich_standard_word_rows_with_gefei_kd(standard_word_rows) if enrich_gefei_kd else {
        "summary": {
            "inputCount": 0,
            "requestCount": 0,
            "successCount": 0,
            "successWithScoreCount": 0,
            "missingScoreCount": 0,
            "failedCount": 0,
        },
        "failures": [],
        "api": {},
    }

    report_rows: List[dict] = []
    report_rows.extend(_subdomain_report_row(row, trend="新增", subdomain_keywords=subdomain_keywords) for row in new_sub)
    report_rows.extend(_subdomain_report_row(row, trend="上涨", subdomain_keywords=subdomain_keywords) for row in rising_sub)
    report_rows.extend(_page_report_row(row, trend="新增") for row in new_page)
    report_rows.extend(_page_report_row(row, trend="上涨") for row in rising_page)

    trend_order = {"上涨": 0, "新增": 1}
    entity_order = {"subdomain": 0, "page": 1}
    report_rows.sort(
        key=lambda row: (
            trend_order.get(row.get("trend", ""), 9),
            entity_order.get(row.get("entityType", ""), 9),
            -safe_float(row.get("clicks")),
            row.get("subdomain", ""),
            row.get("path", ""),
        )
    )

    return {
        "baselineStamp": baseline_stamp,
        "reportRows": report_rows,
        "standardWordRows": standard_word_rows,
        "standardWordSummary": standard_word_summary,
        "subdomainRisingMeta": subdomain_rising_meta,
        "gefeiKD": {
            "summary": gefei_kd_fetch.get("summary") or {},
            "failures": gefei_kd_fetch.get("failures") or [],
            "api": gefei_kd_fetch.get("api") or {},
        },
    }


def _fmt_md_cell(value) -> str:
    if value is None:
        text = "-"
    else:
        text = str(value).strip()
    if not text:
        text = "-"

    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    normalized = re.sub(r"\s*\|\s*", "\n", normalized)
    normalized = normalized.replace(" / ", "\n")
    lines = [line.strip() for line in normalized.split("\n") if line.strip()]
    if not lines:
        lines = ["-"]

    escaped: List[str] = []
    for line in lines:
        line = re.sub(r"([/?&#=_-])", r"\1<wbr>", line)
        escaped.append(line.replace("|", "\\|"))
    return "<br>".join(escaped)


def _md_table(headers: List[str], rows: List[List[str]]) -> List[str]:
    if not rows:
        return ["（无）"]
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(_fmt_md_cell(cell) for cell in row) + " |")
    return lines


def render_report(snapshot: dict) -> str:
    meta = snapshot.get("meta") or {}
    target = meta.get("target") or {}
    comparison = snapshot.get("comparison") or {}
    standard_word_rows = comparison.get("standardWordRows") or []
    standard_word_summary = comparison.get("standardWordSummary") or {}

    report_rows = comparison.get("reportRows") or []
    rising_count = sum(1 for row in report_rows if row.get("trend") == "上涨")
    new_count = sum(1 for row in report_rows if row.get("trend") == "新增")

    def has_sim_metrics(row: dict) -> bool:
        return (
            safe_float(row.get("simWindowVolume")) > 0
            and safe_float(row.get("simCpc")) > 0
            and safe_float(row.get("simKd")) > 0
        )

    def has_sem_metrics(row: dict) -> bool:
        return (
            safe_float(row.get("semVolume")) > 0
            and safe_float(row.get("semCpc")) > 0
            and safe_float(row.get("semKd")) > 0
        )

    def has_gefei_kd(row: dict) -> bool:
        value = row.get("gefeiKD")
        return value not in (None, "")

    def _display_optional_number(value) -> str:
        if value in (None, ""):
            return "-"
        number = safe_float(value)
        if abs(number - round(number)) < 1e-9:
            return str(int(round(number)))
        return f"{number:.2f}"

    sim_ready_count = sum(1 for row in standard_word_rows if has_sim_metrics(row))
    sem_ready_count = sum(1 for row in standard_word_rows if has_sem_metrics(row))
    gefei_kd_ready_count = sum(1 for row in standard_word_rows if has_gefei_kd(row))

    lines: List[str] = []
    lines.append(f"# 子域名落地页监控报告（{target.get('key', '-') }｜{meta.get('stamp', '-') }）")
    lines.append("")
    lines.append("## 摘要")
    lines.append("")
    lines.append(f"- 抓取页数：{meta.get('pagesFetched', 0)}（page={meta.get('startPage', 1)}..{meta.get('endPage', 8)}）")
    lines.append(f"- 原始行数：{meta.get('rawRows', 0)}")
    lines.append(f"- 去重后页面数：{meta.get('dedupedRows', 0)}")
    lines.append(f"- 观测子域名数：{meta.get('subdomainCount', 0)}")

    if meta.get("baselineMode"):
        lines.append("- 基线：无历史快照（本次仅建立基线）")
    else:
        lines.append(f"- 基线：{comparison.get('baselineStamp', '-')}")

    lines.append(f"- 新增数量：{new_count}")
    lines.append(f"- 上涨数量：{rising_count}")
    lines.append(f"- 标准词表行数：{standard_word_summary.get('rowCount', len(standard_word_rows))}（最终完整词表）")
    lines.append(f"- SIM 指标完整行数：{sim_ready_count}")
    lines.append(f"- SEM 指标完整行数：{sem_ready_count}")
    lines.append(f"- gefeiKD 已回填行数：{gefei_kd_ready_count}")
    lines.append("- 标准词表按 keyword 去重，`对应域名` 聚合同词命中的全部域名 / 页面上下文")
    lines.append("")

    lines.append("## 监控结果")
    lines.append("")

    if meta.get("baselineMode"):
        lines.append("本次仅建立基线，不输出新增/上涨结论。")
        lines.append("")
    else:
        table_rows = [
            [
                row.get("subdomain", "-"),
                row.get("path", "-"),
                to_int_if_possible(safe_float(row.get("clicks"))),
                row.get("trendLabel", row.get("trend", "-")) or "-",
                row.get("topKeywords", "-") or "-",
            ]
            for row in report_rows
        ]
        lines.extend(_md_table(["subdomain", "path", "clicks", "trend", "top keywords"], table_rows))
        lines.append("")

    lines.append("## 最终标准词表摘要")
    lines.append("")
    final_table_rows = [
        [
            row.get("keyword", "-") or "-",
            _display_optional_number(row.get("simWindowVolume")),
            _display_optional_number(row.get("semVolume")),
            _display_optional_number(row.get("gefeiKD")),
            _display_optional_number(compute_sim_score(row)),
        ]
        for row in standard_word_rows[:20]
    ]
    lines.extend(_md_table(["keyword", "simVolume", "semVolume", "gefeiKD", "score(sim)"], final_table_rows))
    lines.append("")

    lines.append("## 备注")
    lines.append("")
    lines.append("- 标准词表为最终口径：种子词表经 analyze-words（SIM/SEM）与 check-gefei-kd（gefeiKD）补全后写入 report 与 words 目录")
    lines.append("- 标准词表按 keyword 去重，`对应域名` 聚合同词命中的全部域名 / 页面上下文，`gefeiKD` 为哥飞 KD score")
    for remark in REQUIRED_REMARKS:
        lines.append(f"- {remark}")
    lines.append("")

    return "\n".join(lines)


def build_snapshot(
    *,
    stamp: str,
    args: argparse.Namespace,
    rows: List[dict],
    subdomains: List[dict],
    comparison: dict,
    baseline_mode: bool,
    fetch_meta: dict,
    normalize_stats: dict,
    report_history_path: Path,
    excel_history_path: Path,
    latest_report_path: Path,
    latest_xlsx_path: Path,
) -> dict:
    return {
        "meta": {
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "stamp": stamp,
            "baselineMode": baseline_mode,
            "target": {
                "key": args.key,
                "country": args.country,
                "latest": args.latest,
                "webSource": args.web_source,
                "sourceType": args.source_type,
            },
            "request": {
                "sort": args.sort,
                "asc": bool(args.asc),
                "includeSubDomains": bool(args.include_subdomains),
                "isWindow": bool(args.is_window),
                "searchType": args.search_type,
                "startPage": args.start_page,
                "endPage": args.end_page,
                "standardWordTableVersion": STANDARD_WORD_TABLE_VERSION,
            },
            "api": {
                **fetch_meta,
                "gefeiKD": ((comparison.get("gefeiKD") or {}).get("api") or {}),
            },
            "output": {
                "reportHistoryPath": str(report_history_path),
                "excelHistoryPath": str(excel_history_path),
            },
            "latestReportPath": str(latest_report_path),
            "latestExcelPath": str(latest_xlsx_path),
            "pagesFetched": args.end_page - args.start_page + 1,
            "startPage": args.start_page,
            "endPage": args.end_page,
            "rawRows": normalize_stats["rawRows"],
            "dedupedRows": len(rows),
            "subdomainCount": len(subdomains),
            "invalidUrlRows": normalize_stats["invalidUrlRows"],
            "nonTargetRows": normalize_stats["nonTargetRows"],
            "gefeiKD": {
                "summary": ((comparison.get("gefeiKD") or {}).get("summary") or {}),
                "failures": ((comparison.get("gefeiKD") or {}).get("failures") or []),
            },
        },
        "rows": rows,
        "subdomains": subdomains,
        "comparison": comparison,
    }


def _require_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "缺少依赖 openpyxl。请先执行 `pip3 install openpyxl` 再运行。"
        ) from exc
    return Workbook, Font, Alignment, PatternFill, get_column_letter, load_workbook


MULTILINE_EXPORT_FIELDS = {
    "group",
    "correspondingDomain",
}


def _format_multiline_export_text(field: str, value):
    if value is None:
        return value
    if field not in MULTILINE_EXPORT_FIELDS:
        return value

    text = str(value).strip()
    if not text:
        return text
    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    normalized = re.sub(r"\s*\|\s*", "\n", normalized)
    normalized = normalized.replace(" / ", "\n")
    lines = [line.strip() for line in normalized.split("\n") if line.strip()]
    return "\n".join(lines) if lines else ""



def _transform_export_value(column_def: dict, value, row: Optional[dict] = None):
    transform = column_def.get("transform")
    if transform == "source_presence_label":
        mapping = {
            "both": "both（SIM+SEM）",
            "sim_only": "sim_only（仅 SIM）",
            "sem_only": "sem_only（仅 SEM）",
        }
        transformed = mapping.get(str(value or ""), value)
    elif transform == "sim_score":
        transformed = compute_sim_score(row or {})
    else:
        transformed = value
    field = str(column_def.get("field") or "")
    return _format_multiline_export_text(field, transformed)



def _compute_column_width(values: Sequence[str], width_profile: Optional[dict] = None) -> float:
    profile = width_profile or {}
    fixed = profile.get("fixed")
    if fixed is not None:
        return float(fixed)

    max_length = 0
    for raw in values:
        text = "" if raw is None else str(raw)
        parts = text.splitlines() or [text]
        longest_line = max((len(part) for part in parts), default=0)
        if longest_line > max_length:
            max_length = longest_line

    width = max_length + DEFAULT_COLUMN_PADDING
    min_width = int(profile.get("min", DEFAULT_COLUMN_MIN_WIDTH))
    max_width = int(profile.get("max", DEFAULT_COLUMN_MAX_WIDTH))
    return float(max(min_width, min(width, max_width)))


def write_excel(snapshot: dict, output_path: Path) -> None:
    Workbook, Font, Alignment, PatternFill, get_column_letter, _ = _require_openpyxl()

    meta = snapshot.get("meta") or {}
    target = meta.get("target") or {}
    request = meta.get("request") or {}
    comparison = snapshot.get("comparison") or {}
    output = meta.get("output") or {}
    standard_word_rows = comparison.get("standardWordRows") or []
    standard_word_summary = comparison.get("standardWordSummary") or {}

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    keywords_sheet = workbook.create_sheet("keywords")

    summary_rows = [
        ("key", target.get("key", "")),
        ("generatedAt", meta.get("generatedAt", "")),
        ("stamp", meta.get("stamp", "")),
        ("baselineMode", meta.get("baselineMode", False)),
        ("baselineStamp", comparison.get("baselineStamp", "")),
        ("pagesFetched", meta.get("pagesFetched", 0)),
        ("dedupedRows", meta.get("dedupedRows", 0)),
        ("subdomainCount", meta.get("subdomainCount", 0)),
        ("reportRows", len(comparison.get("reportRows") or [])),
        ("standardWordRows", standard_word_summary.get("rowCount", len(standard_word_rows))),
        ("newPageKeywordRows", standard_word_summary.get("newPageKeywordRows", 0)),
        ("newSubdomainKeywordRows", standard_word_summary.get("newSubdomainKeywordRows", 0)),
        ("standardWordTableVersion", request.get("standardWordTableVersion", STANDARD_WORD_TABLE_VERSION)),
        ("reportHistoryPath", output.get("reportHistoryPath", "")),
        ("excelHistoryPath", output.get("excelHistoryPath", "")),
    ]
    for idx, (name, value) in enumerate(summary_rows, start=1):
        summary_sheet.cell(row=idx, column=1, value=name)
        summary_sheet.cell(row=idx, column=2, value=value)
    summary_sheet.freeze_panes = "A2"

    export_columns = get_keywords_export_columns()
    headers = get_keywords_export_headers()
    keywords_sheet.append(headers)
    for cell in keywords_sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    field_to_index = {str(column.get("field") or ""): index for index, column in enumerate(export_columns, start=1)}
    sim_fill = PatternFill(fill_type="solid", fgColor=SIM_COLUMN_FILL_COLOR)
    sem_fill = PatternFill(fill_type="solid", fgColor=SEM_COLUMN_FILL_COLOR)
    sim_field_indexes = [
        field_to_index[field]
        for field in ("simWindowVolume", "simKd", "simCpc")
        if field in field_to_index
    ]
    sem_field_indexes = [
        field_to_index[field]
        for field in ("semVolume", "semKd", "semCpc")
        if field in field_to_index
    ]

    for row in standard_word_rows:
        export_row = []
        for column in export_columns:
            raw_value = row.get(column["field"])
            export_row.append(_transform_export_value(column, raw_value, row))
        keywords_sheet.append(export_row)

    for row_cells in keywords_sheet.iter_rows(min_row=2, max_row=keywords_sheet.max_row, min_col=1, max_col=len(headers)):
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row_index in range(2, keywords_sheet.max_row + 1):
        for col_index in sim_field_indexes:
            keywords_sheet.cell(row=row_index, column=col_index).fill = sim_fill
        for col_index in sem_field_indexes:
            keywords_sheet.cell(row=row_index, column=col_index).fill = sem_fill

    keywords_sheet.freeze_panes = "A2"
    keywords_sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(keywords_sheet.max_row, 1)}"

    for index, column in enumerate(export_columns, start=1):
        number_format = column.get("number_format")
        if not number_format:
            continue
        col_letter = get_column_letter(index)
        for cell in keywords_sheet[col_letter][1:]:
            cell.number_format = number_format

    for column_index, column in enumerate(export_columns, start=1):
        column_letter = get_column_letter(column_index)
        values = []
        for row_index in range(1, keywords_sheet.max_row + 1):
            value = keywords_sheet.cell(row=row_index, column=column_index).value
            values.append("" if value is None else str(value))
        keywords_sheet.column_dimensions[column_letter].width = _compute_column_width(values, column.get("width_profile"))

    for column_cells in summary_sheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        summary_sheet.column_dimensions[column_cells[0].column_letter].width = _compute_column_width(
            values,
            {"min": 14, "max": 64},
        )

    ensure_dir(output_path.parent)
    workbook.save(output_path)


def write_artifacts(
    *,
    stamp: str,
    snapshot: dict,
    fetch_archive: dict,
    data_dir: Path,
    snapshot_dir: Path,
    report_dir: Path,
    latest_report_path: Path,
    latest_xlsx_path: Path,
) -> dict:
    ensure_dir(data_dir)
    ensure_dir(snapshot_dir)
    ensure_dir(report_dir)
    ensure_dir(latest_report_path.parent)
    ensure_dir(latest_xlsx_path.parent)

    fetch_archive_path = data_dir / f"fetch-{stamp}.json"
    snapshot_path = snapshot_dir / f"snapshot-{stamp}.json"
    report_history_path = report_dir / f"report-{stamp}.md"
    excel_history_path = report_dir / f"keyword-table-{stamp}.xlsx"

    snapshot_meta = snapshot.setdefault("meta", {})
    output_meta = snapshot_meta.setdefault("output", {})
    output_meta["reportHistoryPath"] = str(report_history_path)
    output_meta["excelHistoryPath"] = str(excel_history_path)
    snapshot_meta["latestReportPath"] = str(latest_report_path)
    snapshot_meta["latestExcelPath"] = str(latest_xlsx_path)

    report_text = render_report(snapshot)
    dump_json(fetch_archive_path, fetch_archive)
    dump_json(snapshot_path, snapshot)
    report_history_path.write_text(report_text, encoding="utf-8")
    latest_report_path.write_text(report_text, encoding="utf-8")
    write_excel(snapshot, excel_history_path)
    copyfile(excel_history_path, latest_xlsx_path)

    return {
        "fetchArchivePath": fetch_archive_path,
        "snapshotPath": snapshot_path,
        "reportHistoryPath": report_history_path,
        "excelHistoryPath": excel_history_path,
    }


def rebuild_history_artifacts(snapshot: dict, report_dir: Path, latest_report_path: Path, latest_xlsx_path: Path) -> Tuple[Path, Path, str]:
    ensure_dir(report_dir)
    meta = snapshot.get("meta") or {}
    stamp = str(meta.get("stamp") or "")
    if not stamp:
        raise RuntimeError("snapshot 缺少 meta.stamp")

    report_history_path = report_dir / f"report-{stamp}.md"
    excel_history_path = report_dir / f"keyword-table-{stamp}.xlsx"

    output_meta = meta.setdefault("output", {})
    output_meta["reportHistoryPath"] = str(report_history_path)
    output_meta["excelHistoryPath"] = str(excel_history_path)
    meta["latestReportPath"] = str(latest_report_path)
    meta["latestExcelPath"] = str(latest_xlsx_path)
    request = meta.setdefault("request", {})
    request["standardWordTableVersion"] = STANDARD_WORD_TABLE_VERSION

    report_text = render_report(snapshot)
    report_history_path.write_text(report_text, encoding="utf-8")
    write_excel(snapshot, excel_history_path)
    return report_history_path, excel_history_path, report_text


def _run_python_stage(*, stage_name: str, script_path: Path, extra_args: Sequence[str]) -> None:
    if not script_path.exists():
        raise RuntimeError(f"{stage_name} 脚本不存在: {script_path}")

    command = [sys.executable, str(script_path), *list(extra_args)]
    print(f"[stage] {stage_name}: {' '.join(command)}")

    try:
        subprocess.run(command, cwd=str(REPO_ROOT), check=True)
    except subprocess.CalledProcessError as exc:
        raise RuntimeError(f"{stage_name} 执行失败，退出码: {exc.returncode}") from exc


def _run_analyze_words_stage(*, seed_table_xlsx: Path, chain_root: Path) -> dict:
    stage_root = chain_root / "analyze-words"
    data_dir = stage_root / "data"
    snapshot_dir = stage_root / "_internal" / "snapshots"
    report_dir = stage_root / "report" / "history"
    latest_report_path = stage_root / "report" / "latest.md"
    latest_xlsx_path = stage_root / "report" / "latest.xlsx"
    latest_table_json_path = stage_root / "report" / "latest.json"

    _run_python_stage(
        stage_name="analyze-words",
        script_path=ANALYZE_WORDS_SCRIPT_PATH,
        extra_args=[
            "run",
            "--input-table",
            str(seed_table_xlsx),
            "--data-dir",
            str(data_dir),
            "--snapshot-dir",
            str(snapshot_dir),
            "--report-dir",
            str(report_dir),
            "--latest-report-path",
            str(latest_report_path),
            "--latest-xlsx-path",
            str(latest_xlsx_path),
            "--latest-table-json-path",
            str(latest_table_json_path),
        ],
    )

    if not latest_xlsx_path.exists():
        raise RuntimeError(f"analyze-words 未产出 latest xlsx: {latest_xlsx_path}")
    if not latest_table_json_path.exists():
        raise RuntimeError(f"analyze-words 未产出 latest json: {latest_table_json_path}")

    return {
        "latestReportPath": latest_report_path,
        "latestXlsxPath": latest_xlsx_path,
        "latestJsonPath": latest_table_json_path,
        "root": stage_root,
    }


def _run_check_gefei_kd_stage(*, input_table_xlsx: Path, chain_root: Path) -> dict:
    stage_root = chain_root / "check-gefei-kd"
    data_dir = stage_root / "data"
    snapshot_dir = stage_root / "_internal" / "snapshots"
    report_dir = stage_root / "report" / "history"
    latest_report_path = stage_root / "report" / "latest.md"
    latest_standard_word_table_json = stage_root / "report" / "latest.json"
    latest_standard_word_table_xlsx = stage_root / "report" / "latest.xlsx"

    _run_python_stage(
        stage_name="check-gefei-kd",
        script_path=CHECK_GEFEI_KD_SCRIPT_PATH,
        extra_args=[
            "run",
            "--standard-word-table",
            str(input_table_xlsx),
            "--data-dir",
            str(data_dir),
            "--snapshot-dir",
            str(snapshot_dir),
            "--report-dir",
            str(report_dir),
            "--latest-report-path",
            str(latest_report_path),
            "--latest-standard-word-table-json",
            str(latest_standard_word_table_json),
            "--latest-standard-word-table-xlsx",
            str(latest_standard_word_table_xlsx),
        ],
    )

    if not latest_standard_word_table_xlsx.exists():
        raise RuntimeError(f"check-gefei-kd 未产出 latest xlsx: {latest_standard_word_table_xlsx}")
    if not latest_standard_word_table_json.exists():
        raise RuntimeError(f"check-gefei-kd 未产出 latest json: {latest_standard_word_table_json}")

    return {
        "latestReportPath": latest_report_path,
        "latestXlsxPath": latest_standard_word_table_xlsx,
        "latestJsonPath": latest_standard_word_table_json,
        "root": stage_root,
    }


def _publish_final_words_xlsx(*, final_xlsx: Path, words_dir: Path, stamp: str) -> Path:
    if not final_xlsx.exists():
        raise RuntimeError(f"最终标准词表不存在: {final_xlsx}")

    ensure_dir(words_dir)
    target = words_dir / f"sub-domain-{stamp}.xlsx"
    copyfile(final_xlsx, target)
    return target


def run_pipeline(args: argparse.Namespace) -> int:
    data_dir = Path(args.data_dir).resolve()
    snapshot_dir = Path(args.snapshot_dir).resolve()
    report_dir = Path(args.report_dir).resolve()
    latest_report_path = Path(args.latest_report_path).resolve()
    latest_xlsx_path = Path(args.latest_xlsx_path).resolve()
    token_path = Path(args.token_path).resolve()

    ensure_dir(data_dir)
    ensure_dir(snapshot_dir)
    ensure_dir(report_dir)
    ensure_dir(latest_report_path.parent)
    ensure_dir(latest_xlsx_path.parent)

    stamp = now_stamp()
    report_history_path = report_dir / f"report-{stamp}.md"
    excel_history_path = report_dir / f"keyword-table-{stamp}.xlsx"

    target = {
        "key": args.key,
        "country": args.country,
        "latest": args.latest,
        "sourceType": args.source_type,
        "startPage": args.start_page,
        "endPage": args.end_page,
    }
    recent_baselines = find_recent_baselines(snapshot_dir, target, current_stamp=stamp, limit=2)
    baseline_snapshot, baseline_path = recent_baselines[0] if recent_baselines else (None, None)
    prev_baseline_snapshot, _prev_baseline_path = recent_baselines[1] if len(recent_baselines) > 1 else (None, None)

    baseline_rows = baseline_snapshot.get("rows", []) if baseline_snapshot else []
    baseline_subs = baseline_snapshot.get("subdomains", []) if baseline_snapshot else []
    prev_baseline_subs = prev_baseline_snapshot.get("subdomains", []) if prev_baseline_snapshot else []

    token = read_token(token_path)

    # 先抓取并在内存中处理；任一页失败会抛错并整次退出，不落任何文件。
    page_results, fetch_meta = fetch_pages(args, token)

    normalized_rows, normalize_stats = normalize_rows(page_results, args.key)
    deduped_rows = dedupe_rows(normalized_rows)
    subdomains = aggregate_subdomains(deduped_rows)

    baseline_mode = baseline_snapshot is None
    if baseline_mode:
        comparison = {
            "baselineStamp": None,
            "reportRows": [],
            "standardWordRows": [],
            "standardWordSummary": {
                "rowCount": 0,
                "newPageKeywordRows": 0,
                "newSubdomainKeywordRows": 0,
                "risingSubdomainKeywordRows": 0,
                "qualifiedRisingSubdomainCount": 0,
            },
            "subdomainRisingMeta": {
                "requiredComparisons": 2,
                "availableComparisons": 0,
                "currentPassCount": 0,
                "previousPassCount": 0,
                "qualifiedCount": 0,
                "baselineStamp": None,
                "previousBaselineStamp": None,
            },
        }
    else:
        comparison = build_comparison(
            today_rows=deduped_rows,
            today_subdomains=subdomains,
            baseline_rows=baseline_rows,
            baseline_subdomains=baseline_subs,
            baseline_stamp=((baseline_snapshot.get("meta") or {}).get("stamp") if baseline_snapshot else None),
            prev_baseline_subdomains=prev_baseline_subs,
            prev_baseline_stamp=((prev_baseline_snapshot.get("meta") or {}).get("stamp") if prev_baseline_snapshot else None),
        )

    snapshot = build_snapshot(
        stamp=stamp,
        args=args,
        rows=deduped_rows,
        subdomains=subdomains,
        comparison=comparison,
        baseline_mode=baseline_mode,
        fetch_meta=fetch_meta,
        normalize_stats=normalize_stats,
        report_history_path=report_history_path,
        excel_history_path=excel_history_path,
        latest_report_path=latest_report_path,
        latest_xlsx_path=latest_xlsx_path,
    )

    fetch_archive_payload = {
        "meta": {
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "stamp": stamp,
            "target": target,
            "request": fetch_meta,
        },
        "pages": page_results,
    }

    artifact_paths = write_artifacts(
        stamp=stamp,
        snapshot=snapshot,
        fetch_archive=fetch_archive_payload,
        data_dir=data_dir,
        snapshot_dir=snapshot_dir,
        report_dir=report_dir,
        latest_report_path=latest_report_path,
        latest_xlsx_path=latest_xlsx_path,
    )

    words_dir = Path(args.words_dir).resolve()
    chain_work_dir = Path(args.chain_work_dir).resolve()
    keyword_row_count = comparison.get("standardWordSummary", {}).get("rowCount", 0)

    chain_stage_outputs: dict = {}
    if keyword_row_count > 0:
        chain_root = chain_work_dir / stamp
        ensure_dir(chain_root)

        analyze_outputs = _run_analyze_words_stage(
            seed_table_xlsx=Path(artifact_paths["excelHistoryPath"]),
            chain_root=chain_root,
        )
        check_outputs = _run_check_gefei_kd_stage(
            input_table_xlsx=Path(analyze_outputs["latestXlsxPath"]),
            chain_root=chain_root,
        )

        final_words_xlsx_path = _publish_final_words_xlsx(
            final_xlsx=Path(check_outputs["latestXlsxPath"]),
            words_dir=words_dir,
            stamp=stamp,
        )

        final_table_payload = load_json(Path(check_outputs["latestJsonPath"]))
        final_rows = final_table_payload.get("rows") if isinstance(final_table_payload, dict) else None
        if isinstance(final_rows, list):
            comparison.setdefault("standardWordSummary", {})["rowCount"] = len(final_rows)
            comparison["standardWordRows"] = final_rows

        chain_stage_outputs = {
            "status": "completed",
            "chainRoot": str(chain_root),
            "analyzeWords": {
                "latestXlsxPath": str(analyze_outputs["latestXlsxPath"]),
                "latestJsonPath": str(analyze_outputs["latestJsonPath"]),
            },
            "checkGefeiKd": {
                "latestXlsxPath": str(check_outputs["latestXlsxPath"]),
                "latestJsonPath": str(check_outputs["latestJsonPath"]),
            },
        }
    else:
        final_words_xlsx_path = _publish_final_words_xlsx(
            final_xlsx=Path(artifact_paths["excelHistoryPath"]),
            words_dir=words_dir,
            stamp=stamp,
        )
        chain_stage_outputs = {
            "status": "skipped",
            "reason": "standardWordRows=0",
        }

    report_excel_history_path = Path(artifact_paths["excelHistoryPath"])
    copyfile(final_words_xlsx_path, report_excel_history_path)
    copyfile(final_words_xlsx_path, latest_xlsx_path)

    snapshot_meta = snapshot.setdefault("meta", {})
    output_meta = snapshot_meta.setdefault("output", {})
    output_meta["excelHistoryPath"] = str(report_excel_history_path)
    output_meta["finalWordsXlsxPath"] = str(final_words_xlsx_path)
    output_meta["chainStages"] = chain_stage_outputs

    report_history_path = Path(artifact_paths["reportHistoryPath"])
    final_report_text = render_report(snapshot)
    report_history_path.write_text(final_report_text, encoding="utf-8")
    latest_report_path.write_text(final_report_text, encoding="utf-8")

    dump_json(Path(artifact_paths["snapshotPath"]), snapshot)

    print(f"[done] fetch archive : {artifact_paths['fetchArchivePath']}")
    print(f"[done] snapshot      : {artifact_paths['snapshotPath']}")
    print(f"[done] report history: {artifact_paths['reportHistoryPath']}")
    print(f"[done] report latest : {latest_report_path}")
    print(f"[done] excel history : {artifact_paths['excelHistoryPath']}")
    print(f"[done] excel latest  : {latest_xlsx_path}")
    print(f"[done] words final   : {final_words_xlsx_path}")
    print(f"[done] deduped rows  : {len(deduped_rows)}")
    print(f"[done] subdomains    : {len(subdomains)}")
    print(f"[done] keyword rows  : {comparison.get('standardWordSummary', {}).get('rowCount', 0)}")
    if baseline_mode:
        print("[done] baseline      : none (first run)")
    else:
        print(f"[done] baseline      : {baseline_path}")

    return 0


def rebuild_reports(args: argparse.Namespace) -> int:
    snapshot_dir = Path(args.snapshot_dir).resolve()
    report_dir = Path(args.report_dir).resolve()
    latest_report_path = Path(args.latest_report_path).resolve()
    latest_xlsx_path = Path(args.latest_xlsx_path).resolve()

    ensure_dir(snapshot_dir)
    ensure_dir(report_dir)
    ensure_dir(latest_report_path.parent)
    ensure_dir(latest_xlsx_path.parent)

    files = list_snapshot_files(snapshot_dir)
    if not files:
        raise SystemExit(f"未找到快照文件: {snapshot_dir}")

    latest_report_text = ""
    latest_excel_path: Optional[Path] = None
    history_for_target: Dict[str, List[dict]] = {}

    for path in files:
        snapshot = load_json(path)
        meta = snapshot.setdefault("meta", {})
        target = meta.get("target") or {}
        request_meta = meta.setdefault("request", {})
        target_key = json.dumps(
            {
                "key": target.get("key"),
                "country": target.get("country"),
                "latest": target.get("latest"),
                "sourceType": target.get("sourceType"),
                "startPage": request_meta.get("startPage", meta.get("startPage", 1)),
                "endPage": request_meta.get("endPage", meta.get("endPage", 8)),
            },
            sort_keys=True,
            ensure_ascii=False,
        )

        target_history = history_for_target.setdefault(target_key, [])
        baseline = target_history[-1] if target_history else None
        prev_baseline = target_history[-2] if len(target_history) >= 2 else None
        if baseline is None:
            comparison = {
                "baselineStamp": None,
                "reportRows": [],
                "standardWordRows": [],
                "standardWordSummary": {
                    "rowCount": 0,
                    "newPageKeywordRows": 0,
                    "newSubdomainKeywordRows": 0,
                    "risingSubdomainKeywordRows": 0,
                    "qualifiedRisingSubdomainCount": 0,
                },
                "subdomainRisingMeta": {
                    "requiredComparisons": 2,
                    "availableComparisons": 0,
                    "currentPassCount": 0,
                    "previousPassCount": 0,
                    "qualifiedCount": 0,
                    "baselineStamp": None,
                    "previousBaselineStamp": None,
                },
                "gefeiKD": {
                    "summary": {
                        "inputCount": 0,
                        "requestCount": 0,
                        "successCount": 0,
                        "successWithScoreCount": 0,
                        "missingScoreCount": 0,
                        "failedCount": 0,
                    },
                    "failures": [],
                    "api": {},
                },
            }
            meta["baselineMode"] = True
        else:
            comparison = build_comparison(
                today_rows=snapshot.get("rows", []),
                today_subdomains=snapshot.get("subdomains", []),
                baseline_rows=baseline.get("rows", []),
                baseline_subdomains=baseline.get("subdomains", []),
                baseline_stamp=(baseline.get("meta") or {}).get("stamp"),
                prev_baseline_subdomains=(prev_baseline.get("subdomains", []) if prev_baseline else None),
                prev_baseline_stamp=((prev_baseline.get("meta") or {}).get("stamp") if prev_baseline else None),
                enrich_gefei_kd=False,
            )
            meta["baselineMode"] = False

        request_meta["standardWordTableVersion"] = STANDARD_WORD_TABLE_VERSION
        snapshot["comparison"] = comparison
        report_history_path, excel_history_path, report_text = rebuild_history_artifacts(
            snapshot,
            report_dir,
            latest_report_path,
            latest_xlsx_path,
        )
        dump_json(path, snapshot)

        latest_report_text = report_text
        latest_excel_path = excel_history_path
        print(f"[rebuild] report: {report_history_path}")
        print(f"[rebuild] excel : {excel_history_path}")

        target_history.append(snapshot)
        if len(target_history) > 2:
            del target_history[0]

    latest_report_path.write_text(latest_report_text, encoding="utf-8")
    if latest_excel_path is not None:
        copyfile(latest_excel_path, latest_xlsx_path)

    print(f"[done] latest report: {latest_report_path}")
    print(f"[done] latest excel : {latest_xlsx_path}")
    return 0


def validate_report(args: argparse.Namespace) -> int:
    report_path = Path(args.report).resolve()
    xlsx_path = Path(args.xlsx).resolve()
    if not report_path.exists():
        raise SystemExit(f"报告不存在: {report_path}")
    text = report_path.read_text(encoding="utf-8")

    required_sections = [
        "## 摘要",
        "## 监控结果",
        "## 备注",
    ]

    missing = [item for item in required_sections if item not in text]
    missing += [item for item in REQUIRED_REMARKS if item not in text]
    if "- 标准词表为最终口径：种子词表经 analyze-words（SIM/SEM）与 check-gefei-kd（gefeiKD）补全后写入 report 与 words 目录" not in text:
        missing.append("- 标准词表为最终口径：种子词表经 analyze-words（SIM/SEM）与 check-gefei-kd（gefeiKD）补全后写入 report 与 words 目录")

    if (
        "无历史快照（本次仅建立基线）" not in text
        and "| subdomain | path | clicks | trend | top keywords |" not in text
        and "## 监控结果\n\n（无）" not in text
    ):
        missing.append("| subdomain | path | clicks | trend | top keywords |")

    if missing:
        print("[failed] 报告校验失败，缺少内容：")
        for item in missing:
            print(f"- {item}")
        raise SystemExit(1)

    if not xlsx_path.exists():
        raise SystemExit(f"Excel 不存在: {xlsx_path}")

    _Workbook, _Font, _Alignment, _PatternFill, _get_column_letter, load_workbook = _require_openpyxl()
    workbook = load_workbook(xlsx_path, data_only=True)
    try:
        if "keywords" not in workbook.sheetnames:
            raise SystemExit("Excel 缺少 keywords sheet")
        sheet = workbook["keywords"]

        expected_headers = get_keywords_export_headers()
        actual_headers = [sheet.cell(row=1, column=index).value for index in range(1, sheet.max_column + 1)]
        if actual_headers != expected_headers:
            print("Excel 表头不匹配：")
            max_len = max(len(expected_headers), len(actual_headers))
            for index in range(1, max_len + 1):
                expected = expected_headers[index - 1] if index <= len(expected_headers) else None
                actual = actual_headers[index - 1] if index <= len(actual_headers) else None
                if expected != actual:
                    print(f"- 第 {index} 列 expected={expected!r} actual={actual!r}")
            raise SystemExit(1)

        keyword_col = expected_headers.index(get_header_for_field("keyword")) + 1
        domain_col = expected_headers.index(get_header_for_field("correspondingDomain")) + 1
        gefei_kd_col = expected_headers.index(get_header_for_field("gefeiKD")) + 1
        score_col = get_score_column_index_from_headers(expected_headers) + 1
        numeric_fields = [
            "simWindowVolume",
            "simKd",
            "simCpc",
            "semVolume",
            "semKd",
            "semCpc",
            "gefeiKD",
        ]
        numeric_indexes = [expected_headers.index(get_header_for_field(field)) + 1 for field in numeric_fields]

        def _is_numeric_cell(value) -> bool:
            return isinstance(value, (int, float)) and not isinstance(value, bool)

        for row_index in range(2, sheet.max_row + 1):
            keyword = sheet.cell(row=row_index, column=keyword_col).value
            domain = sheet.cell(row=row_index, column=domain_col).value
            gefei_kd_value = sheet.cell(row=row_index, column=gefei_kd_col).value
            score_value = sheet.cell(row=row_index, column=score_col).value

            if keyword in (None, ""):
                raise SystemExit(f"keyword 列为空（row={row_index}）")
            if domain in (None, ""):
                raise SystemExit(f"对应域名 列为空（row={row_index}）")

            if score_value not in (None, "") and not _is_numeric_cell(score_value):
                raise SystemExit(f"score 列应为数值类型而非文本（row={row_index}）")

            if gefei_kd_value not in (None, "") and not _is_numeric_cell(gefei_kd_value):
                raise SystemExit(f"gefeiKD 列应为数值类型而非文本（row={row_index}）")

            for col_index in numeric_indexes:
                value = sheet.cell(row=row_index, column=col_index).value
                if value in (None, ""):
                    continue
                if not _is_numeric_cell(value):
                    raise SystemExit(f"数字列应为数值类型而非文本（row={row_index}, col={col_index}）")

            domain_items = [item.strip() for item in str(domain).split("|") if item.strip()]
            url_hosts: set[str] = set()
            bare_hosts: set[str] = set()
            for item in domain_items:
                parts = item.split()
                if not parts:
                    continue

                if len(parts) == 1:
                    bare_host = _normalize_host_token(parts[0])
                    if bare_host:
                        bare_hosts.add(bare_host)
                    continue

                host = parts[0].strip().lower().rstrip(".")
                url_candidate = parts[-1].strip()
                parsed_host = _extract_host_from_url_candidate(url_candidate)
                if parsed_host and host and parsed_host == host:
                    raise SystemExit(f"对应域名重复写入子域名与同 host URL（row={row_index}）")

                full_item_host = _extract_host_from_url_candidate(item)
                if full_item_host:
                    url_hosts.add(full_item_host)
                if parsed_host:
                    url_hosts.add(parsed_host)

            duplicated_hosts = sorted(bare_hosts & url_hosts)
            if duplicated_hosts:
                raise SystemExit(
                    f"对应域名存在 host 与同 host URL 重复项（row={row_index}, hosts={','.join(duplicated_hosts)}）"
                )
    finally:
        workbook.close()

    print(f"[ok] report validated: {report_path}")
    print(f"[ok] xlsx exists      : {xlsx_path}")
    print("[ok] xlsx headers / keyword / 对应域名 validated")
    return 0


def main() -> int:
    args = parse_args()
    if args.command == "run":
        return run_pipeline(args)
    if args.command == "rebuild-reports":
        return rebuild_reports(args)
    if args.command == "validate-report":
        return validate_report(args)
    raise SystemExit(f"未知命令: {args.command}")


if __name__ == "__main__":
    raise SystemExit(main())
