#!/usr/bin/env python3
"""按词根抓取 SIM / SEM 关键词并导出 Excel。

Usage:
  python3 word-from-root/_internal/scripts/word_from_root.py run --keyword "image to text"
  python3 word-from-root/_internal/scripts/word_from_root.py rebuild-reports
  python3 word-from-root/_internal/scripts/word_from_root.py validate-report
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import subprocess
import time
from datetime import datetime
from pathlib import Path
from shutil import copyfile
from typing import Dict, Iterable, List, Optional, Sequence, Tuple
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

PROJECT_DIR = Path(__file__).resolve().parents[2]
REPO_ROOT = PROJECT_DIR.parent
LOCAL_SERVICE_TOKEN_PATH = PROJECT_DIR / "local-service" / "bridge_token.txt"
LOCAL_SERVICE_GMITM_PATH = PROJECT_DIR / "local-service" / "__gmitm.txt"
LEGACY_LOCAL_SERVICE_GMITM_PATH = PROJECT_DIR / "local-service" / "gmitm.txt"

STANDARD_WORD_TABLE_VERSION = "v1"
STANDARD_WORD_TABLE_SPEC_PATH = REPO_ROOT / "standard-word-analysis" / "spec" / f"standard-word-table.{STANDARD_WORD_TABLE_VERSION}.json"

DEFAULT_API_BASE = "http://127.0.0.1:17311"
SIM_ENDPOINT = "/sim/api/KeywordGenerator/google/suggest"
SEM_KEYWORDS_ENDPOINT = "/sem/kmtgw/v2/webapi/ideas.GetKeywords"
SEM_SUMMARY_ENDPOINT = "/sem/kmtgw/v2/webapi/ideas.GetKeywordsSummary"

SIM_RANGE_FILTER = "cpc,0.1,|difficulty,1,70"
SIM_SORT = "windowVolume"
SIM_TYPE = "Broad"
SIM_WEBSOURCE = "Total"
SIM_ROWS_PER_PAGE = 100
SIM_MAX_KEYWORDS = 300

SEM_PAGE_SIZE = 100
SEM_MAX_KEYWORDS = 300
SEM_DATABASE = "us"
SEM_CURRENCY = "USD"

REQUEST_TIMEOUT_SECONDS = 180
SNAPSHOT_RE = re.compile(r"^snapshot-(\d{8}-\d{6})\.json$")
SPACE_RE = re.compile(r"\s+")
HYPHEN_SPACE_RE = re.compile(r"[-_]+")
TOP_PREVIEW_ROWS = 15
SCORE_FORMULA = "simWindowVolume * simCpc / simKd"
GROUPING_PROMPT_VERSION = "v1"
GROUPING_SCHEMA = {
    "type": "object",
    "additionalProperties": False,
    "required": ["groups"],
    "properties": {
        "groups": {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": False,
                "required": ["memberKeys", "reason"],
                "properties": {
                    "memberKeys": {
                        "type": "array",
                        "minItems": 1,
                        "items": {"type": "string", "minLength": 1},
                    },
                    "reason": {"type": "string", "minLength": 1},
                    "confidence": {"type": "number", "minimum": 0, "maximum": 1},
                },
            },
        },
    },
}

REQUIRED_SECTIONS = [
    "## 摘要",
    "## 抓取概览",
    "## 合并结果概览",
    "## Top 关键词预览",
    "## 产物路径",
    "## 备注",
]

REQUIRED_REMARKS = [
    "Excel 是完整结果，Markdown 仅做预览",
    "排序值仅基于 SIM 的 windowVolume * cpc / kd",
    "sem_only 关键词由于缺少完整 SIM 指标，score 可能为空",
    "近义合并按 source 内先分组，再做 SIM/SEM 合并；group 列记录组内全部原词",
]

# Excel 导出列由标准词表规范驱动（standard-word-analysis/spec/standard-word-table.v1.json）。

DEFAULT_COLUMN_PADDING = 4
DEFAULT_COLUMN_MIN_WIDTH = 12
DEFAULT_COLUMN_MAX_WIDTH = 72


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="按词根抓取 SIM/SEM 关键词并导出 Excel")
    subparsers = parser.add_subparsers(dest="command", required=True)

    run_parser = subparsers.add_parser("run", help="抓取关键词 + 合并 + 快照 + 报告 + Excel")
    run_parser.add_argument("--keyword", required=True, help="词根，例如 'image to text'")
    run_parser.add_argument("--api-base", default=DEFAULT_API_BASE)
    run_parser.add_argument("--token-path", default=str(LOCAL_SERVICE_TOKEN_PATH))
    run_parser.add_argument("--gmitm-path", default=str(LOCAL_SERVICE_GMITM_PATH))

    run_parser.add_argument("--country", default="999")
    run_parser.add_argument("--latest", default="28d")
    run_parser.add_argument("--sim-rows-per-page", type=int, default=SIM_ROWS_PER_PAGE)
    run_parser.add_argument("--sim-max-keywords", type=int, default=SIM_MAX_KEYWORDS)
    run_parser.add_argument("--database", default=SEM_DATABASE)
    run_parser.add_argument("--currency", default=SEM_CURRENCY)
    run_parser.add_argument("--sem-page-size", type=int, default=SEM_PAGE_SIZE)
    run_parser.add_argument("--sem-max-keywords", type=int, default=SEM_MAX_KEYWORDS)

    run_parser.add_argument("--grouping-model", default="", help="用于近义分组的 Claude CLI model（留空使用 CLI 默认模型）")
    run_parser.add_argument("--grouping-temperature", type=float, default=0.0)
    run_parser.add_argument("--grouping-timeout-seconds", type=int, default=120)
    run_parser.add_argument("--grouping-max-retries", type=int, default=2)
    run_parser.add_argument("--grouping-retry-backoff-seconds", type=float, default=1.5)
    run_parser.add_argument("--grouping-chunk-size", type=int, default=40)
    run_parser.add_argument("--grouping-max-prompt-chars", type=int, default=12000)
    run_parser.add_argument("--grouping-cache-dir", default=str(PROJECT_DIR / "_internal" / "grouping-cache"))

    run_parser.add_argument("--data-dir", default=str(PROJECT_DIR / "data"))
    run_parser.add_argument("--snapshot-dir", default=str(PROJECT_DIR / "_internal" / "snapshots"))
    run_parser.add_argument("--report-dir", default=str(PROJECT_DIR / "report" / "history"))
    run_parser.add_argument("--latest-report-path", default=str(PROJECT_DIR / "report" / "latest.md"))
    run_parser.add_argument("--latest-xlsx-path", default=str(PROJECT_DIR / "report" / "latest.xlsx"))

    rebuild_parser = subparsers.add_parser("rebuild-reports", help="根据快照重建 Markdown 和 Excel")
    rebuild_parser.add_argument("--snapshot-dir", default=str(PROJECT_DIR / "_internal" / "snapshots"))
    rebuild_parser.add_argument("--report-dir", default=str(PROJECT_DIR / "report" / "history"))
    rebuild_parser.add_argument("--latest-report-path", default=str(PROJECT_DIR / "report" / "latest.md"))
    rebuild_parser.add_argument("--latest-xlsx-path", default=str(PROJECT_DIR / "report" / "latest.xlsx"))

    validate_parser = subparsers.add_parser("validate-report", help="校验 Markdown 报告结构与 latest.xlsx")
    validate_parser.add_argument("--report", default=str(PROJECT_DIR / "report" / "latest.md"))
    validate_parser.add_argument("--xlsx", default=str(PROJECT_DIR / "report" / "latest.xlsx"))

    return parser.parse_args()


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d-%H%M%S")


def load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def dump_json(path: Path, payload: dict) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _normalize_export_column(raw: dict) -> dict:
    if not isinstance(raw, dict):
        raise RuntimeError("标准词表列配置项必须是对象")

    header = str(raw.get("header") or "").strip()
    field = str(raw.get("field") or "").strip()
    if not header or not field:
        raise RuntimeError("标准词表列配置缺少 header/field")

    column = {
        "header": header,
        "field": field,
    }

    transform = raw.get("transform")
    if transform not in (None, ""):
        column["transform"] = str(transform)

    number_format = raw.get("number_format")
    if number_format not in (None, ""):
        column["number_format"] = str(number_format)

    width_profile = raw.get("width_profile")
    if isinstance(width_profile, dict):
        normalized_width: dict = {}
        if width_profile.get("fixed") is not None:
            normalized_width["fixed"] = width_profile.get("fixed")
        if width_profile.get("min") is not None:
            normalized_width["min"] = width_profile.get("min")
        if width_profile.get("max") is not None:
            normalized_width["max"] = width_profile.get("max")
        if normalized_width:
            column["width_profile"] = normalized_width

    return column


def load_standard_word_table_spec() -> dict:
    if not STANDARD_WORD_TABLE_SPEC_PATH.exists():
        raise RuntimeError(
            "标准词表规范缺失: "
            f"{STANDARD_WORD_TABLE_SPEC_PATH}。"
            "请先补齐 standard-word-analysis/spec/standard-word-table.v1.json。"
        )

    try:
        spec = load_json(STANDARD_WORD_TABLE_SPEC_PATH)
    except Exception as exc:  # pragma: no cover - defensive path
        raise RuntimeError(f"读取标准词表规范失败: {STANDARD_WORD_TABLE_SPEC_PATH}") from exc

    if not isinstance(spec, dict):
        raise RuntimeError("标准词表规范格式错误：顶层必须是对象")

    version = str(spec.get("version") or "").strip()
    if version != STANDARD_WORD_TABLE_VERSION:
        raise RuntimeError(
            f"标准词表版本不匹配：expected={STANDARD_WORD_TABLE_VERSION} actual={version or '-'}"
        )

    raw_columns = spec.get("excelColumns")
    if not isinstance(raw_columns, list) or not raw_columns:
        raise RuntimeError("标准词表规范缺少 excelColumns 数组")

    columns = [_normalize_export_column(item) for item in raw_columns]
    fields = [str(column.get("field") or "") for column in columns]
    if len(fields) != len(set(fields)):
        raise RuntimeError("标准词表规范字段重复：excelColumns.field 必须唯一")

    score_header = str(spec.get("scoreColumnHeader") or "").strip()
    if not score_header:
        raise RuntimeError("标准词表规范缺少 scoreColumnHeader")

    score_field = str(spec.get("scoreField") or "score").strip() or "score"
    score_candidates = [column for column in columns if column.get("field") == score_field]
    if len(score_candidates) != 1:
        raise RuntimeError("标准词表规范中 scoreField 对应列必须且仅能有一列")

    if score_candidates[0].get("header") != score_header:
        raise RuntimeError("标准词表规范 scoreColumnHeader 与 scoreField 对应列表头不一致")

    score_formula = str(spec.get("scoreFormula") or "").strip() or SCORE_FORMULA
    if score_formula != SCORE_FORMULA:
        raise RuntimeError(
            "标准词表规范 scoreFormula 与 word-from-root 当前公式不一致："
            f"expected={SCORE_FORMULA} actual={score_formula}"
        )

    return {
        "version": version,
        "columns": columns,
        "scoreColumnHeader": score_header,
        "scoreField": score_field,
        "scoreFormula": score_formula,
    }


def get_standard_word_table_spec() -> dict:
    return STANDARD_WORD_TABLE_SPEC


def get_keywords_export_columns() -> List[dict]:
    return list(get_standard_word_table_spec()["columns"])


def get_keywords_export_headers() -> List[str]:
    return [column["header"] for column in get_keywords_export_columns()]


def get_score_column_index_from_headers(headers: Sequence[str]) -> int:
    score_header = get_standard_word_table_spec()["scoreColumnHeader"]
    try:
        return list(headers).index(score_header)
    except ValueError as exc:
        raise RuntimeError(f"标准词表缺少 score 列: {score_header}") from exc


STANDARD_WORD_TABLE_SPEC = load_standard_word_table_spec()


def safe_float(value) -> float:
    try:
        if value is None or value == "":
            return 0.0
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def safe_int(value) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def to_int_if_possible(v: float) -> str:
    if v is None:
        return "-"
    if abs(v - round(v)) < 1e-9:
        return str(int(round(v)))
    return f"{v:.2f}"


def to_display_number(value) -> str:
    if value in (None, ""):
        return "-"
    return to_int_if_possible(safe_float(value))


def compute_sim_score(row: dict) -> Optional[float]:
    sim_window_volume = row.get("simWindowVolume")
    sim_cpc = row.get("simCpc")
    sim_kd = row.get("simKd")
    if sim_window_volume is None or sim_cpc is None or sim_kd is None:
        return None

    window_volume = safe_float(sim_window_volume)
    cpc = safe_float(sim_cpc)
    kd = safe_float(sim_kd)
    if window_volume <= 0 or cpc <= 0 or kd <= 0:
        return None
    return round(window_volume * cpc / kd, 6)


def score_sort_key(row: dict) -> Tuple[bool, float, float, str]:
    score = compute_sim_score(row)
    return (
        score is None,
        -(score if score is not None else -1),
        -safe_float(row.get("simWindowVolume")),
        str(row.get("keyword", "")),
    )


def read_text_required(path: Path, label: str) -> str:
    if not path.exists():
        raise RuntimeError(f"{label} 不存在: {path}")
    value = path.read_text(encoding="utf-8").strip()
    if not value:
        raise RuntimeError(f"{label} 为空: {path}")
    return value


def read_gmitm(path: Path) -> str:
    if path.exists():
        return read_text_required(path, "gmitm 文件")
    if path.name == "gmitm.txt" and LEGACY_LOCAL_SERVICE_GMITM_PATH.exists():
        return read_text_required(LEGACY_LOCAL_SERVICE_GMITM_PATH, "gmitm 文件")
    raise RuntimeError(
        f"gmitm 文件不存在: {path}。请在 word-from-root/local-service/__gmitm.txt 中写入当前会话的 __gmitm 值。"
    )


def normalize_keyword_text(text: str) -> str:
    return SPACE_RE.sub(" ", str(text or "").strip().lower())


def normalize_group_text(text: str) -> str:
    normalized = normalize_keyword_text(text)
    normalized = HYPHEN_SPACE_RE.sub(" ", normalized)
    normalized = SPACE_RE.sub(" ", normalized).strip()
    return normalized


def normalize_token_for_group(token: str) -> str:
    text = str(token or "").strip().lower()
    if len(text) > 4 and text.endswith("ies"):
        return text[:-3] + "y"
    if len(text) > 4 and text.endswith("es"):
        return text[:-2]
    if len(text) > 3 and text.endswith("s"):
        return text[:-1]
    return text


def keyword_signature(keyword: str) -> str:
    tokens = [normalize_token_for_group(part) for part in normalize_group_text(keyword).split(" ") if part.strip()]
    deduped = sorted({token for token in tokens if token})
    return " ".join(deduped)


def stable_unique_texts(values: Sequence[str]) -> List[str]:
    seen: set[str] = set()
    result: List[str] = []
    for value in values:
        item = str(value or "").strip()
        if not item:
            continue
        key = normalize_keyword_text(item)
        if key in seen:
            continue
        seen.add(key)
        result.append(item)
    return result


def _source_volume_field(source: str) -> str:
    return "simWindowVolume" if source == "sim" else "semVolume"


def grouping_model_label(args: argparse.Namespace) -> str:
    model_name = str(args.grouping_model or "").strip()
    return model_name if model_name else "default"


def _weighted_average(rows: Sequence[dict], *, value_field: str, weight_field: str) -> Optional[float]:
    numerator = 0.0
    denominator = 0.0
    for row in rows:
        raw_value = row.get(value_field)
        if raw_value in (None, ""):
            continue
        value = safe_float(raw_value)
        weight = safe_float(row.get(weight_field))
        if weight <= 0:
            continue
        numerator += value * weight
        denominator += weight
    if denominator <= 0:
        return None
    return round(numerator / denominator, 6)


def _grouping_payload_row(row: dict, source: str) -> dict:
    return {
        "key": row.get("keywordNormalized"),
        "keyword": row.get("keyword"),
        "signature": keyword_signature(str(row.get("keyword") or "")),
        "volume": safe_float(row.get(_source_volume_field(source))),
    }


def build_grouping_prompt(rows: Sequence[dict], source: str) -> str:
    source_label = "SIM" if source == "sim" else "SEM"
    payload_rows = [_grouping_payload_row(row, source) for row in rows]
    payload_rows.sort(key=lambda item: str(item.get("key") or ""))

    instructions = {
        "task": f"对 {source_label} 关键词做近义分组。",
        "rules": [
            "仅在词序变化、空格/连字符差异、单复数、无意义重复时合并。",
            "严禁将语义相反词合并，例如 image to text 与 text to image。",
            "每个 key 必须且只能出现一次，不能遗漏、不能重复。",
            "仅输出符合 schema 的 JSON：{groups:[{memberKeys:[], reason:'', confidence:0~1}]}",
        ],
        "rows": payload_rows,
    }

    return "你是关键词近义聚类专家。仅返回满足 schema 的 JSON。\n" + json.dumps(instructions, ensure_ascii=False)


def _grouping_cache_file(cache_dir: Path, source: str, rows: Sequence[dict], args: argparse.Namespace, chunk_label: str = "all") -> Path:
    serializable_rows = [_grouping_payload_row(row, source) for row in rows]
    serializable_rows.sort(key=lambda item: str(item.get("key") or ""))
    payload = {
        "source": source,
        "chunkLabel": chunk_label,
        "promptVersion": GROUPING_PROMPT_VERSION,
        "model": grouping_model_label(args),
        "temperature": args.grouping_temperature,
        "rows": serializable_rows,
    }
    cache_key = hashlib.sha256(json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest()
    return cache_dir / f"grouping-{source}-{chunk_label}-{cache_key}.json"


def _parse_structured_grouping_output(stdout_text: str) -> dict:
    text = str(stdout_text or "").strip()
    if not text:
        raise RuntimeError("AI 分组返回为空")

    parsed = json.loads(text)
    if isinstance(parsed, dict) and isinstance(parsed.get("structured_output"), dict):
        return parsed["structured_output"]
    if isinstance(parsed, dict) and isinstance(parsed.get("result"), str):
        result_text = parsed.get("result") or ""
        try:
            decoded = json.loads(result_text)
            if isinstance(decoded, dict):
                return decoded
        except json.JSONDecodeError:
            pass
    if isinstance(parsed, dict):
        return parsed
    raise RuntimeError("AI 分组输出无法解析为 JSON 对象")


def call_grouping_model(prompt: str, args: argparse.Namespace) -> dict:
    command = [
        "claude",
        "-p",
        "--output-format",
        "json",
        "--json-schema",
        json.dumps(GROUPING_SCHEMA, ensure_ascii=False),
        "--settings",
        json.dumps({"modelParameters": {"temperature": float(args.grouping_temperature)}}),
    ]
    model_name = str(args.grouping_model or "").strip()
    if model_name:
        command.extend(["--model", model_name])
    command.append(prompt)
    try:
        result = subprocess.run(
            command,
            capture_output=True,
            text=True,
            timeout=max(int(args.grouping_timeout_seconds), 30),
            check=False,
        )
    except subprocess.TimeoutExpired as exc:
        raise RuntimeError(f"AI 分组调用超时（{int(args.grouping_timeout_seconds)}s）") from exc

    if result.returncode != 0:
        detail = (result.stderr or result.stdout or "").strip()
        raise RuntimeError(f"AI 分组调用失败: {detail}")
    return _parse_structured_grouping_output(result.stdout)


def call_grouping_model_with_retry(prompt: str, args: argparse.Namespace) -> Tuple[dict, dict]:
    max_retries = max(int(args.grouping_max_retries), 0)
    base_backoff = max(float(args.grouping_retry_backoff_seconds), 0.1)
    attempts = 0
    retries = 0
    last_error = ""

    for attempt_index in range(max_retries + 1):
        attempts += 1
        try:
            result = call_grouping_model(prompt, args)
            return result, {"attempts": attempts, "retries": retries, "error": ""}
        except RuntimeError as exc:
            last_error = str(exc)
            if attempt_index >= max_retries:
                break
            retries += 1
            sleep_seconds = base_backoff * (2 ** attempt_index)
            time.sleep(sleep_seconds)

    raise RuntimeError(f"AI 分组重试后仍失败（attempts={attempts}）: {last_error}")


def validate_grouping_result(rows: Sequence[dict], grouping_result: dict) -> List[dict]:
    groups = grouping_result.get("groups") if isinstance(grouping_result, dict) else None
    if not isinstance(groups, list):
        raise RuntimeError("AI 分组结果缺少 groups 数组")

    available_keys = {str(row.get("keywordNormalized") or "") for row in rows}
    available_keys = {item for item in available_keys if item}
    seen: set[str] = set()
    normalized_groups: List[dict] = []

    for group in groups:
        if not isinstance(group, dict):
            raise RuntimeError("AI 分组结果中存在非法 group")
        members = group.get("memberKeys")
        if not isinstance(members, list) or not members:
            raise RuntimeError("AI 分组结果存在空 memberKeys")
        normalized_members: List[str] = []
        for member in members:
            key = normalize_keyword_text(str(member or ""))
            if not key:
                raise RuntimeError("AI 分组结果出现空 key")
            if key not in available_keys:
                raise RuntimeError(f"AI 分组结果包含未知 key: {key}")
            if key in seen:
                raise RuntimeError(f"AI 分组结果存在重复 key: {key}")
            seen.add(key)
            normalized_members.append(key)

        reason = str(group.get("reason") or "").strip()
        if not reason:
            raise RuntimeError("AI 分组结果缺少 reason")

        confidence = group.get("confidence")
        if confidence in (None, ""):
            normalized_confidence = None
        else:
            normalized_confidence = max(0.0, min(1.0, safe_float(confidence)))

        normalized_groups.append(
            {
                "memberKeys": sorted(normalized_members),
                "reason": reason,
                "confidence": normalized_confidence,
            }
        )

    missing = sorted(available_keys - seen)
    if missing:
        raise RuntimeError(f"AI 分组结果遗漏 key: {', '.join(missing[:8])}")
    return normalized_groups


def aggregate_group_rows(rows: Sequence[dict], *, source: str, reason: str, confidence: Optional[float]) -> dict:
    if not rows:
        raise RuntimeError("分组聚合收到空 rows")

    source_volume_field = _source_volume_field(source)
    sorted_rows = sorted(
        rows,
        key=lambda row: (
            -safe_float(row.get(source_volume_field)),
            str(row.get("keyword") or ""),
        ),
    )
    canonical_row = dict(sorted_rows[0])

    keywords = stable_unique_texts([str(item.get("keyword") or "") for item in sorted_rows])
    group_members = keywords
    group_text = " | ".join(group_members)

    merge_key = min(
        (
            keyword_signature(str(item.get("keyword") or ""))
            for item in sorted_rows
            if str(item.get("keyword") or "").strip()
        ),
        default="",
    )

    grouped = dict(canonical_row)
    grouped["keyword"] = canonical_row.get("keyword")
    grouped["keywordNormalized"] = str(canonical_row.get("keywordNormalized") or normalize_keyword_text(grouped.get("keyword") or ""))
    grouped["mergeKey"] = merge_key or grouped["keywordNormalized"]
    grouped["groupMembers"] = group_members
    grouped["group"] = group_text
    grouped["groupReason"] = reason
    grouped["groupConfidence"] = confidence
    grouped["groupSize"] = len(group_members)

    group_volume = sum(max(0.0, safe_float(item.get(source_volume_field))) for item in sorted_rows)

    if source == "sim":
        grouped["simKeyword"] = grouped["keyword"]
        grouped["simWindowVolume"] = round(group_volume, 6)
        grouped["simAverageVolume"] = _weighted_average(sorted_rows, value_field="simAverageVolume", weight_field="simWindowVolume")
        grouped["simCpc"] = _weighted_average(sorted_rows, value_field="simCpc", weight_field="simWindowVolume")
        grouped["simKd"] = _weighted_average(sorted_rows, value_field="simKd", weight_field="simWindowVolume")
        grouped["simGroupedCount"] = len(sorted_rows)
    else:
        grouped["semKeyword"] = grouped["keyword"]
        grouped["semVolume"] = round(group_volume, 6)
        grouped["semCpc"] = _weighted_average(sorted_rows, value_field="semCpc", weight_field="semVolume")
        grouped["semKd"] = _weighted_average(sorted_rows, value_field="semKd", weight_field="semVolume")
        grouped["semGroupedCount"] = len(sorted_rows)

    return grouped


def plan_grouping_chunks(rows: Sequence[dict], source: str, args: argparse.Namespace) -> List[dict]:
    if not rows:
        return []

    chunk_size = max(int(args.grouping_chunk_size), 1)
    max_prompt_chars = max(int(args.grouping_max_prompt_chars), 1200)
    bucket_count = max(1, math.ceil(len(rows) / chunk_size))
    buckets: Dict[int, List[dict]] = {idx: [] for idx in range(bucket_count)}

    for row in rows:
        signature = keyword_signature(str(row.get("keyword") or ""))
        digest = hashlib.sha256(signature.encode("utf-8")).hexdigest()
        bucket_id = int(digest, 16) % bucket_count
        buckets[bucket_id].append(row)

    chunks: List[dict] = []

    def _append_chunk(bucket_id: int, part_id: int, part_rows: List[dict]) -> None:
        sorted_rows = sorted(part_rows, key=lambda item: str(item.get("keywordNormalized") or item.get("keyword") or ""))
        prompt_text = build_grouping_prompt(sorted_rows, source)
        chunk_id = f"b{bucket_id:03d}-p{part_id:03d}"
        chunks.append(
            {
                "id": chunk_id,
                "rows": sorted_rows,
                "inputCount": len(sorted_rows),
                "promptChars": len(prompt_text),
            }
        )

    for bucket_id in sorted(buckets.keys()):
        bucket_rows = sorted(buckets[bucket_id], key=lambda item: str(item.get("keywordNormalized") or item.get("keyword") or ""))
        if not bucket_rows:
            continue

        part_id = 1
        current_rows: List[dict] = []
        for row in bucket_rows:
            candidate_rows = current_rows + [row]
            candidate_prompt_chars = len(build_grouping_prompt(candidate_rows, source))
            if current_rows and (len(candidate_rows) > chunk_size or candidate_prompt_chars > max_prompt_chars):
                _append_chunk(bucket_id, part_id, current_rows)
                part_id += 1
                current_rows = [row]
                continue
            current_rows = candidate_rows

        if current_rows:
            _append_chunk(bucket_id, part_id, current_rows)

    chunks.sort(key=lambda item: str(item.get("id") or ""))
    return chunks


def group_source_rows_with_ai(rows: List[dict], *, source: str, args: argparse.Namespace) -> Tuple[List[dict], dict]:
    if not rows:
        return [], {
            "source": source,
            "promptVersion": GROUPING_PROMPT_VERSION,
            "model": grouping_model_label(args),
            "temperature": args.grouping_temperature,
            "groupCount": 0,
            "inputCount": 0,
            "chunkCount": 0,
            "chunkCacheHits": 0,
            "chunkCacheMisses": 0,
            "totalRetries": 0,
        }

    cache_dir = Path(args.grouping_cache_dir).resolve()
    ensure_dir(cache_dir)

    chunks = plan_grouping_chunks(rows, source, args)
    all_input_keys = {str(row.get("keywordNormalized") or "") for row in rows if str(row.get("keywordNormalized") or "").strip()}
    seen_keys: set[str] = set()
    grouped_rows: List[dict] = []

    chunk_cache_hits = 0
    chunk_cache_misses = 0
    total_retries = 0
    chunk_stats: List[dict] = []

    for chunk in chunks:
        chunk_rows = chunk["rows"]
        chunk_id = str(chunk["id"])
        cache_file = _grouping_cache_file(cache_dir, source, chunk_rows, args, chunk_label=chunk_id)

        cache_hit = False
        attempts = 0
        retries = 0
        start_at = time.monotonic()

        if cache_file.exists():
            try:
                grouping_result = load_json(cache_file)
                cache_hit = True
                chunk_cache_hits += 1
            except Exception:
                grouping_result = None
        else:
            grouping_result = None

        if grouping_result is None:
            chunk_cache_misses += 1
            prompt = build_grouping_prompt(chunk_rows, source)
            grouping_result, retry_meta = call_grouping_model_with_retry(prompt, args)
            attempts = int(retry_meta.get("attempts") or 0)
            retries = int(retry_meta.get("retries") or 0)
            total_retries += retries
            dump_json(cache_file, grouping_result)

        normalized_groups = validate_grouping_result(chunk_rows, grouping_result)
        key_to_row = {str(row.get("keywordNormalized") or ""): row for row in chunk_rows}

        for group in normalized_groups:
            member_keys = [key for key in group["memberKeys"] if key in key_to_row]
            duplicate_keys = [key for key in member_keys if key in seen_keys]
            if duplicate_keys:
                raise RuntimeError(f"分块分组结果重复 key（source={source}, chunk={chunk_id}）: {', '.join(duplicate_keys[:8])}")
            for key in member_keys:
                seen_keys.add(key)
            member_rows = [key_to_row[key] for key in member_keys]
            if not member_rows:
                continue
            grouped_rows.append(
                aggregate_group_rows(
                    member_rows,
                    source=source,
                    reason=group["reason"],
                    confidence=group.get("confidence"),
                )
            )

        elapsed_ms = int(round((time.monotonic() - start_at) * 1000))
        chunk_stats.append(
            {
                "chunkId": chunk_id,
                "inputCount": len(chunk_rows),
                "groupCount": len(normalized_groups),
                "promptChars": chunk.get("promptChars"),
                "cacheHit": cache_hit,
                "attempts": attempts,
                "retries": retries,
                "elapsedMs": elapsed_ms,
            }
        )

    missing_keys = sorted(all_input_keys - seen_keys)
    if missing_keys:
        raise RuntimeError(f"分块分组结果遗漏 key（source={source}）: {', '.join(missing_keys[:8])}")

    grouped_rows.sort(key=score_sort_key)
    grouping_meta = {
        "source": source,
        "promptVersion": GROUPING_PROMPT_VERSION,
        "model": grouping_model_label(args),
        "temperature": args.grouping_temperature,
        "groupCount": len(grouped_rows),
        "inputCount": len(rows),
        "chunkCount": len(chunks),
        "chunkCacheHits": chunk_cache_hits,
        "chunkCacheMisses": chunk_cache_misses,
        "totalRetries": total_retries,
        "cacheHit": bool(chunks) and chunk_cache_hits == len(chunks),
        "chunkStats": chunk_stats,
    }
    return grouped_rows, grouping_meta


def merge_group_members(left: Sequence[str], right: Sequence[str]) -> Tuple[List[str], str]:
    merged = stable_unique_texts(list(left) + list(right))
    return merged, " | ".join(merged)


def list_snapshot_files(snapshot_dir: Path) -> List[Path]:
    if not snapshot_dir.exists():
        return []
    files = []
    for path in snapshot_dir.iterdir():
        if path.is_file() and SNAPSHOT_RE.match(path.name):
            files.append(path)
    files.sort(key=lambda p: p.name)
    return files


def _post_json(url: str, headers: dict, payload: dict, timeout_seconds: int) -> Tuple[int, str]:
    req = Request(
        url,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers=headers,
        method="POST",
    )
    try:
        with urlopen(req, timeout=timeout_seconds) as resp:
            body = resp.read().decode("utf-8", errors="replace")
            return int(getattr(resp, "status", 0) or 0), body
    except HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace") if hasattr(exc, "read") else ""
        return int(getattr(exc, "code", 0) or 0), body


def post_local_service(*, api_url: str, headers: dict, payload: dict, expect_jsonrpc: bool = False) -> dict:
    try:
        status_code, raw_body = _post_json(
            api_url,
            headers=headers,
            payload=payload,
            timeout_seconds=REQUEST_TIMEOUT_SECONDS,
        )
    except URLError as exc:
        raise RuntimeError(f"请求本地服务失败: {exc}") from exc

    try:
        wrapper = json.loads(raw_body)
    except json.JSONDecodeError as exc:
        raise RuntimeError("本地服务返回非 JSON") from exc

    if not isinstance(wrapper, dict):
        raise RuntimeError("本地服务返回非 JSON 对象")
    if not wrapper.get("ok"):
        err = wrapper.get("error") or {}
        raise RuntimeError(f"本地服务错误: {err.get('code', 'UNKNOWN')} {err.get('message', '')}".strip())

    data = wrapper.get("data") or {}
    upstream_status = int(data.get("status", status_code) or 0)
    if upstream_status != 200:
        raise RuntimeError(f"上游状态码非 200: {upstream_status}")

    upstream_body_raw = data.get("body")
    if not isinstance(upstream_body_raw, str):
        raise RuntimeError("上游 body 缺失或类型错误")

    try:
        upstream = json.loads(upstream_body_raw)
    except json.JSONDecodeError as exc:
        raise RuntimeError("上游 body 非合法 JSON") from exc

    if expect_jsonrpc and not isinstance(upstream, dict):
        raise RuntimeError("上游 JSON-RPC 响应不是对象")

    return {
        "wrapper": wrapper,
        "upstream": upstream,
        "status": upstream_status,
    }


def build_headers(token: str) -> dict:
    return {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }


def sim_request_payload(args: argparse.Namespace, page: int) -> dict:
    return {
        "keyword": args.keyword,
        "country": args.country,
        "latest": args.latest,
        "isWindow": True,
        "websource": SIM_WEBSOURCE,
        "webSource": SIM_WEBSOURCE,
        "sort": SIM_SORT,
        "asc": False,
        "rangeFilter": SIM_RANGE_FILTER,
        "rowsPerPage": args.sim_rows_per_page,
        "type": SIM_TYPE,
        "page": page,
    }


def build_sem_filter() -> dict:
    return {
        "phrase": [],
        "competition_level": [],
        "cpc": [
            {
                "inverted": False,
                "operation": 5,
                "value": 0.1,
            }
        ],
        "difficulty": [
            {
                "inverted": False,
                "operation": 4,
                "value": 60,
            }
        ],
        "results": [],
        "serp_features": [
            {
                "inverted": False,
                "value": [],
            }
        ],
        "volume": [],
        "words_count": [],
        "phrase_include_logic": 0,
    }


def sem_request_body(args: argparse.Namespace, method: str, page_number: Optional[int] = None) -> dict:
    params = {
        "mode": 0,
        "currency": args.currency,
        "database": args.database,
        "filter": build_sem_filter(),
        "groups": [],
        "order": {
            "direction": 1,
            "field": "volume",
        },
        "groups_order": {
            "direction": 1,
            "field": "count",
        },
        "phrase": args.keyword,
        "questions_only": False,
    }
    if page_number is not None:
        params["page"] = {
            "number": page_number,
            "size": args.sem_page_size,
        }

    request_id = 15 if method == "ideas.GetKeywordsSummary" else 14 + int(page_number or 0)
    return {
        "id": request_id,
        "jsonrpc": "2.0",
        "method": method,
        "params": params,
    }


def fetch_sim_keywords(args: argparse.Namespace, headers: dict) -> Tuple[List[dict], dict]:
    api_url = args.api_base.rstrip("/") + SIM_ENDPOINT
    first_page = post_local_service(
        api_url=api_url,
        headers=headers,
        payload=sim_request_payload(args, page=1),
    )
    upstream = first_page["upstream"]
    records = upstream.get("records")
    if not isinstance(records, list):
        raise RuntimeError("SIM 响应 records 非数组")

    total_records = safe_int(upstream.get("totalRecords"))
    effective_total = min(total_records, args.sim_max_keywords)
    total_pages = max(1, math.ceil(effective_total / max(args.sim_rows_per_page, 1)))

    page_results = [
        {
            "page": 1,
            "requestPayload": sim_request_payload(args, page=1),
            "wrapper": first_page["wrapper"],
            "upstream": upstream,
            "rowsCount": len(records),
        }
    ]

    for page in range(2, total_pages + 1):
        result = post_local_service(
            api_url=api_url,
            headers=headers,
            payload=sim_request_payload(args, page=page),
        )
        page_upstream = result["upstream"]
        page_records = page_upstream.get("records")
        if not isinstance(page_records, list):
            raise RuntimeError(f"SIM 第 {page} 页 records 非数组")
        page_results.append(
            {
                "page": page,
                "requestPayload": sim_request_payload(args, page=page),
                "wrapper": result["wrapper"],
                "upstream": page_upstream,
                "rowsCount": len(page_records),
            }
        )

    normalized = normalize_sim_records(page_results, args.sim_max_keywords)
    meta = {
        "apiUrl": api_url,
        "totalRecords": total_records,
        "effectiveTotal": effective_total,
        "rowsPerPage": args.sim_rows_per_page,
        "pagesFetched": len(page_results),
        "pagesRequested": [item["page"] for item in page_results],
        "rangeFilter": SIM_RANGE_FILTER,
        "sort": SIM_SORT,
    }
    return normalized, {
        "meta": meta,
        "pages": page_results,
    }


def normalize_sim_records(page_results: List[dict], max_keywords: int) -> List[dict]:
    normalized: List[dict] = []
    seen: set[str] = set()

    for page_result in page_results:
        page_num = safe_int(page_result.get("page"))
        records = (page_result.get("upstream") or {}).get("records") or []
        for idx, item in enumerate(records, start=1):
            if not isinstance(item, dict):
                continue
            keyword = str(item.get("keyword") or "").strip()
            if not keyword:
                continue
            normalized_key = normalize_keyword_text(keyword)
            if normalized_key in seen:
                continue
            seen.add(normalized_key)
            normalized.append(
                {
                    "keyword": keyword,
                    "keywordNormalized": normalized_key,
                    "simKeyword": keyword,
                    "simWindowVolume": safe_float(item.get("windowVolume")),
                    "simAverageVolume": safe_float(item.get("averageVolume")),
                    "simCpc": safe_float(item.get("cpc")),
                    "simKd": safe_float(item.get("difficulty")),
                    "simRank": len(normalized) + 1,
                    "simPage": page_num,
                    "simRankInPage": idx,
                }
            )
            if len(normalized) >= max_keywords:
                return normalized

    return normalized


def fetch_sem_keywords(args: argparse.Namespace, headers: dict, gmitm: str) -> Tuple[List[dict], dict]:
    summary_url = args.api_base.rstrip("/") + SEM_SUMMARY_ENDPOINT
    keywords_url = args.api_base.rstrip("/") + SEM_KEYWORDS_ENDPOINT

    summary_payload = {
        "__gmitm": gmitm,
        "requestBody": sem_request_body(args, "ideas.GetKeywordsSummary"),
    }
    summary_result = post_local_service(
        api_url=summary_url,
        headers=headers,
        payload=summary_payload,
        expect_jsonrpc=True,
    )
    summary_upstream = summary_result["upstream"]
    summary_data = (summary_upstream.get("result") or {}) if isinstance(summary_upstream, dict) else {}
    total = safe_int(summary_data.get("total"))
    effective_total = min(total, args.sem_max_keywords)
    total_pages = max(1, math.ceil(effective_total / max(args.sem_page_size, 1))) if effective_total else 1

    page_results: List[dict] = []
    normalized: List[dict] = []
    seen: set[str] = set()

    for page in range(1, total_pages + 1):
        page_request_body = sem_request_body(args, "ideas.GetKeywords", page_number=page)
        page_payload = {
            "__gmitm": gmitm,
            "requestBody": page_request_body,
        }
        page_result = post_local_service(
            api_url=keywords_url,
            headers=headers,
            payload=page_payload,
            expect_jsonrpc=True,
        )
        page_upstream = page_result["upstream"]
        keywords = (((page_upstream.get("result") or {}) if isinstance(page_upstream, dict) else {}).get("keywords") or [])
        if not isinstance(keywords, list):
            raise RuntimeError(f"SEM 第 {page} 页 keywords 非数组")

        page_results.append(
            {
                "page": page,
                "requestPayload": page_payload,
                "wrapper": page_result["wrapper"],
                "upstream": page_upstream,
                "rowsCount": len(keywords),
            }
        )

        for idx, item in enumerate(keywords, start=1):
            if not isinstance(item, dict):
                continue
            keyword = str(item.get("phrase") or "").strip()
            if not keyword:
                continue
            normalized_key = normalize_keyword_text(keyword)
            if normalized_key in seen:
                continue
            seen.add(normalized_key)
            normalized.append(
                {
                    "keyword": keyword,
                    "keywordNormalized": normalized_key,
                    "semKeyword": keyword,
                    "semVolume": safe_float(item.get("volume")),
                    "semCpc": safe_float(item.get("cpc")),
                    "semKd": safe_float(item.get("difficulty")),
                    "semCompetitionLevel": safe_float(item.get("competition_level")),
                    "semResults": safe_int(item.get("results")),
                    "semSnapshotDate": str(item.get("snapshot_date") or ""),
                    "semTrend": item.get("trend") if isinstance(item.get("trend"), list) else [],
                    "semRank": len(normalized) + 1,
                    "semPage": page,
                    "semRankInPage": idx,
                }
            )
            if len(normalized) >= args.sem_max_keywords:
                break
        if len(normalized) >= args.sem_max_keywords:
            break

    meta = {
        "summaryApiUrl": summary_url,
        "keywordsApiUrl": keywords_url,
        "totalRecords": total,
        "effectiveTotal": effective_total,
        "pageSize": args.sem_page_size,
        "pagesFetched": len(page_results),
        "pagesRequested": [item["page"] for item in page_results],
        "summary": {
            "total": total,
            "totalVolume": safe_int(summary_data.get("total_volume")),
            "totalKeywordsWithDifficulty": safe_int(summary_data.get("total_keywords_with_difficulty")),
            "totalDifficulty": safe_int(summary_data.get("total_difficulty")),
        },
    }
    return normalized, {
        "meta": meta,
        "summary": {
            "requestPayload": summary_payload,
            "wrapper": summary_result["wrapper"],
            "upstream": summary_upstream,
        },
        "pages": page_results,
    }


def merge_rows(sim_rows: List[dict], sem_rows: List[dict]) -> List[dict]:
    merged: Dict[str, dict] = {}

    for row in sim_rows:
        key = str(row.get("mergeKey") or row.get("keywordNormalized") or "")
        if not key:
            key = normalize_keyword_text(str(row.get("keyword") or ""))
        item = merged.setdefault(
            key,
            {
                "keyword": row.get("keyword"),
                "keywordNormalized": row.get("keywordNormalized") or key,
                "mergeKey": key,
                "sourcePresence": "sim_only",
                "score": None,
                "groupMembers": list(row.get("groupMembers") or []),
                "group": str(row.get("group") or ""),
            },
        )
        item.update({k: v for k, v in row.items() if k not in ("keyword", "keywordNormalized")})

    for row in sem_rows:
        key = str(row.get("mergeKey") or row.get("keywordNormalized") or "")
        if not key:
            key = normalize_keyword_text(str(row.get("keyword") or ""))
        item = merged.setdefault(
            key,
            {
                "keyword": row.get("keyword"),
                "keywordNormalized": row.get("keywordNormalized") or key,
                "mergeKey": key,
                "sourcePresence": "sem_only",
                "score": None,
                "groupMembers": list(row.get("groupMembers") or []),
                "group": str(row.get("group") or ""),
            },
        )
        if item.get("sourcePresence") == "sim_only":
            item["sourcePresence"] = "both"

        left_members = item.get("groupMembers") or []
        right_members = row.get("groupMembers") or []
        merged_members, merged_group_text = merge_group_members(left_members, right_members)
        item["groupMembers"] = merged_members
        item["group"] = merged_group_text

        sim_volume = safe_float(item.get("simWindowVolume"))
        sem_volume = safe_float(row.get("semVolume"))
        current_keyword = str(item.get("keyword") or "")
        incoming_keyword = str(row.get("keyword") or "")
        if (sem_volume > sim_volume) or (abs(sem_volume - sim_volume) < 1e-9 and incoming_keyword and incoming_keyword < current_keyword):
            item["keyword"] = incoming_keyword
            item["keywordNormalized"] = row.get("keywordNormalized") or item.get("keywordNormalized")

        item.update({k: v for k, v in row.items() if k not in ("keyword", "keywordNormalized", "groupMembers", "group")})

    rows = list(merged.values())
    for row in rows:
        if not row.get("group"):
            members = row.get("groupMembers") or []
            row["group"] = " | ".join(stable_unique_texts(members))
        row["score"] = compute_sim_score(row)

    rows.sort(key=score_sort_key)
    return rows


def build_summary_counts(*, sim_raw_rows: List[dict], sem_raw_rows: List[dict], sim_grouped_rows: List[dict], sem_grouped_rows: List[dict], merged_rows: List[dict]) -> dict:
    counts = {
        "simRawCount": len(sim_raw_rows),
        "semRawCount": len(sem_raw_rows),
        "simCount": len(sim_grouped_rows),
        "semCount": len(sem_grouped_rows),
        "simGroupedCount": len(sim_grouped_rows),
        "semGroupedCount": len(sem_grouped_rows),
        "mergedCount": len(merged_rows),
        "bothCount": 0,
        "simOnlyCount": 0,
        "semOnlyCount": 0,
        "scoredCount": 0,
    }
    for row in merged_rows:
        source_presence = row.get("sourcePresence")
        if source_presence == "both":
            counts["bothCount"] += 1
        elif source_presence == "sim_only":
            counts["simOnlyCount"] += 1
        elif source_presence == "sem_only":
            counts["semOnlyCount"] += 1
        if row.get("score") is not None:
            counts["scoredCount"] += 1
    return counts


def _md_table(headers: Sequence[str], rows: Sequence[Sequence[str]]) -> List[str]:
    if not rows:
        return ["（无）"]
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(row) + " |")
    return lines


def build_snapshot(*, stamp: str, args: argparse.Namespace, sim_fetch: dict, sem_fetch: dict, sim_raw_rows: List[dict], sem_raw_rows: List[dict], sim_rows: List[dict], sem_rows: List[dict], merged_rows: List[dict], summary_counts: dict, grouping_meta: dict, excel_history_path: Path, report_history_path: Path) -> dict:
    return {
        "meta": {
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "stamp": stamp,
            "target": {
                "keyword": args.keyword,
                "country": args.country,
                "latest": args.latest,
                "database": args.database,
                "currency": args.currency,
            },
            "request": {
                "simRowsPerPage": args.sim_rows_per_page,
                "simMaxKeywords": args.sim_max_keywords,
                "semPageSize": args.sem_page_size,
                "semMaxKeywords": args.sem_max_keywords,
                "simRangeFilter": SIM_RANGE_FILTER,
                "simSort": SIM_SORT,
                "scoreFormula": SCORE_FORMULA,
                "standardWordTableVersion": STANDARD_WORD_TABLE_VERSION,
                "grouping": {
                    "enabled": True,
                    "promptVersion": GROUPING_PROMPT_VERSION,
                    "model": grouping_model_label(args),
                    "temperature": args.grouping_temperature,
                    "timeoutSeconds": args.grouping_timeout_seconds,
                    "maxRetries": args.grouping_max_retries,
                    "retryBackoffSeconds": args.grouping_retry_backoff_seconds,
                    "chunkSize": args.grouping_chunk_size,
                    "maxPromptChars": args.grouping_max_prompt_chars,
                    "cacheDir": str(Path(args.grouping_cache_dir).resolve()),
                },
            },
            "api": {
                "baseUrl": args.api_base,
                "sim": sim_fetch["meta"],
                "sem": sem_fetch["meta"],
            },
            "grouping": grouping_meta,
            "output": {
                "reportHistoryPath": str(report_history_path),
                "excelHistoryPath": str(excel_history_path),
            },
            "summary": summary_counts,
        },
        "sim": {
            "rawRows": sim_raw_rows,
            "rows": sim_rows,
            "raw": sim_fetch,
        },
        "sem": {
            "rawRows": sem_raw_rows,
            "rows": sem_rows,
            "raw": sem_fetch,
        },
        "mergedRows": merged_rows,
    }


def render_report(snapshot: dict) -> str:
    meta = snapshot.get("meta") or {}
    target = meta.get("target") or {}
    api = meta.get("api") or {}
    summary = meta.get("summary") or {}
    output = meta.get("output") or {}
    merged_rows = snapshot.get("mergedRows") or []

    sim_meta = api.get("sim") or {}
    sem_meta = api.get("sem") or {}
    score_formula = (meta.get("request") or {}).get("scoreFormula", SCORE_FORMULA)

    preview_rows = []
    for row in merged_rows[:TOP_PREVIEW_ROWS]:
        preview_rows.append(
            [
                str(row.get("keyword") or "-"),
                str(row.get("group") or "-")[:120],
                str(_transform_export_value({"transform": "source_presence_label"}, row.get("sourcePresence"), row) or "-"),
                to_display_number(compute_sim_score(row)) if compute_sim_score(row) is not None else "-",
                to_display_number(row.get("simWindowVolume")),
                to_display_number(row.get("simKd")),
                to_display_number(row.get("simCpc")),
                to_display_number(row.get("semVolume")),
                to_display_number(row.get("semKd")),
                to_display_number(row.get("semCpc")),
            ]
        )

    lines: List[str] = []
    lines.append(f"# 词根扩词报告（{target.get('keyword', '-')}｜{meta.get('stamp', '-')}）")
    lines.append("")

    lines.append("## 摘要")
    lines.append("")
    lines.append(f"- 关键词总数：{summary.get('mergedCount', 0)}（可计算 score：{summary.get('scoredCount', 0)}）")
    lines.append(
        f"- 分组后来源词数：SIM={summary.get('simGroupedCount', summary.get('simCount', 0))} / SEM={summary.get('semGroupedCount', summary.get('semCount', 0))}"
    )
    lines.append(
        f"- 来源分布：both={summary.get('bothCount', 0)} / sim_only={summary.get('simOnlyCount', 0)} / sem_only={summary.get('semOnlyCount', 0)}"
    )
    lines.append(f"- 排序公式：{score_formula}")
    lines.append("- 来源说明：SIM = Similarweb Keyword Generator；SEM = Semrush Keyword Magic")
    lines.append("")

    lines.append("## 抓取概览")
    lines.append("")
    lines.append(f"- SIM：totalRecords={sim_meta.get('totalRecords', 0)}，raw={summary.get('simRawCount', 0)}，grouped={summary.get('simCount', 0)}")
    lines.append(f"- SEM：totalRecords={sem_meta.get('totalRecords', 0)}，raw={summary.get('semRawCount', 0)}，grouped={summary.get('semCount', 0)}")
    lines.append("")

    lines.append("## 合并结果概览")
    lines.append("")
    lines.append("- 近义合并按 source 内先分组，再做 SIM/SEM 合并；group 列记录组内全部原词")

    lines.append("## Top 关键词预览")
    lines.append("")
    lines.extend(
        _md_table(
            [
                "keyword",
                "group",
                "source",
                "score(sim)",
                "volume(sim)",
                "kd(sim)",
                "cpc(sim)",
                "volume(sem)",
                "kd(sem)",
                "cpc(sem)",
            ],
            preview_rows,
        )
    )
    lines.append("")

    lines.append("## 产物路径")
    lines.append("")
    lines.append(f"- Markdown：{output.get('reportHistoryPath', '-')}")
    lines.append(f"- Excel：{output.get('excelHistoryPath', '-')}")
    lines.append("")

    lines.append("## 备注")
    lines.append("")
    for remark in REQUIRED_REMARKS:
        lines.append(f"- {remark}")
    lines.append("")

    return "\n".join(lines)

def _require_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "缺少依赖 openpyxl。请先执行 `pip3 install -r word-from-root/requirements.txt` 再运行。"
        ) from exc
    return Workbook, Font, get_column_letter, load_workbook


def _transform_export_value(column_def: dict, value, row: Optional[dict] = None):
    transform = column_def.get("transform")
    if transform == "source_presence_label":
        mapping = {
            "both": "both（SIM+SEM）",
            "sim_only": "sim_only（仅 SIM）",
            "sem_only": "sem_only（仅 SEM）",
        }
        return mapping.get(str(value or ""), value)
    if transform == "sim_score":
        return compute_sim_score(row or {})
    if transform == "sem_trend_json":
        if isinstance(value, list):
            return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
        if value in (None, ""):
            return "[]"
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return "[]"
            try:
                parsed = json.loads(text)
            except json.JSONDecodeError:
                return text
            if isinstance(parsed, list):
                return json.dumps(parsed, ensure_ascii=False, separators=(",", ":"))
            return text
        return json.dumps(value, ensure_ascii=False)
    return value


def _compute_column_width(values: Sequence[str], width_profile: Optional[dict] = None) -> float:
    profile = width_profile or {}
    fixed = profile.get("fixed")
    if fixed is not None:
        return float(fixed)

    max_length = max((len(v) for v in values), default=0)
    width = max_length + DEFAULT_COLUMN_PADDING
    min_width = int(profile.get("min", DEFAULT_COLUMN_MIN_WIDTH))
    max_width = int(profile.get("max", DEFAULT_COLUMN_MAX_WIDTH))
    return float(max(min_width, min(width, max_width)))


def write_excel(snapshot: dict, output_path: Path) -> None:
    Workbook, Font, get_column_letter, _ = _require_openpyxl()

    meta = snapshot.get("meta") or {}
    target = meta.get("target") or {}
    summary = meta.get("summary") or {}
    merged_rows = snapshot.get("mergedRows") or []

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    keywords_sheet = workbook.create_sheet("keywords")

    summary_rows = [
        ("keyword", target.get("keyword", "")),
        ("generatedAt", meta.get("generatedAt", "")),
        ("simRawCount", summary.get("simRawCount", summary.get("simCount", 0))),
        ("semRawCount", summary.get("semRawCount", summary.get("semCount", 0))),
        ("simCount", summary.get("simCount", 0)),
        ("semCount", summary.get("semCount", 0)),
        ("mergedCount", summary.get("mergedCount", 0)),
        ("bothCount", summary.get("bothCount", 0)),
        ("simOnlyCount", summary.get("simOnlyCount", 0)),
        ("semOnlyCount", summary.get("semOnlyCount", 0)),
        ("scoredCount", summary.get("scoredCount", 0)),
        ("scoreFormula", SCORE_FORMULA),
        ("standardWordTableVersion", STANDARD_WORD_TABLE_VERSION),
        ("sourceLegendSIM", "SIM = Similarweb Keyword Generator 来源"),
        ("sourceLegendSEM", "SEM = Semrush Keyword Magic 来源"),
    ]
    for idx, (name, value) in enumerate(summary_rows, start=1):
        summary_sheet.cell(row=idx, column=1, value=name)
        summary_sheet.cell(row=idx, column=2, value=value)
    summary_sheet.freeze_panes = "A2"

    export_columns = get_keywords_export_columns()
    headers = get_keywords_export_headers()
    keywords_sheet.append(headers)
    for cell in keywords_sheet[1]:
        cell.font = Font(bold=True)

    for row in merged_rows:
        export_row = []
        for column in export_columns:
            raw_value = row.get(column["field"])
            export_row.append(_transform_export_value(column, raw_value, row))
        keywords_sheet.append(export_row)

    keywords_sheet.freeze_panes = "A2"
    keywords_sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(keywords_sheet.max_row, 1)}"

    for index, column in enumerate(export_columns, start=1):
        number_format = column.get("number_format")
        if not number_format:
            continue
        col_letter = get_column_letter(index)
        for cell in keywords_sheet[col_letter][1:]:
            cell.number_format = number_format

    for column_index, column in enumerate(export_columns, start=1):
        column_letter = get_column_letter(column_index)
        values = []
        for row_index in range(1, keywords_sheet.max_row + 1):
            value = keywords_sheet.cell(row=row_index, column=column_index).value
            values.append("" if value is None else str(value))
        keywords_sheet.column_dimensions[column_letter].width = _compute_column_width(values, column.get("width_profile"))

    for column_cells in summary_sheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        summary_sheet.column_dimensions[column_cells[0].column_letter].width = _compute_column_width(
            values,
            {"min": 14, "max": 40},
        )

    ensure_dir(output_path.parent)
    workbook.save(output_path)


def write_artifacts(*, stamp: str, snapshot: dict, fetch_archive: dict, data_dir: Path, snapshot_dir: Path, report_dir: Path, latest_report_path: Path, latest_xlsx_path: Path) -> dict:
    ensure_dir(data_dir)
    ensure_dir(snapshot_dir)
    ensure_dir(report_dir)
    ensure_dir(latest_report_path.parent)
    ensure_dir(latest_xlsx_path.parent)

    fetch_archive_path = data_dir / f"fetch-{stamp}.json"
    snapshot_path = snapshot_dir / f"snapshot-{stamp}.json"
    report_history_path = report_dir / f"report-{stamp}.md"
    excel_history_path = report_dir / f"keyword-table-{stamp}.xlsx"

    snapshot_meta = snapshot.setdefault("meta", {})
    output_meta = snapshot_meta.setdefault("output", {})
    output_meta["reportHistoryPath"] = str(report_history_path)
    output_meta["excelHistoryPath"] = str(excel_history_path)
    snapshot_meta["latestReportPath"] = str(latest_report_path)
    snapshot_meta["latestExcelPath"] = str(latest_xlsx_path)

    report_text = render_report(snapshot)
    dump_json(fetch_archive_path, fetch_archive)
    dump_json(snapshot_path, snapshot)
    report_history_path.write_text(report_text, encoding="utf-8")
    latest_report_path.write_text(report_text, encoding="utf-8")
    write_excel(snapshot, excel_history_path)
    copyfile(excel_history_path, latest_xlsx_path)

    return {
        "fetchArchivePath": fetch_archive_path,
        "snapshotPath": snapshot_path,
        "reportHistoryPath": report_history_path,
        "excelHistoryPath": excel_history_path,
    }


def rebuild_history_artifacts(snapshot: dict, report_dir: Path) -> Tuple[Path, Path, str]:
    ensure_dir(report_dir)
    meta = snapshot.get("meta") or {}
    stamp = str(meta.get("stamp") or "")
    if not stamp:
        raise RuntimeError("snapshot 缺少 meta.stamp")

    report_history_path = report_dir / f"report-{stamp}.md"
    excel_history_path = report_dir / f"keyword-table-{stamp}.xlsx"
    report_text = render_report(snapshot)
    report_history_path.write_text(report_text, encoding="utf-8")
    write_excel(snapshot, excel_history_path)
    return report_history_path, excel_history_path, report_text


def run_pipeline(args: argparse.Namespace) -> int:
    data_dir = Path(args.data_dir).resolve()
    snapshot_dir = Path(args.snapshot_dir).resolve()
    report_dir = Path(args.report_dir).resolve()
    latest_report_path = Path(args.latest_report_path).resolve()
    latest_xlsx_path = Path(args.latest_xlsx_path).resolve()
    token_path = Path(args.token_path).resolve()
    gmitm_path = Path(args.gmitm_path).resolve()

    keyword = str(args.keyword or "").strip()
    if not keyword:
        raise SystemExit("--keyword 不能为空")
    args.keyword = keyword

    token = read_text_required(token_path, "token 文件")
    gmitm = read_gmitm(gmitm_path)
    headers = build_headers(token)

    stamp = now_stamp()
    sim_raw_rows, sim_fetch = fetch_sim_keywords(args, headers)
    sem_raw_rows, sem_fetch = fetch_sem_keywords(args, headers, gmitm)

    sim_rows, sim_grouping_meta = group_source_rows_with_ai(sim_raw_rows, source="sim", args=args)
    sem_rows, sem_grouping_meta = group_source_rows_with_ai(sem_raw_rows, source="sem", args=args)

    merged_rows = merge_rows(sim_rows, sem_rows)
    summary_counts = build_summary_counts(
        sim_raw_rows=sim_raw_rows,
        sem_raw_rows=sem_raw_rows,
        sim_grouped_rows=sim_rows,
        sem_grouped_rows=sem_rows,
        merged_rows=merged_rows,
    )

    placeholder_report_history_path = report_dir / f"report-{stamp}.md"
    placeholder_excel_history_path = report_dir / f"keyword-table-{stamp}.xlsx"
    snapshot = build_snapshot(
        stamp=stamp,
        args=args,
        sim_fetch=sim_fetch,
        sem_fetch=sem_fetch,
        sim_raw_rows=sim_raw_rows,
        sem_raw_rows=sem_raw_rows,
        sim_rows=sim_rows,
        sem_rows=sem_rows,
        merged_rows=merged_rows,
        summary_counts=summary_counts,
        grouping_meta={
            "sim": sim_grouping_meta,
            "sem": sem_grouping_meta,
        },
        excel_history_path=placeholder_excel_history_path,
        report_history_path=placeholder_report_history_path,
    )

    fetch_archive = {
        "meta": {
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "stamp": stamp,
            "keyword": keyword,
            "grouping": {
                "promptVersion": GROUPING_PROMPT_VERSION,
                "model": grouping_model_label(args),
                "temperature": args.grouping_temperature,
            },
        },
        "sim": sim_fetch,
        "sem": sem_fetch,
    }

    artifact_paths = write_artifacts(
        stamp=stamp,
        snapshot=snapshot,
        fetch_archive=fetch_archive,
        data_dir=data_dir,
        snapshot_dir=snapshot_dir,
        report_dir=report_dir,
        latest_report_path=latest_report_path,
        latest_xlsx_path=latest_xlsx_path,
    )

    print(f"[done] fetch archive : {artifact_paths['fetchArchivePath']}")
    print(f"[done] snapshot      : {artifact_paths['snapshotPath']}")
    print(f"[done] report history: {artifact_paths['reportHistoryPath']}")
    print(f"[done] report latest : {latest_report_path}")
    print(f"[done] excel history : {artifact_paths['excelHistoryPath']}")
    print(f"[done] excel latest  : {latest_xlsx_path}")
    print(f"[done] sim rows(raw) : {len(sim_raw_rows)}")
    print(f"[done] sim rows(grp) : {len(sim_rows)}")
    print(f"[done] sem rows(raw) : {len(sem_raw_rows)}")
    print(f"[done] sem rows(grp) : {len(sem_rows)}")
    print(f"[done] merged rows   : {len(merged_rows)}")
    print(f"[done] scored rows   : {summary_counts['scoredCount']}")
    return 0


def rebuild_reports(args: argparse.Namespace) -> int:
    snapshot_dir = Path(args.snapshot_dir).resolve()
    report_dir = Path(args.report_dir).resolve()
    latest_report_path = Path(args.latest_report_path).resolve()
    latest_xlsx_path = Path(args.latest_xlsx_path).resolve()

    ensure_dir(snapshot_dir)
    ensure_dir(report_dir)
    ensure_dir(latest_report_path.parent)
    ensure_dir(latest_xlsx_path.parent)

    files = list_snapshot_files(snapshot_dir)
    if not files:
        raise SystemExit(f"未找到快照文件: {snapshot_dir}")

    latest_report_text = ""
    latest_excel_path: Optional[Path] = None

    for path in files:
        snapshot = load_json(path)
        report_history_path, excel_history_path, report_text = rebuild_history_artifacts(snapshot, report_dir)
        latest_report_text = report_text
        latest_excel_path = excel_history_path
        print(f"[rebuild] report: {report_history_path}")
        print(f"[rebuild] excel : {excel_history_path}")

    latest_report_path.write_text(latest_report_text, encoding="utf-8")
    if latest_excel_path is not None:
        copyfile(latest_excel_path, latest_xlsx_path)

    print(f"[done] latest report: {latest_report_path}")
    print(f"[done] latest excel : {latest_xlsx_path}")
    return 0


def validate_report(args: argparse.Namespace) -> int:
    report_path = Path(args.report).resolve()
    xlsx_path = Path(args.xlsx).resolve()
    if not report_path.exists():
        raise SystemExit(f"报告不存在: {report_path}")
    text = report_path.read_text(encoding="utf-8")

    missing_sections = [section for section in REQUIRED_SECTIONS if section not in text]
    missing_remarks = [remark for remark in REQUIRED_REMARKS if remark not in text]
    if missing_sections or missing_remarks:
        if missing_sections:
            print("缺少 section:")
            for item in missing_sections:
                print(f"- {item}")
        if missing_remarks:
            print("缺少 remark:")
            for item in missing_remarks:
                print(f"- {item}")
        raise SystemExit(1)

    if not xlsx_path.exists():
        raise SystemExit(f"Excel 不存在: {xlsx_path}")

    _Workbook, _Font, _get_column_letter, load_workbook = _require_openpyxl()
    workbook = load_workbook(xlsx_path, data_only=True)
    try:
        if "keywords" not in workbook.sheetnames:
            raise SystemExit("Excel 缺少 keywords sheet")
        sheet = workbook["keywords"]

        expected_headers = get_keywords_export_headers()
        actual_headers = [sheet.cell(row=1, column=index).value for index in range(1, sheet.max_column + 1)]
        if actual_headers != expected_headers:
            print("Excel 表头不匹配：")
            max_len = max(len(expected_headers), len(actual_headers))
            for index in range(1, max_len + 1):
                expected = expected_headers[index - 1] if index <= len(expected_headers) else None
                actual = actual_headers[index - 1] if index <= len(actual_headers) else None
                if expected != actual:
                    print(f"- 第 {index} 列 expected={expected!r} actual={actual!r}")
            raise SystemExit(1)

        if sheet.max_row >= 2:
            score_index = get_score_column_index_from_headers(expected_headers) + 1
            for row_index in range(2, sheet.max_row + 1):
                value = sheet.cell(row=row_index, column=score_index).value
                if value in (None, ""):
                    continue
                try:
                    float(value)
                except (TypeError, ValueError) as exc:
                    raise SystemExit(f"score 列不是数字（row={row_index}）") from exc
    finally:
        workbook.close()

    print(f"[ok] report validated: {report_path}")
    print(f"[ok] xlsx exists      : {xlsx_path}")
    print("[ok] xlsx headers and score column validated")
    return 0


def main() -> int:
    args = parse_args()
    if args.command == "run":
        return run_pipeline(args)
    if args.command == "rebuild-reports":
        return rebuild_reports(args)
    if args.command == "validate-report":
        return validate_report(args)
    raise SystemExit(f"未知命令: {args.command}")


if __name__ == "__main__":
    raise SystemExit(main())
