#!/usr/bin/env python3
"""监控多站点 sitemap，发现新增内页并提取关键词候选（MVP）。

Usage:
  python3 word-monitor-sitemap/_internal/scripts/word_monitor_sitemap.py run
  python3 word-monitor-sitemap/_internal/scripts/word_monitor_sitemap.py run --all-sites
  python3 word-monitor-sitemap/_internal/scripts/word_monitor_sitemap.py run --site onlinegames_io
  python3 word-monitor-sitemap/_internal/scripts/word_monitor_sitemap.py rebuild-reports
  python3 word-monitor-sitemap/_internal/scripts/word_monitor_sitemap.py validate-report --report word-monitor-sitemap/report/latest.md
"""

from __future__ import annotations

import argparse
import gzip
import json
import re
import shutil
import subprocess
import sys
import time
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple
from urllib.error import HTTPError, URLError
from urllib.parse import urlsplit, urlunsplit
from urllib.request import Request, urlopen

PROJECT_DIR = Path(__file__).resolve().parents[2]
REPO_ROOT = PROJECT_DIR.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from shared_gefei_kd import normalize_keyword_text

DEFAULT_SITES_CONFIG = PROJECT_DIR / "data" / "sites.json"
DEFAULT_DATA_ROOT = PROJECT_DIR / "data"
DEFAULT_SNAPSHOT_ROOT = PROJECT_DIR / "_internal" / "snapshots"
DEFAULT_REPORT_ROOT = PROJECT_DIR / "report"
DEFAULT_REPORT_HISTORY_ROOT = DEFAULT_REPORT_ROOT / "history"
DEFAULT_REPORT_LATEST_ROOT = DEFAULT_REPORT_ROOT

SITEMAP_RETRY = 2
RETRY_BACKOFF_SECONDS = [0.8, 1.6]
SITEMAP_TIMEOUT_SECONDS = 30
MAX_SITEMAPS_PER_SITE = 300

TOP_NEW_URLS_IN_REPORT = 80
TOP_NEW_PATTERNS_IN_REPORT = 40
TOP_KEYWORDS_IN_REPORT = 60

STANDARD_WORD_TABLE_VERSION = "v1"
STANDARD_WORD_TABLE_SPEC_PATH = REPO_ROOT / "standard-word-analysis" / "spec" / f"standard-word-table.{STANDARD_WORD_TABLE_VERSION}.json"
ANALYZE_WORDS_SCRIPT_PATH = REPO_ROOT / "analyze-words" / "_internal" / "scripts" / "analyze_words.py"
CHECK_GEFEI_KD_SCRIPT_PATH = REPO_ROOT / "check-gefei-kd" / "_internal" / "scripts" / "check_gefei_kd.py"
DEFAULT_WORDS_DIR = REPO_ROOT / "words"
DEFAULT_CHAIN_WORK_DIR = PROJECT_DIR / "_internal" / "chained"
DEFAULT_COLUMN_MIN_WIDTH = 12
DEFAULT_COLUMN_MAX_WIDTH = 72
DEFAULT_COLUMN_PADDING = 2

SIM_COLUMN_FILL_COLOR = "FFEAF4FF"
SEM_COLUMN_FILL_COLOR = "FFF4EAFF"
MERGED_LATEST_FILE = "latest.md"
MERGED_LATEST_XLSX = "latest.xlsx"

SNAPSHOT_RE = re.compile(r"^snapshot-(\d{8}-\d{6})\.json$")
TOKEN_SPLIT_RE = re.compile(r"[^a-z0-9]+")
NUMERIC_RE = re.compile(r"^\d+$")
HEXISH_RE = re.compile(r"^[a-f0-9]{16,}$", re.IGNORECASE)
UUIDISH_RE = re.compile(r"^[a-f0-9]{8}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{12}$", re.IGNORECASE)

REQUIRED_SECTIONS = [
    "## 摘要",
    "## Sitemap 概览",
    "## 新增内页",
    "## 新增路由模式",
    "## 新关键词候选",
    "## 备注",
]

REQUIRED_REMARKS = [
    "本报告基于 sitemap 可见 URL 与路由推断，不等于全站完整索引",
    "新增内页基于与最近一次同站点快照对比得出",
    "关键词候选来自新增 URL slug 分词，为机会线索而非搜索引擎全量词库",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="监控 sitemap 并输出新增内页关键词报告")
    subparsers = parser.add_subparsers(dest="command", required=True)

    run_parser = subparsers.add_parser("run", help="抓 sitemap + 快照 + 对比 + 报告（默认运行 sites.json 中 enabled=true 的站点）")
    run_group = run_parser.add_mutually_exclusive_group(required=False)
    run_group.add_argument("--site", help="仅运行单站（site id）")
    run_group.add_argument("--all-sites", action="store_true", help="运行 sites.json 中全部站点（包含 enabled=false）")
    run_parser.add_argument("--sites-config", default=str(DEFAULT_SITES_CONFIG))
    run_parser.add_argument("--data-root", default=str(DEFAULT_DATA_ROOT))
    run_parser.add_argument("--snapshot-root", default=str(DEFAULT_SNAPSHOT_ROOT))
    run_parser.add_argument("--report-history-root", default=str(DEFAULT_REPORT_HISTORY_ROOT))
    run_parser.add_argument("--report-latest-root", default=str(DEFAULT_REPORT_LATEST_ROOT))
    run_parser.add_argument("--words-dir", default=str(DEFAULT_WORDS_DIR))
    run_parser.add_argument("--chain-work-dir", default=str(DEFAULT_CHAIN_WORK_DIR))

    rebuild_parser = subparsers.add_parser("rebuild-reports", help="根据快照重建历史报告")
    rebuild_parser.add_argument("--site", help="仅重建单站（默认重建所有）")
    rebuild_parser.add_argument("--sites-config", default=str(DEFAULT_SITES_CONFIG))
    rebuild_parser.add_argument("--snapshot-root", default=str(DEFAULT_SNAPSHOT_ROOT))
    rebuild_parser.add_argument("--report-history-root", default=str(DEFAULT_REPORT_HISTORY_ROOT))
    rebuild_parser.add_argument("--report-latest-root", default=str(DEFAULT_REPORT_LATEST_ROOT))
    rebuild_parser.add_argument("--words-dir", default=str(DEFAULT_WORDS_DIR))
    rebuild_parser.add_argument("--chain-work-dir", default=str(DEFAULT_CHAIN_WORK_DIR))

    validate_parser = subparsers.add_parser("validate-report", help="校验 Markdown 报告结构")
    validate_parser.add_argument(
        "--report",
        default=str(DEFAULT_REPORT_LATEST_ROOT / MERGED_LATEST_FILE),
        help="报告路径（默认校验 report/latest.md）",
    )
    validate_parser.add_argument(
        "--xlsx",
        default=str(DEFAULT_REPORT_LATEST_ROOT / MERGED_LATEST_XLSX),
        help="Excel 路径（默认校验 report/latest.xlsx）",
    )

    return parser.parse_args()
def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d-%H%M%S")


def load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def dump_json(path: Path, payload: dict) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def safe_int(value) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def safe_float(value) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0



def to_int_if_possible(v: float) -> str:
    value = safe_float(v)
    if abs(value - round(value)) < 1e-9:
        return str(int(round(value)))
    return f"{value:.2f}"


def compute_sim_score(row: dict) -> Optional[float]:
    sim_window_volume = safe_float(row.get("simWindowVolume"))
    sim_cpc = safe_float(row.get("simCpc"))
    sim_kd = safe_float(row.get("simKd"))
    if sim_window_volume <= 0 or sim_cpc <= 0 or sim_kd <= 0:
        return None
    return sim_window_volume * sim_cpc / sim_kd


def has_sim_metrics(row: dict) -> bool:
    return (
        safe_float(row.get("simWindowVolume")) > 0
        and safe_float(row.get("simCpc")) > 0
        and safe_float(row.get("simKd")) > 0
    )


def has_sem_metrics(row: dict) -> bool:
    return (
        safe_float(row.get("semVolume")) > 0
        and safe_float(row.get("semCpc")) > 0
        and safe_float(row.get("semKd")) > 0
    )


def has_gefei_kd(row: dict) -> bool:
    value = row.get("gefeiKD")
    return value not in (None, "")


def _display_optional_number(value) -> str:
    if value in (None, ""):
        return "-"
    number = safe_float(value)
    if abs(number - round(number)) < 1e-9:
        return str(int(round(number)))
    return f"{number:.2f}"


def _fmt_md_cell(value) -> str:
    if value is None:
        text = "-"
    else:
        text = str(value).strip()
    if not text:
        text = "-"

    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    normalized = re.sub(r"\s*\|\s*", "\n", normalized)
    normalized = normalized.replace(" / ", "\n")
    lines = [line.strip() for line in normalized.split("\n") if line.strip()]
    if not lines:
        lines = ["-"]

    escaped: List[str] = []
    for line in lines:
        line = re.sub(r"([/?&#=_-])", r"\1<wbr>", line)
        escaped.append(line.replace("|", "\\|"))
    return "<br>".join(escaped)


def _md_table(headers: Sequence[str], rows: Sequence[Sequence[str]]) -> List[str]:

    if not rows:
        return ["（无）"]
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(_fmt_md_cell(cell) for cell in row) + " |")
    return lines


def _normalize_export_column(raw: dict) -> dict:
    if not isinstance(raw, dict):
        raise RuntimeError("标准词表列定义必须是对象")

    header = str(raw.get("header") or "").strip()
    field = str(raw.get("field") or "").strip()
    if not header or not field:
        raise RuntimeError("标准词表列定义缺少 header 或 field")

    column = {
        "header": header,
        "field": field,
        "type": str(raw.get("type") or "string").strip() or "string",
        "nullable": bool(raw.get("nullable", True)),
    }

    transform = raw.get("transform")
    if transform not in (None, ""):
        column["transform"] = str(transform)

    number_format = raw.get("number_format")
    if number_format not in (None, ""):
        column["number_format"] = str(number_format)

    width_profile = raw.get("width_profile")
    if isinstance(width_profile, dict):
        normalized_width: dict = {}
        if width_profile.get("fixed") is not None:
            normalized_width["fixed"] = width_profile.get("fixed")
        if width_profile.get("min") is not None:
            normalized_width["min"] = width_profile.get("min")
        if width_profile.get("max") is not None:
            normalized_width["max"] = width_profile.get("max")
        if normalized_width:
            column["width_profile"] = normalized_width

    return column


def load_standard_word_table_spec() -> dict:
    if not STANDARD_WORD_TABLE_SPEC_PATH.exists():
        raise RuntimeError(
            "标准词表规范缺失: "
            f"{STANDARD_WORD_TABLE_SPEC_PATH}。"
            "请先补齐 standard-word-analysis/spec/standard-word-table.v1.json。"
        )

    spec = load_json(STANDARD_WORD_TABLE_SPEC_PATH)
    if not isinstance(spec, dict):
        raise RuntimeError("标准词表规范格式错误：顶层必须是对象")

    version = str(spec.get("version") or "").strip()
    if version != STANDARD_WORD_TABLE_VERSION:
        raise RuntimeError(
            f"标准词表版本不匹配：expected={STANDARD_WORD_TABLE_VERSION} actual={version or '-'}"
        )

    raw_columns = spec.get("excelColumns")
    if not isinstance(raw_columns, list) or not raw_columns:
        raise RuntimeError("标准词表规范缺少 excelColumns 数组")

    columns = [_normalize_export_column(item) for item in raw_columns]
    fields = [str(column.get("field") or "") for column in columns]
    if len(fields) != len(set(fields)):
        raise RuntimeError("标准词表规范字段重复：excelColumns.field 必须唯一")

    return {
        "version": version,
        "columns": columns,
    }


STANDARD_WORD_TABLE_SPEC = load_standard_word_table_spec()


def get_standard_word_table_spec() -> dict:
    return STANDARD_WORD_TABLE_SPEC


def get_keywords_export_columns() -> List[dict]:
    return list(get_standard_word_table_spec()["columns"])


def get_keywords_export_headers() -> List[str]:
    return [column["header"] for column in get_keywords_export_columns()]


def get_header_for_field(field_name: str) -> str:
    for column in get_keywords_export_columns():
        if column.get("field") == field_name:
            return str(column.get("header") or "")
    raise RuntimeError(f"标准词表缺少字段: {field_name}")


def get_score_column_index_from_headers(headers: Sequence[str]) -> int:
    score_header = str(get_standard_word_table_spec().get("scoreColumnHeader") or "").strip()
    if not score_header:
        score_header = "score(simWindowVolume*cpc/kd)"
    try:
        return list(headers).index(score_header)
    except ValueError as exc:
        raise RuntimeError(f"标准词表缺少 score 列: {score_header}") from exc


def _require_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "缺少依赖 openpyxl。请先执行 `pip3 install openpyxl` 再运行。"
        ) from exc
    return Workbook, Font, Alignment, PatternFill, get_column_letter, load_workbook


MULTILINE_EXPORT_FIELDS = {
    "group",
    "correspondingDomain",
}


def _format_multiline_export_text(field: str, value):
    if value is None:
        return value
    if field not in MULTILINE_EXPORT_FIELDS:
        return value
    text = str(value).strip()
    if not text:
        return text
    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    normalized = re.sub(r"\s*\|\s*", "\n", normalized)
    normalized = normalized.replace(" / ", "\n")
    lines = [line.strip() for line in normalized.split("\n") if line.strip()]
    return "\n".join(lines) if lines else ""


def _compute_column_width(values: Sequence[str], width_profile: Optional[dict] = None) -> float:
    profile = width_profile or {}
    fixed = profile.get("fixed")
    if fixed is not None:
        return float(fixed)

    max_length = 0
    for raw in values:
        text = "" if raw is None else str(raw)
        parts = text.splitlines() or [text]
        longest_line = max((len(part) for part in parts), default=0)
        if longest_line > max_length:
            max_length = longest_line

    width = max_length + DEFAULT_COLUMN_PADDING
    min_width = int(profile.get("min", DEFAULT_COLUMN_MIN_WIDTH))
    max_width = int(profile.get("max", DEFAULT_COLUMN_MAX_WIDTH))
    return float(max(min_width, min(width, max_width)))


def _transform_export_value(column_def: dict, value):
    transform = column_def.get("transform")
    if transform == "source_presence_label":
        mapping = {
            "both": "both（SIM+SEM）",
            "sim_only": "sim_only（仅 SIM）",
            "sem_only": "sem_only（仅 SEM）",
        }
        transformed = mapping.get(str(value or ""), value)
    else:
        transformed = value
    field = str(column_def.get("field") or "")
    return _format_multiline_export_text(field, transformed)

def stable_unique_texts(values: Sequence[str]) -> List[str]:
    seen: set[str] = set()
    result: List[str] = []
    for value in values:
        item = str(value or "").strip()
        if not item:
            continue
        key = normalize_keyword_text(item)
        if key in seen:
            continue
        seen.add(key)
        result.append(item)
    return result


def _local_name(tag: str) -> str:
    if not isinstance(tag, str):
        return ""
    if tag.startswith("{") and "}" in tag:
        return tag.split("}", 1)[1]
    return tag


def _iter_loc_texts(root: ET.Element, mode: str) -> List[str]:
    if mode == "urlset":
        xpath = ".//{*}url/{*}loc"
    elif mode == "sitemapindex":
        xpath = ".//{*}sitemap/{*}loc"
    else:
        xpath = ".//{*}loc"

    locs = []
    for node in root.findall(xpath):
        text = (node.text or "").strip()
        if text:
            locs.append(text)

    # 兜底：某些站点命名空间异常时，按本地名扫描。
    if locs:
        return locs

    if mode == "urlset":
        parent_name = "url"
    elif mode == "sitemapindex":
        parent_name = "sitemap"
    else:
        parent_name = None

    for parent in root.iter():
        if parent_name and _local_name(parent.tag) != parent_name:
            continue
        for child in list(parent):
            if _local_name(child.tag) != "loc":
                continue
            text = (child.text or "").strip()
            if text:
                locs.append(text)
    return locs


def _decode_http_body(raw: bytes, headers: dict, url: str) -> str:
    data = raw
    content_encoding = (headers.get("Content-Encoding") or "").lower()
    if "gzip" in content_encoding or url.lower().endswith(".gz") or data[:2] == b"\x1f\x8b":
        try:
            data = gzip.decompress(data)
        except OSError:
            pass

    for enc in ("utf-8", "utf-8-sig", "gb18030", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def fetch_url_text_with_retry(url: str, timeout_seconds: int = SITEMAP_TIMEOUT_SECONDS) -> dict:
    last_error: Optional[Exception] = None
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/125.0 Safari/537.36 sitemap-monitor/1.0",
        "Accept": "application/xml,text/xml,application/xhtml+xml,text/html;q=0.8,*/*;q=0.5",
        "Accept-Language": "en-US,en;q=0.9",
    }

    for attempt in range(SITEMAP_RETRY + 1):
        req = Request(url, headers=headers, method="GET")
        try:
            with urlopen(req, timeout=timeout_seconds) as resp:
                status = int(getattr(resp, "status", 200) or 200)
                body_bytes = resp.read()
                body_text = _decode_http_body(body_bytes, dict(resp.headers), str(resp.geturl()))
                return {
                    "url": url,
                    "finalUrl": str(resp.geturl()),
                    "status": status,
                    "attempt": attempt + 1,
                    "headers": {k: v for k, v in resp.headers.items()},
                    "bytes": len(body_bytes),
                    "body": body_text,
                }
        except HTTPError as exc:
            body = b""
            if hasattr(exc, "read"):
                try:
                    body = exc.read()
                except Exception:
                    body = b""
            decoded = _decode_http_body(body, dict(getattr(exc, "headers", {}) or {}), url)
            last_error = RuntimeError(f"HTTP {exc.code} {exc.reason}; body={decoded[:200]}")
        except (URLError, TimeoutError, OSError) as exc:
            last_error = exc

        if attempt < SITEMAP_RETRY:
            time.sleep(RETRY_BACKOFF_SECONDS[min(attempt, len(RETRY_BACKOFF_SECONDS) - 1)])

    raise RuntimeError(f"抓取 sitemap 失败: {url}: {last_error}")


def normalize_url_for_queue(raw_url: str) -> Optional[str]:
    if not isinstance(raw_url, str):
        return None
    text = raw_url.strip()
    if not text:
        return None
    if "://" not in text:
        text = f"https://{text}"

    try:
        parsed = urlsplit(text)
    except ValueError:
        return None

    host = (parsed.hostname or "").strip().lower().rstrip(".")
    if not host:
        return None

    path = parsed.path or "/"
    query = parsed.query or ""
    scheme = (parsed.scheme or "https").lower()
    if scheme not in {"http", "https"}:
        scheme = "https"

    normalized = urlunsplit((scheme, host, path, query, ""))
    return normalized


def crawl_sitemaps(site: dict) -> Tuple[List[str], dict]:
    site_id = str(site.get("id", "")).strip()
    start_urls = site.get("sitemapUrls") or []
    if not isinstance(start_urls, list) or not start_urls:
        raise RuntimeError(f"站点 {site_id} 未配置 sitemapUrls")

    queue: List[Tuple[str, int, Optional[str]]] = []
    visited = set()
    fetches: List[dict] = []
    page_urls: List[str] = []

    for raw in start_urls:
        normalized = normalize_url_for_queue(str(raw))
        if normalized and normalized not in visited:
            queue.append((normalized, 0, None))
            visited.add(normalized)

    if not queue:
        raise RuntimeError(f"站点 {site_id} 的 sitemapUrls 无有效 URL")

    discovered_sitemap_urls = 0

    while queue:
        if len(fetches) >= MAX_SITEMAPS_PER_SITE:
            raise RuntimeError(
                f"站点 {site_id} sitemap 数量超过上限 {MAX_SITEMAPS_PER_SITE}，可能存在异常膨胀或循环引用"
            )

        sitemap_url, depth, parent = queue.pop(0)
        fetched = fetch_url_text_with_retry(sitemap_url)

        try:
            root = ET.fromstring(fetched["body"])
        except ET.ParseError as exc:
            raise RuntimeError(f"解析 XML 失败: {sitemap_url}: {exc}") from exc

        root_name = _local_name(root.tag)
        if root_name == "urlset":
            mode = "urlset"
            locs = _iter_loc_texts(root, "urlset")
            page_urls.extend(locs)
            child_sitemaps = 0
        elif root_name == "sitemapindex":
            mode = "sitemapindex"
            locs = _iter_loc_texts(root, "sitemapindex")
            child_sitemaps = len(locs)
            for raw in locs:
                normalized = normalize_url_for_queue(raw)
                if not normalized or normalized in visited:
                    continue
                queue.append((normalized, depth + 1, sitemap_url))
                visited.add(normalized)
                discovered_sitemap_urls += 1
        else:
            mode = "unknown"
            locs = _iter_loc_texts(root, "generic")
            page_urls.extend(locs)
            child_sitemaps = 0

        fetches.append(
            {
                "url": sitemap_url,
                "finalUrl": fetched.get("finalUrl"),
                "status": fetched.get("status"),
                "attempt": fetched.get("attempt"),
                "depth": depth,
                "parent": parent,
                "type": mode,
                "locCount": len(locs),
                "childSitemaps": child_sitemaps,
                "bytes": fetched.get("bytes", 0),
            }
        )

    deduped = sorted({u.strip() for u in page_urls if isinstance(u, str) and u.strip()})

    crawl_meta = {
        "siteId": site_id,
        "startSitemapUrls": start_urls,
        "sitemapsFetched": len(fetches),
        "sitemapUrlsDiscovered": discovered_sitemap_urls,
        "ignoredByDepthLimit": 0,
        "rawPageUrlCount": len(page_urls),
        "dedupedPageUrlCount": len(deduped),
        "fetches": fetches,
        "limits": {
            "maxSitemapsPerSite": MAX_SITEMAPS_PER_SITE,
        },
    }
    return deduped, crawl_meta


def normalize_page_url(raw_url: str) -> Optional[dict]:
    if not isinstance(raw_url, str):
        return None
    text = raw_url.strip()
    if not text:
        return None

    if "://" not in text:
        text = f"https://{text}"

    try:
        parsed = urlsplit(text)
    except ValueError:
        return None

    host = (parsed.hostname or "").strip().lower().rstrip(".")
    if not host:
        return None

    path = parsed.path or "/"
    if path != "/":
        path = "/" + path.lstrip("/")
        path = path.rstrip("/") or "/"

    canonical = f"https://{host}{path}"
    segments = [seg for seg in path.split("/") if seg]
    slug = segments[-1] if segments else ""

    return {
        "url": canonical,
        "host": host,
        "path": path,
        "segments": segments,
        "slug": slug,
        "depth": len(segments),
    }


def compile_path_rules(regex_list: Iterable[str]) -> List[re.Pattern]:
    compiled = []
    for idx, item in enumerate(regex_list, start=1):
        pattern = str(item or "").strip()
        if not pattern:
            continue
        try:
            compiled.append(re.compile(pattern))
        except re.error as exc:
            raise RuntimeError(f"excludePathRegexes 第 {idx} 条正则非法: {pattern}; {exc}") from exc
    return compiled


def normalize_site_config(site: dict) -> dict:
    site_id = str(site.get("id", "")).strip()
    if not site_id:
        raise RuntimeError("site.id 不能为空")

    sitemap_urls = site.get("sitemapUrls") or []
    if not isinstance(sitemap_urls, list) or not sitemap_urls:
        raise RuntimeError(f"site={site_id} 缺少 sitemapUrls")

    include_hosts = [str(x).strip().lower() for x in (site.get("includeHosts") or []) if str(x).strip()]
    if not include_hosts:
        raise RuntimeError(f"site={site_id} 缺少 includeHosts")

    exclude_regexes = site.get("excludePathRegexes") or []
    if not isinstance(exclude_regexes, list):
        raise RuntimeError(f"site={site_id} excludePathRegexes 必须是数组")

    keyword_rules = site.get("keywordRules") or {}
    stop_tokens = keyword_rules.get("stopTokens") or []
    min_token_length = safe_int(keyword_rules.get("minTokenLength") or 2)
    if min_token_length <= 0:
        min_token_length = 2

    return {
        "id": site_id,
        "enabled": bool(site.get("enabled", True)),
        "displayName": str(site.get("displayName") or site_id),
        "sitemapUrls": sitemap_urls,
        "includeHosts": include_hosts,
        "excludePathRegexes": [str(x) for x in exclude_regexes],
        "compiledExcludePathRegexes": compile_path_rules(exclude_regexes),
        "keywordRules": {
            "stopTokens": [str(t).strip().lower() for t in stop_tokens if str(t).strip()],
            "minTokenLength": min_token_length,
            "dropNumericOnlyToken": bool(keyword_rules.get("dropNumericOnlyToken", True)),
        },
    }


def load_sites_config(path: Path) -> Dict[str, dict]:
    payload = load_json(path)
    sites = payload.get("sites")
    if not isinstance(sites, list) or not sites:
        raise RuntimeError(f"sites 配置为空: {path}")

    result: Dict[str, dict] = {}
    for raw_site in sites:
        if not isinstance(raw_site, dict):
            continue
        site = normalize_site_config(raw_site)
        site_id = site["id"]
        if site_id in result:
            raise RuntimeError(f"sites 配置存在重复 id: {site_id}")
        result[site_id] = site
    if not result:
        raise RuntimeError(f"sites 配置无有效站点: {path}")
    return result


def filter_and_normalize_urls(raw_urls: List[str], site: dict) -> Tuple[List[dict], dict]:
    include_hosts = set(site.get("includeHosts") or [])
    exclude_rules: List[re.Pattern] = site.get("compiledExcludePathRegexes") or []

    stats = {
        "rawUrls": len(raw_urls),
        "invalidUrls": 0,
        "nonTargetHostUrls": 0,
        "excludedByPathRules": 0,
        "dedupedUrls": 0,
    }

    deduped: Dict[str, dict] = {}

    for raw in raw_urls:
        item = normalize_page_url(raw)
        if item is None:
            stats["invalidUrls"] += 1
            continue

        if item["host"] not in include_hosts:
            stats["nonTargetHostUrls"] += 1
            continue

        excluded = False
        for rule in exclude_rules:
            if rule.search(item["path"]):
                excluded = True
                break
        if excluded:
            stats["excludedByPathRules"] += 1
            continue

        deduped[item["url"]] = item

    rows = list(deduped.values())
    rows.sort(key=lambda x: x["url"])

    stats["dedupedUrls"] = len(rows)
    return rows, stats


def segment_shape(seg: str) -> str:
    text = str(seg or "").strip().lower()
    if not text:
        return "{seg}"
    if NUMERIC_RE.match(text):
        return "{num}"
    if HEXISH_RE.match(text) or UUIDISH_RE.match(text):
        return "{id}"

    if re.match(r"^[a-z0-9]+(?:-[a-z0-9]+)*$", text):
        tokens = text.split("-")
        if len(tokens) == 1:
            return "{token}"
        return "-".join(["{token}"] * len(tokens))

    return "{seg}"


def infer_patterns(rows: List[dict]) -> List[dict]:
    grouped: Dict[str, dict] = {}

    for row in rows:
        segments = row.get("segments") or []
        if not segments:
            pattern = "/"
        else:
            shaped = [segment_shape(seg) for seg in segments]
            pattern = "/" + "/".join(shaped) + "/"

        item = grouped.setdefault(
            pattern,
            {
                "pattern": pattern,
                "count": 0,
                "exampleUrls": [],
            },
        )
        item["count"] += 1
        if len(item["exampleUrls"]) < 3:
            item["exampleUrls"].append(row.get("url", ""))

    patterns = list(grouped.values())
    patterns.sort(key=lambda x: (-safe_int(x.get("count")), x.get("pattern", "")))
    return patterns


def build_comparison(today_rows: List[dict], baseline_rows: List[dict], today_patterns: List[dict], baseline_patterns: List[dict]) -> dict:
    baseline_map = {row.get("url"): row for row in baseline_rows}
    today_map = {row.get("url"): row for row in today_rows}

    newly_added_urls = [row for row in today_rows if row.get("url") not in baseline_map]
    removed_urls = [row for row in baseline_rows if row.get("url") not in today_map]

    baseline_pattern_map = {item.get("pattern"): item for item in baseline_patterns}
    new_patterns = [item for item in today_patterns if item.get("pattern") not in baseline_pattern_map]

    return {
        "newlyAddedUrls": newly_added_urls,
        "removedUrls": removed_urls,
        "newPatterns": new_patterns,
    }


def _clean_tokens(slug_text: str, keyword_rules: dict) -> List[str]:
    min_token_length = safe_int(keyword_rules.get("minTokenLength") or 2)
    drop_numeric_only = bool(keyword_rules.get("dropNumericOnlyToken", True))

    raw_tokens = [t for t in TOKEN_SPLIT_RE.split((slug_text or "").lower()) if t]
    cleaned = []
    for t in raw_tokens:
        if drop_numeric_only and NUMERIC_RE.match(t):
            continue
        if len(t) < min_token_length:
            continue
        cleaned.append(t)
    return cleaned


def extract_keywords(new_urls: List[dict], keyword_rules: dict) -> dict:
    phrase_counter: Counter = Counter()
    phrase_urls: Dict[str, set] = defaultdict(set)
    display_phrase_by_key: Dict[str, str] = {}

    for row in new_urls:
        slug = str(row.get("slug") or "").strip().lower()
        if not slug:
            continue

        cleaned = _clean_tokens(slug, keyword_rules)
        if not cleaned:
            continue

        phrase = " ".join(cleaned).strip()
        normalized_phrase = normalize_keyword_text(phrase)
        if not normalized_phrase:
            continue

        source_url = str(row.get("url") or "").strip()
        display_phrase_by_key.setdefault(normalized_phrase, phrase)
        phrase_counter[normalized_phrase] += 1
        if source_url:
            phrase_urls[normalized_phrase].add(source_url)

    entries = []
    for key, count in phrase_counter.items():
        urls = sorted({u for u in phrase_urls.get(key, set()) if u})
        url_count = len(urls)
        score = round(float(count) * 1.2 + float(url_count) * 0.3, 3)
        entries.append(
            {
                "type": "phrase",
                "keyword": display_phrase_by_key.get(key, key),
                "count": int(count),
                "urlCount": url_count,
                "score": score,
                "exampleUrls": urls[:3],
            }
        )

    entries.sort(key=lambda x: (-float(x.get("score", 0.0)), -safe_int(x.get("urlCount")), x.get("keyword", ""), x.get("type", "")))

    return {
        "topKeywordsFromNewUrls": entries[:TOP_KEYWORDS_IN_REPORT],
        "allKeywordsFromNewUrls": entries,
        "counters": {
            "phrase": [(display_phrase_by_key.get(key, key), count) for key, count in phrase_counter.most_common(120)],
            "bigram": [],
            "token": [],
        },
    }


def build_snapshot(
    *,
    stamp: str,
    site: dict,
    crawl_meta: dict,
    normalize_stats: dict,
    rows: List[dict],
    patterns: List[dict],
    baseline_stamp: Optional[str],
    baseline_mode: bool,
    comparison: dict,
    keyword_result: dict,
) -> dict:
    standard_word = keyword_result.get("standardWord") or {}
    return {
        "meta": {
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "stamp": stamp,
            "baselineMode": baseline_mode,
            "site": {
                "id": site.get("id"),
                "displayName": site.get("displayName"),
                "sitemapUrls": site.get("sitemapUrls"),
                "includeHosts": site.get("includeHosts"),
            },
            "crawl": {
                "sitemapsFetched": crawl_meta.get("sitemapsFetched", 0),
                "sitemapUrlsDiscovered": crawl_meta.get("sitemapUrlsDiscovered", 0),
                "ignoredByDepthLimit": crawl_meta.get("ignoredByDepthLimit", 0),
                "rawPageUrlCount": crawl_meta.get("rawPageUrlCount", 0),
                "dedupedPageUrlCount": crawl_meta.get("dedupedPageUrlCount", 0),
            },
            "normalize": normalize_stats,
            "effectiveUrlCount": len(rows),
            "patternCount": len(patterns),
            "baselineStamp": baseline_stamp,
            "standardWordTableVersion": STANDARD_WORD_TABLE_VERSION,
        },
        "urls": rows,
        "patterns": patterns,
        "comparison": {
            "baselineStamp": baseline_stamp,
            "newlyAddedUrls": comparison.get("newlyAddedUrls", []),
            "removedUrls": comparison.get("removedUrls", []),
            "newPatterns": comparison.get("newPatterns", []),
        },
        "keywords": {
            "topKeywordsFromNewUrls": keyword_result.get("topKeywordsFromNewUrls", []),
            "counters": keyword_result.get("counters", {}),
        },
        "standardWord": {
            "rows": standard_word.get("rows", []),
            "summary": standard_word.get("summary", {}),
        },
    }


def render_report(snapshot: dict) -> str:
    meta = snapshot.get("meta") or {}
    site_meta = meta.get("site") or {}
    comparison = snapshot.get("comparison") or {}

    urls = snapshot.get("urls") or []
    patterns = snapshot.get("patterns") or []
    new_urls = comparison.get("newlyAddedUrls") or []
    removed_urls = comparison.get("removedUrls") or []
    new_patterns = comparison.get("newPatterns") or []
    top_keywords = (snapshot.get("keywords") or {}).get("topKeywordsFromNewUrls") or []

    depth_counter = Counter()
    for row in urls:
        depth_counter[safe_int(row.get("depth"))] += 1

    lines: List[str] = []
    lines.append(f"# Sitemap 监控报告（{site_meta.get('id', '-')}｜{meta.get('stamp', '-')}）")
    lines.append("")

    lines.append("## 摘要")
    lines.append("")
    lines.append(f"- 站点：{site_meta.get('displayName', '-')}")
    lines.append(f"- sitemap 抓取数：{(meta.get('crawl') or {}).get('sitemapsFetched', 0)}")
    lines.append(f"- URL 原始数：{(meta.get('crawl') or {}).get('rawPageUrlCount', 0)}")
    lines.append(f"- URL 去重数：{(meta.get('crawl') or {}).get('dedupedPageUrlCount', 0)}")
    lines.append(f"- 有效内页数：{meta.get('effectiveUrlCount', 0)}")
    lines.append(f"- 路由模式数：{meta.get('patternCount', 0)}")

    if meta.get("baselineMode"):
        lines.append("- 基线：无历史快照（本次仅建立基线）")
    else:
        lines.append(f"- 基线：{meta.get('baselineStamp', '-')}")

    lines.append(f"- 新增内页数：{len(new_urls)}")
    lines.append(f"- 移除内页数：{len(removed_urls)}")
    lines.append(f"- 新增路由模式数：{len(new_patterns)}")
    lines.append(f"- 新关键词候选数：{len(top_keywords)}")
    lines.append("")

    lines.append("## Sitemap 概览")
    lines.append("")
    lines.append(f"- includeHosts: {', '.join(site_meta.get('includeHosts') or [])}")
    lines.append(
        "- pathDepth 分布: "
        + ", ".join([f"depth={depth}: {count}" for depth, count in sorted(depth_counter.items())])
        if depth_counter
        else "- pathDepth 分布: （无）"
    )
    lines.append("")

    lines.append("## 新增内页")
    lines.append("")
    if meta.get("baselineMode"):
        lines.append("本次仅建立基线，不输出新增内页。")
    else:
        rows = []
        for row in new_urls[:TOP_NEW_URLS_IN_REPORT]:
            rows.append(
                [
                    row.get("url", "-"),
                    row.get("path", "-"),
                    row.get("slug", "-") or "-",
                    str(row.get("depth", "-")),
                ]
            )
        lines.extend(_md_table(["url", "path", "slug", "depth"], rows))
    lines.append("")

    lines.append("## 新增路由模式")
    lines.append("")
    if meta.get("baselineMode"):
        lines.append("本次仅建立基线，不输出新增路由模式。")
    else:
        rows = []
        for row in new_patterns[:TOP_NEW_PATTERNS_IN_REPORT]:
            rows.append(
                [
                    row.get("pattern", "-"),
                    str(row.get("count", 0)),
                    "<br>".join(row.get("exampleUrls") or []),
                ]
            )
        lines.extend(_md_table(["pattern", "count", "examples"], rows))
    lines.append("")

    lines.append("## 新关键词候选")
    lines.append("")
    if meta.get("baselineMode"):
        lines.append("本次仅建立基线，不输出关键词候选。")
    else:
        rows = []
        for kw in top_keywords[:TOP_KEYWORDS_IN_REPORT]:
            rows.append(
                [
                    kw.get("type", "-"),
                    kw.get("keyword", "-"),
                    f"{float(kw.get('score', 0.0)):.2f}",
                    str(kw.get("urlCount", 0)),
                    "<br>".join(kw.get("exampleUrls") or []),
                ]
            )
        lines.extend(_md_table(["type", "keyword", "score", "urlCount", "examples"], rows))
    lines.append("")

    lines.append("## 备注")
    lines.append("")
    for remark in REQUIRED_REMARKS:
        lines.append(f"- {remark}")
    lines.append("")

    lines.append("## 附录：主路由模式（Top 10）")
    lines.append("")
    rows = []
    for p in patterns[:10]:
        rows.append(
            [
                p.get("pattern", "-"),
                str(p.get("count", 0)),
                "<br>".join(p.get("exampleUrls") or []),
            ]
        )
    lines.extend(_md_table(["pattern", "count", "examples"], rows))
    lines.append("")

    return "\n".join(lines)


def render_merged_report(*, stamp: str, site_results: List[dict], final_standard_word_rows: Optional[List[dict]] = None) -> str:
    total_effective = sum(safe_int(r.get("effectiveUrlCount")) for r in site_results)
    total_new = sum(safe_int(r.get("newlyAddedCount")) for r in site_results)
    total_removed = sum(safe_int(r.get("removedCount")) for r in site_results)
    total_patterns = sum(safe_int(r.get("newPatternCount")) for r in site_results)
    total_keywords = sum(safe_int(r.get("keywordCount")) for r in site_results)

    if final_standard_word_rows is None:
        merged_standard_word_rows = build_merged_standard_word_rows(site_results).get("rows") or []
    else:
        merged_standard_word_rows = final_standard_word_rows

    sim_ready_count = sum(1 for row in merged_standard_word_rows if has_sim_metrics(row))
    sem_ready_count = sum(1 for row in merged_standard_word_rows if has_sem_metrics(row))
    gefei_kd_ready_count = sum(1 for row in merged_standard_word_rows if has_gefei_kd(row))

    lines: List[str] = []
    lines.append(f"# Sitemap 合并监控报告（{stamp}）")
    lines.append("")

    lines.append("## 摘要")
    lines.append("")
    lines.append(f"- 站点数：{len(site_results)}")
    lines.append("- 站点列表：" + ", ".join([r.get("siteId", "-") for r in site_results]))
    lines.append(f"- 有效内页总数：{total_effective}")
    lines.append(f"- 新增内页总数：{total_new}")
    lines.append(f"- 移除内页总数：{total_removed}")
    lines.append(f"- 新增路由模式总数：{total_patterns}")
    lines.append(f"- 新关键词候选总数：{total_keywords}")
    lines.append(f"- 标准词表行数：{len(merged_standard_word_rows)}（最终完整词表）")
    lines.append(f"- SIM 指标完整行数：{sim_ready_count}")
    lines.append(f"- SEM 指标完整行数：{sem_ready_count}")
    lines.append(f"- gefeiKD 已回填行数：{gefei_kd_ready_count}")
    lines.append("")

    lines.append("## Sitemap 概览")
    lines.append("")
    overview_rows = []
    for r in site_results:
        overview_rows.append(
            [
                r.get("siteId", "-"),
                str(r.get("effectiveUrlCount", 0)),
                str(r.get("patternCount", 0)),
                "yes" if r.get("baselineMode") else str(r.get("baselineStamp") or "no"),
            ]
        )
    lines.extend(_md_table(["siteId", "effectiveUrls", "patterns", "baseline"], overview_rows))
    lines.append("")

    lines.append("## 新增内页")
    lines.append("")
    new_rows = []
    for r in site_results:
        for item in (r.get("newlyAddedUrls") or [])[:TOP_NEW_URLS_IN_REPORT]:
            new_rows.append(
                [
                    r.get("siteId", "-"),
                    item.get("url", "-"),
                    item.get("path", "-"),
                    item.get("slug", "-") or "-",
                    str(item.get("depth", "-")),
                ]
            )
    lines.extend(_md_table(["siteId", "url", "path", "slug", "depth"], new_rows))
    lines.append("")

    lines.append("## 新增路由模式")
    lines.append("")
    pattern_rows = []
    for r in site_results:
        for item in (r.get("newPatterns") or [])[:TOP_NEW_PATTERNS_IN_REPORT]:
            pattern_rows.append(
                [
                    r.get("siteId", "-"),
                    item.get("pattern", "-"),
                    str(item.get("count", 0)),
                    "<br>".join(item.get("exampleUrls") or []),
                ]
            )
    lines.extend(_md_table(["siteId", "pattern", "count", "examples"], pattern_rows))
    lines.append("")

    lines.append("## 新关键词候选")
    lines.append("")
    keyword_rows = []
    for r in site_results:
        for kw in (r.get("topKeywords") or [])[:TOP_KEYWORDS_IN_REPORT]:
            keyword_rows.append(
                [
                    r.get("siteId", "-"),
                    kw.get("type", "-"),
                    kw.get("keyword", "-"),
                    f"{float(kw.get('score', 0.0)):.2f}",
                    str(kw.get("urlCount", 0)),
                    "<br>".join(kw.get("exampleUrls") or []),
                ]
            )
    lines.extend(_md_table(["siteId", "type", "keyword", "score", "urlCount", "examples"], keyword_rows))
    lines.append("")

    lines.append("## 最终标准词表摘要")
    lines.append("")
    final_table_rows = [
        [
            row.get("keyword", "-") or "-",
            _display_optional_number(row.get("simWindowVolume")),
            _display_optional_number(row.get("semVolume")),
            _display_optional_number(row.get("gefeiKD")),
            _display_optional_number(compute_sim_score(row)),
        ]
        for row in merged_standard_word_rows[:20]
    ]
    lines.extend(_md_table(["keyword", "simVolume", "semVolume", "gefeiKD", "score(sim)"], final_table_rows))
    lines.append("")

    lines.append("## 备注")
    lines.append("")
    lines.append("- 标准词表为最终口径：种子词表经 analyze-words（SIM/SEM）与 check-gefei-kd（gefeiKD）补全后写入 report 与 words 目录")
    for remark in REQUIRED_REMARKS:
        lines.append(f"- {remark}")
    lines.append("")

    return "\n".join(lines)


def _make_standard_word_row(keyword: str, corresponding_domain: str) -> dict:
    return {
        "keyword": keyword,
        "correspondingDomain": corresponding_domain,
        "group": "",
        "sourcePresence": "",
        "score": None,
        "simWindowVolume": None,
        "simKd": None,
        "simCpc": None,
        "semVolume": None,
        "semKd": None,
        "semCpc": None,
        "gefeiKD": None,
    }


def build_standard_word_rows(top_keywords: List[dict], newly_added_urls: List[dict], keyword_rules: dict) -> Tuple[List[dict], dict]:
    url_map: Dict[str, List[str]] = defaultdict(list)
    for row in newly_added_urls:
        source_url = str(row.get("url") or "").strip()
        phrase = normalize_keyword_text(" ".join(_clean_tokens(str(row.get("slug") or ""), keyword_rules)))
        if source_url and phrase:
            url_map[phrase].append(source_url)

    rows: List[dict] = []
    for entry in top_keywords:
        keyword = str(entry.get("keyword") or "").strip()
        normalized_keyword = normalize_keyword_text(keyword)
        if not normalized_keyword:
            continue
        urls = " | ".join(stable_unique_texts(url_map.get(normalized_keyword, [])))
        rows.append(_make_standard_word_row(keyword, urls))

    return rows, {
        "rowCount": len(rows),
    }


def write_standard_word_excel(*, stamp: str, site_results: List[dict], output_path: Path) -> None:
    Workbook, Font, Alignment, PatternFill, get_column_letter, _ = _require_openpyxl()

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    keywords_sheet = workbook.create_sheet("keywords")

    merged_rows = build_merged_standard_word_rows(site_results)
    standard_word_rows = merged_rows.get("rows") or []
    summary_rows = [
        ("generatedAt", datetime.now().isoformat(timespec="seconds")),
        ("stamp", stamp),
        ("siteCount", len(site_results)),
        ("standardWordRows", len(standard_word_rows)),
        ("standardWordTableVersion", STANDARD_WORD_TABLE_VERSION),
    ]
    for idx, (name, value) in enumerate(summary_rows, start=1):
        summary_sheet.cell(row=idx, column=1, value=name)
        summary_sheet.cell(row=idx, column=2, value=value)
    summary_sheet.freeze_panes = "A2"

    export_columns = get_keywords_export_columns()
    headers = get_keywords_export_headers()
    keywords_sheet.append(headers)
    for cell in keywords_sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    field_to_index = {str(column.get("field") or ""): index for index, column in enumerate(export_columns, start=1)}
    sim_fill = PatternFill(fill_type="solid", fgColor=SIM_COLUMN_FILL_COLOR)
    sem_fill = PatternFill(fill_type="solid", fgColor=SEM_COLUMN_FILL_COLOR)
    sim_field_indexes = [
        field_to_index[field]
        for field in ("simWindowVolume", "simKd", "simCpc")
        if field in field_to_index
    ]
    sem_field_indexes = [
        field_to_index[field]
        for field in ("semVolume", "semKd", "semCpc")
        if field in field_to_index
    ]

    for row in standard_word_rows:
        export_row = []
        for column in export_columns:
            raw_value = row.get(column["field"])
            export_row.append(_transform_export_value(column, raw_value))
        keywords_sheet.append(export_row)

    for row_cells in keywords_sheet.iter_rows(min_row=2, max_row=keywords_sheet.max_row, min_col=1, max_col=len(headers)):
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row_index in range(2, keywords_sheet.max_row + 1):
        for col_index in sim_field_indexes:
            keywords_sheet.cell(row=row_index, column=col_index).fill = sim_fill
        for col_index in sem_field_indexes:
            keywords_sheet.cell(row=row_index, column=col_index).fill = sem_fill

    keywords_sheet.freeze_panes = "A2"
    keywords_sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(keywords_sheet.max_row, 1)}"

    for index, column in enumerate(export_columns, start=1):
        number_format = column.get("number_format")
        if not number_format:
            continue
        col_letter = get_column_letter(index)
        for cell in keywords_sheet[col_letter][1:]:
            cell.number_format = number_format

    for column_index, column in enumerate(export_columns, start=1):
        column_letter = get_column_letter(column_index)
        values = []
        for row_index in range(1, keywords_sheet.max_row + 1):
            value = keywords_sheet.cell(row=row_index, column=column_index).value
            values.append("" if value is None else str(value))
        keywords_sheet.column_dimensions[column_letter].width = _compute_column_width(values, column.get("width_profile"))

    for column_cells in summary_sheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        summary_sheet.column_dimensions[column_cells[0].column_letter].width = _compute_column_width(
            values,
            {"min": 14, "max": 64},
        )

    ensure_dir(output_path.parent)
    workbook.save(output_path)


def build_merged_standard_word_rows(site_results: List[dict]) -> dict:
    keyword_map: Dict[str, dict] = {}
    for site_result in site_results:
        for row in site_result.get("standardWordRows") or []:
            keyword = str(row.get("keyword") or "").strip()
            normalized_keyword = normalize_keyword_text(keyword)
            if not normalized_keyword:
                continue
            domains = stable_unique_texts(str(row.get("correspondingDomain") or "").split("|"))
            if normalized_keyword not in keyword_map:
                keyword_map[normalized_keyword] = _make_standard_word_row(keyword, " | ".join(domains))
                continue
            merged_domains = stable_unique_texts(
                keyword_map[normalized_keyword]["correspondingDomain"].split("|") + domains
            )
            keyword_map[normalized_keyword]["correspondingDomain"] = " | ".join(merged_domains)

    rows = list(keyword_map.values())
    rows.sort(key=lambda item: normalize_keyword_text(str(item.get("keyword") or "")))
    return {
        "rows": rows,
        "summary": {
            "rowCount": len(rows),
        },
    }


def write_merged_standard_word_table(*, stamp: str, site_results: List[dict], report_history_root: Path, report_latest_root: Path) -> Tuple[Path, Path]:
    history_path = report_history_root / f"keyword-table-{stamp}.xlsx"
    latest_path = report_latest_root / MERGED_LATEST_XLSX
    ensure_dir(history_path.parent)
    ensure_dir(latest_path.parent)
    write_standard_word_excel(stamp=stamp, site_results=site_results, output_path=history_path)
    shutil.copyfile(history_path, latest_path)
    return history_path, latest_path


def write_merged_report(
    *,
    stamp: str,
    site_results: List[dict],
    report_history_root: Path,
    report_latest_root: Path,
    final_standard_word_rows: Optional[List[dict]] = None,
) -> Tuple[Path, Path]:
    merged_history_dir = report_history_root
    merged_latest_path = report_latest_root / MERGED_LATEST_FILE

    ensure_dir(merged_history_dir)
    ensure_dir(merged_latest_path.parent)

    merged_text = render_merged_report(
        stamp=stamp,
        site_results=site_results,
        final_standard_word_rows=final_standard_word_rows,
    )
    history_path = merged_history_dir / f"report-{stamp}.md"

    history_path.write_text(merged_text, encoding="utf-8")
    merged_latest_path.write_text(merged_text, encoding="utf-8")

    return history_path, merged_latest_path


def _run_python_stage(*, stage_name: str, script_path: Path, extra_args: Sequence[str]) -> None:
    if not script_path.exists():
        raise RuntimeError(f"{stage_name} 脚本不存在: {script_path}")

    command = [sys.executable, str(script_path), *list(extra_args)]
    print(f"[stage] {stage_name}: {' '.join(command)}")

    try:
        subprocess.run(command, cwd=str(REPO_ROOT), check=True)
    except subprocess.CalledProcessError as exc:
        raise RuntimeError(f"{stage_name} 执行失败，退出码: {exc.returncode}") from exc


def _run_analyze_words_stage(*, seed_table_xlsx: Path, chain_root: Path) -> dict:
    stage_root = chain_root / "analyze-words"
    data_dir = stage_root / "data"
    snapshot_dir = stage_root / "_internal" / "snapshots"
    report_dir = stage_root / "report" / "history"
    latest_report_path = stage_root / "report" / "latest.md"
    latest_xlsx_path = stage_root / "report" / "latest.xlsx"
    latest_table_json_path = stage_root / "report" / "latest.json"

    _run_python_stage(
        stage_name="analyze-words",
        script_path=ANALYZE_WORDS_SCRIPT_PATH,
        extra_args=[
            "run",
            "--input-table",
            str(seed_table_xlsx),
            "--data-dir",
            str(data_dir),
            "--snapshot-dir",
            str(snapshot_dir),
            "--report-dir",
            str(report_dir),
            "--latest-report-path",
            str(latest_report_path),
            "--latest-xlsx-path",
            str(latest_xlsx_path),
            "--latest-table-json-path",
            str(latest_table_json_path),
        ],
    )

    if not latest_xlsx_path.exists():
        raise RuntimeError(f"analyze-words 未产出 latest xlsx: {latest_xlsx_path}")
    if not latest_table_json_path.exists():
        raise RuntimeError(f"analyze-words 未产出 latest json: {latest_table_json_path}")

    return {
        "latestReportPath": latest_report_path,
        "latestXlsxPath": latest_xlsx_path,
        "latestJsonPath": latest_table_json_path,
        "root": stage_root,
    }


def _run_check_gefei_kd_stage(*, input_table_xlsx: Path, chain_root: Path) -> dict:
    stage_root = chain_root / "check-gefei-kd"
    data_dir = stage_root / "data"
    snapshot_dir = stage_root / "_internal" / "snapshots"
    report_dir = stage_root / "report" / "history"
    latest_report_path = stage_root / "report" / "latest.md"
    latest_standard_word_table_json = stage_root / "report" / "latest.json"
    latest_standard_word_table_xlsx = stage_root / "report" / "latest.xlsx"

    _run_python_stage(
        stage_name="check-gefei-kd",
        script_path=CHECK_GEFEI_KD_SCRIPT_PATH,
        extra_args=[
            "run",
            "--standard-word-table",
            str(input_table_xlsx),
            "--data-dir",
            str(data_dir),
            "--snapshot-dir",
            str(snapshot_dir),
            "--report-dir",
            str(report_dir),
            "--latest-report-path",
            str(latest_report_path),
            "--latest-standard-word-table-json",
            str(latest_standard_word_table_json),
            "--latest-standard-word-table-xlsx",
            str(latest_standard_word_table_xlsx),
        ],
    )

    if not latest_standard_word_table_xlsx.exists():
        raise RuntimeError(f"check-gefei-kd 未产出 latest xlsx: {latest_standard_word_table_xlsx}")
    if not latest_standard_word_table_json.exists():
        raise RuntimeError(f"check-gefei-kd 未产出 latest json: {latest_standard_word_table_json}")

    return {
        "latestReportPath": latest_report_path,
        "latestXlsxPath": latest_standard_word_table_xlsx,
        "latestJsonPath": latest_standard_word_table_json,
        "root": stage_root,
    }


def _publish_final_words_xlsx(*, final_xlsx: Path, words_dir: Path, stamp: str) -> Path:
    if not final_xlsx.exists():
        raise RuntimeError(f"最终标准词表不存在: {final_xlsx}")

    ensure_dir(words_dir)
    target = words_dir / f"sitemap-{stamp}.xlsx"
    shutil.copyfile(final_xlsx, target)
    return target


def run_keyword_enrichment_chain(
    *,
    stamp: str,
    seed_table_xlsx: Path,
    chain_work_dir: Path,
    words_dir: Path,
    publish_stamp: Optional[str] = None,
) -> dict:
    chain_root = chain_work_dir / stamp
    ensure_dir(chain_root)

    analyze_outputs = _run_analyze_words_stage(
        seed_table_xlsx=seed_table_xlsx,
        chain_root=chain_root,
    )
    check_outputs = _run_check_gefei_kd_stage(
        input_table_xlsx=Path(analyze_outputs["latestXlsxPath"]),
        chain_root=chain_root,
    )

    final_words_xlsx_path = _publish_final_words_xlsx(
        final_xlsx=Path(check_outputs["latestXlsxPath"]),
        words_dir=words_dir,
        stamp=publish_stamp or stamp,
    )

    final_table_payload = load_json(Path(check_outputs["latestJsonPath"]))
    final_rows = final_table_payload.get("rows") if isinstance(final_table_payload, dict) else []
    if not isinstance(final_rows, list):
        final_rows = []

    return {
        "rows": final_rows,
        "finalWordsXlsxPath": final_words_xlsx_path,
        "chainStages": {
            "status": "completed",
            "chainRoot": str(chain_root),
            "analyzeWords": {
                "latestXlsxPath": str(analyze_outputs["latestXlsxPath"]),
                "latestJsonPath": str(analyze_outputs["latestJsonPath"]),
            },
            "checkGefeiKd": {
                "latestXlsxPath": str(check_outputs["latestXlsxPath"]),
                "latestJsonPath": str(check_outputs["latestJsonPath"]),
            },
        },
    }


def cleanup_legacy_report_layout(*, report_history_root: Path, report_latest_root: Path) -> None:
    legacy_latest_dir = report_latest_root / "latest"
    legacy_latest_merged = legacy_latest_dir / "merged.md"
    latest_path = report_latest_root / MERGED_LATEST_FILE

    if legacy_latest_merged.exists() and not latest_path.exists():
        latest_path.write_text(legacy_latest_merged.read_text(encoding="utf-8"), encoding="utf-8")

    legacy_merged_history = report_history_root / "merged"
    if legacy_merged_history.exists():
        for p in sorted(legacy_merged_history.glob("report-*.md")):
            target = report_history_root / p.name
            if not target.exists():
                shutil.move(str(p), str(target))

    if legacy_latest_dir.exists():
        shutil.rmtree(legacy_latest_dir, ignore_errors=True)
    if legacy_merged_history.exists():
        shutil.rmtree(legacy_merged_history, ignore_errors=True)


def purge_per_site_reports(*, site_ids: Sequence[str], report_history_root: Path, report_latest_root: Path) -> None:
    for site_id in site_ids:
        history_dir = report_history_root / site_id
        latest_path = report_latest_root / f"{site_id}.md"

        if history_dir.exists():
            for p in history_dir.glob("report-*.md"):
                p.unlink(missing_ok=True)
            shutil.rmtree(history_dir, ignore_errors=True)
        latest_path.unlink(missing_ok=True)


def list_snapshot_files(snapshot_dir: Path) -> List[Path]:
    if not snapshot_dir.exists():
        return []
    files = []
    for p in snapshot_dir.iterdir():
        if p.is_file() and SNAPSHOT_RE.match(p.name):
            files.append(p)
    files.sort(key=lambda p: p.name)
    return files


def find_latest_baseline(snapshot_dir: Path, current_stamp: str) -> Tuple[Optional[dict], Optional[Path]]:
    files = list_snapshot_files(snapshot_dir)
    for path in reversed(files):
        m = SNAPSHOT_RE.match(path.name)
        if not m:
            continue
        stamp = m.group(1)
        if stamp >= current_stamp:
            continue
        snapshot = load_json(path)
        return snapshot, path
    return None, None


def build_site_result_from_snapshot(site_id: str, site: Optional[dict], snapshot: Optional[dict]) -> dict:
    display_name = (site or {}).get("displayName") or site_id
    if not snapshot:
        return {
            "siteId": site_id,
            "siteDisplayName": display_name,
            "effectiveUrlCount": 0,
            "patternCount": 0,
            "newlyAddedCount": 0,
            "removedCount": 0,
            "newPatternCount": 0,
            "keywordCount": 0,
            "baselineMode": True,
            "baselineStamp": None,
            "newlyAddedUrls": [],
            "newPatterns": [],
            "topKeywords": [],
            "standardWordRows": [],
            "standardWordSummary": {"rowCount": 0},
        }

    meta = snapshot.get("meta") or {}
    comparison = snapshot.get("comparison") or {}
    keywords = snapshot.get("keywords") or {}
    standard_word = snapshot.get("standardWord") or {}

    newly = comparison.get("newlyAddedUrls") or []
    removed = comparison.get("removedUrls") or []
    new_patterns = comparison.get("newPatterns") or []
    top_keywords = keywords.get("topKeywordsFromNewUrls") or []
    standard_word_rows = standard_word.get("rows") or []
    standard_word_summary = standard_word.get("summary") or {"rowCount": len(standard_word_rows)}

    return {
        "siteId": site_id,
        "siteDisplayName": display_name,
        "effectiveUrlCount": safe_int(meta.get("effectiveUrlCount")),
        "patternCount": safe_int(meta.get("patternCount")),
        "newlyAddedCount": len(newly),
        "removedCount": len(removed),
        "newPatternCount": len(new_patterns),
        "keywordCount": safe_int(standard_word_summary.get("rowCount", len(standard_word_rows))),
        "baselineMode": bool(meta.get("baselineMode", False)),
        "baselineStamp": meta.get("baselineStamp"),
        "newlyAddedUrls": newly,
        "newPatterns": new_patterns,
        "topKeywords": top_keywords,
        "standardWordRows": standard_word_rows,
        "standardWordSummary": standard_word_summary,
    }


def load_latest_snapshot_for_site(snapshot_root: Path, site_id: str) -> Tuple[Optional[dict], Optional[str]]:
    snapshot_dir = snapshot_root / site_id
    files = list_snapshot_files(snapshot_dir)
    if not files:
        return None, None

    path = files[-1]
    snapshot = load_json(path)
    meta = snapshot.get("meta") or {}
    stamp = meta.get("stamp")
    if not isinstance(stamp, str) or not stamp:
        m = SNAPSHOT_RE.match(path.name)
        stamp = m.group(1) if m else None
    return snapshot, stamp


def run_single_site(
    *,
    site: dict,
    stamp: str,
    data_root: Path,
    snapshot_root: Path,
    report_latest_root: Path,
) -> dict:
    site_id = site["id"]

    site_data_dir = data_root / site_id
    site_snapshot_dir = snapshot_root / site_id
    site_latest_path = report_latest_root / f"{site_id}.md"

    ensure_dir(site_data_dir)
    ensure_dir(site_snapshot_dir)
    ensure_dir(site_latest_path.parent)

    baseline_snapshot, baseline_path = find_latest_baseline(site_snapshot_dir, stamp)
    baseline_mode = baseline_snapshot is None
    keyword_rules = site.get("keywordRules") or {}

    raw_urls, crawl_meta = crawl_sitemaps(site)
    rows, normalize_stats = filter_and_normalize_urls(raw_urls, site)
    patterns = infer_patterns(rows)

    if baseline_mode:
        comparison = {
            "newlyAddedUrls": [],
            "removedUrls": [],
            "newPatterns": [],
        }
        standard_word_rows: List[dict] = []
        standard_word_summary = {"rowCount": 0}
        keyword_result = {
            "topKeywordsFromNewUrls": [],
            "allKeywordsFromNewUrls": [],
            "counters": {"phrase": [], "bigram": [], "token": []},
        }
        baseline_stamp = None
    else:
        baseline_rows = (baseline_snapshot.get("urls") or [])
        baseline_patterns = (baseline_snapshot.get("patterns") or [])

        comparison = build_comparison(rows, baseline_rows, patterns, baseline_patterns)
        keyword_result = extract_keywords(comparison.get("newlyAddedUrls") or [], keyword_rules)
        standard_word_rows, standard_word_summary = build_standard_word_rows(
            keyword_result.get("allKeywordsFromNewUrls") or [],
            comparison.get("newlyAddedUrls") or [],
            keyword_rules,
        )
        baseline_stamp = ((baseline_snapshot.get("meta") or {}).get("stamp") if baseline_snapshot else None)

    keyword_result["standardWord"] = {
        "rows": standard_word_rows,
        "summary": standard_word_summary,
    }

    snapshot = build_snapshot(
        stamp=stamp,
        site=site,
        crawl_meta=crawl_meta,
        normalize_stats=normalize_stats,
        rows=rows,
        patterns=patterns,
        baseline_stamp=baseline_stamp,
        baseline_mode=baseline_mode,
        comparison=comparison,
        keyword_result=keyword_result,
    )

    fetch_archive_path = site_data_dir / f"fetch-{stamp}.json"
    snapshot_path = site_snapshot_dir / f"snapshot-{stamp}.json"

    dump_json(
        fetch_archive_path,
        {
            "meta": {
                "generatedAt": datetime.now().isoformat(timespec="seconds"),
                "stamp": stamp,
                "siteId": site_id,
            },
            "crawl": crawl_meta,
            "rawUrls": raw_urls,
        },
    )
    dump_json(snapshot_path, snapshot)

    site_latest_path.unlink(missing_ok=True)

    return {
        "siteId": site_id,
        "siteDisplayName": site.get("displayName") or site_id,
        "fetchArchive": fetch_archive_path,
        "snapshot": snapshot_path,
        "reportHistory": None,
        "reportLatest": None,
        "effectiveUrlCount": len(rows),
        "patternCount": len(patterns),
        "newlyAddedCount": len(comparison.get("newlyAddedUrls") or []),
        "removedCount": len(comparison.get("removedUrls") or []),
        "newPatternCount": len(comparison.get("newPatterns") or []),
        "keywordCount": len((keyword_result.get("topKeywordsFromNewUrls") or [])),
        "baselineMode": baseline_mode,
        "baselineStamp": baseline_stamp,
        "baselinePath": baseline_path,
        "newlyAddedUrls": comparison.get("newlyAddedUrls") or [],
        "newPatterns": comparison.get("newPatterns") or [],
        "topKeywords": keyword_result.get("topKeywordsFromNewUrls") or [],
        "standardWordRows": standard_word_rows,
        "standardWordSummary": standard_word_summary,
    }


def run_pipeline(args: argparse.Namespace) -> int:
    sites_config = Path(args.sites_config).resolve()
    data_root = Path(args.data_root).resolve()
    snapshot_root = Path(args.snapshot_root).resolve()
    report_history_root = Path(args.report_history_root).resolve()
    report_latest_root = Path(args.report_latest_root).resolve()
    words_dir = Path(args.words_dir).resolve()
    chain_work_dir = Path(args.chain_work_dir).resolve()

    ensure_dir(data_root)
    ensure_dir(snapshot_root)
    ensure_dir(report_history_root)
    ensure_dir(report_latest_root)
    ensure_dir(words_dir)

    cleanup_legacy_report_layout(report_history_root=report_history_root, report_latest_root=report_latest_root)

    sites = load_sites_config(sites_config)

    if args.site:
        if args.site not in sites:
            raise SystemExit(f"未找到 site: {args.site}")
        selected = [sites[args.site]]
    elif args.all_sites:
        selected = [site for site in sites.values()]
    else:
        selected = [site for site in sites.values() if site.get("enabled", True)]

    if not selected:
        raise SystemExit("没有可运行的站点：请检查 sites.json 的 enabled 配置")

    stamp = now_stamp()

    results = []
    for site in selected:
        site_id = site["id"]
        print(f"[run] site={site_id} ...")
        result = run_single_site(
            site=site,
            stamp=stamp,
            data_root=data_root,
            snapshot_root=snapshot_root,
            report_latest_root=report_latest_root,
        )
        results.append(result)

        print(f"[done] site          : {result['siteId']}")
        print(f"[done] fetch archive : {result['fetchArchive']}")
        print(f"[done] snapshot      : {result['snapshot']}")
        print(f"[done] effective urls: {result['effectiveUrlCount']}")
        print(f"[done] newly added   : {result['newlyAddedCount']}")
        print(f"[done] new patterns  : {result['newPatternCount']}")
        print(f"[done] keywords      : {result['keywordCount']}")
        if result["baselineMode"]:
            print("[done] baseline      : none (first run)")
        else:
            print(f"[done] baseline      : {result['baselinePath']}")

    excel_history_path, excel_latest_path = write_merged_standard_word_table(
        stamp=stamp,
        site_results=results,
        report_history_root=report_history_root,
        report_latest_root=report_latest_root,
    )

    merged_standard_word_rows = build_merged_standard_word_rows(results).get("rows") or []
    final_standard_word_rows = merged_standard_word_rows

    keyword_row_count = len(merged_standard_word_rows)
    if keyword_row_count > 0:
        chain_result = run_keyword_enrichment_chain(
            stamp=stamp,
            seed_table_xlsx=excel_history_path,
            chain_work_dir=chain_work_dir,
            words_dir=words_dir,
        )
        final_words_xlsx_path = Path(chain_result["finalWordsXlsxPath"])
        candidate_rows = chain_result.get("rows")
        if isinstance(candidate_rows, list):
            final_standard_word_rows = candidate_rows
    else:
        final_words_xlsx_path = _publish_final_words_xlsx(
            final_xlsx=excel_history_path,
            words_dir=words_dir,
            stamp=stamp,
        )

    shutil.copyfile(final_words_xlsx_path, excel_history_path)
    shutil.copyfile(final_words_xlsx_path, excel_latest_path)

    merged_history_path, merged_latest_path = write_merged_report(
        stamp=stamp,
        site_results=results,
        report_history_root=report_history_root,
        report_latest_root=report_latest_root,
        final_standard_word_rows=final_standard_word_rows,
    )

    purge_per_site_reports(
        site_ids=[site["id"] for site in selected],
        report_history_root=report_history_root,
        report_latest_root=report_latest_root,
    )

    print(f"[done] merged history: {merged_history_path}")
    print(f"[done] merged latest : {merged_latest_path}")
    print(f"[done] excel history : {excel_history_path}")
    print(f"[done] excel latest  : {excel_latest_path}")
    print(f"[done] words final   : {final_words_xlsx_path}")
    print(f"[done] sites total: {len(results)}")
    return 0
def rebuild_single_site_reports(
    *,
    site: Optional[dict],
    site_id: str,
    snapshot_root: Path,
    report_history_root: Path,
    report_latest_root: Path,
) -> None:
    snapshot_dir = snapshot_root / site_id
    history_dir = report_history_root / site_id
    latest_path = report_latest_root / f"{site_id}.md"

    ensure_dir(snapshot_dir)
    ensure_dir(latest_path.parent)

    files = list_snapshot_files(snapshot_dir)
    if not files:
        print(f"[skip] site={site_id} 无快照: {snapshot_dir}")
        if history_dir.exists():
            for p in history_dir.glob("report-*.md"):
                p.unlink(missing_ok=True)
        latest_path.unlink(missing_ok=True)
        return

    keyword_rules = (site or {}).get("keywordRules") or {}
    baseline_snapshot: Optional[dict] = None

    for path in files:
        snapshot = load_json(path)

        today_rows = snapshot.get("urls") or []
        today_patterns = snapshot.get("patterns") or []

        if baseline_snapshot is None:
            baseline_mode = True
            baseline_stamp = None
            comparison = {
                "newlyAddedUrls": [],
                "removedUrls": [],
                "newPatterns": [],
            }
            standard_word_rows = []
            standard_word_summary = {"rowCount": 0}
            keyword_result = {
                "topKeywordsFromNewUrls": [],
                "allKeywordsFromNewUrls": [],
                "counters": {"phrase": [], "bigram": [], "token": []},
            }
        else:
            baseline_mode = False
            baseline_stamp = (baseline_snapshot.get("meta") or {}).get("stamp")
            comparison = build_comparison(
                today_rows,
                baseline_snapshot.get("urls") or [],
                today_patterns,
                baseline_snapshot.get("patterns") or [],
            )
            keyword_result = extract_keywords(comparison.get("newlyAddedUrls") or [], keyword_rules)
            standard_word_rows, standard_word_summary = build_standard_word_rows(
                keyword_result.get("allKeywordsFromNewUrls") or [],
                comparison.get("newlyAddedUrls") or [],
                keyword_rules,
            )

        keyword_result["standardWord"] = {
            "rows": standard_word_rows,
            "summary": standard_word_summary,
        }

        meta = snapshot.get("meta") or {}
        meta["baselineMode"] = baseline_mode
        meta["baselineStamp"] = baseline_stamp
        meta["standardWordTableVersion"] = STANDARD_WORD_TABLE_VERSION
        snapshot["meta"] = meta
        snapshot["comparison"] = {
            "baselineStamp": baseline_stamp,
            "newlyAddedUrls": comparison.get("newlyAddedUrls") or [],
            "removedUrls": comparison.get("removedUrls") or [],
            "newPatterns": comparison.get("newPatterns") or [],
        }
        snapshot["keywords"] = {
            "topKeywordsFromNewUrls": keyword_result.get("topKeywordsFromNewUrls") or [],
            "counters": keyword_result.get("counters") or {},
        }
        snapshot["standardWord"] = {
            "rows": standard_word_rows,
            "summary": standard_word_summary,
        }

        dump_json(path, snapshot)
        baseline_snapshot = snapshot

    if history_dir.exists():
        for p in history_dir.glob("report-*.md"):
            p.unlink(missing_ok=True)
    latest_path.unlink(missing_ok=True)


def rebuild_reports(args: argparse.Namespace) -> int:
    sites_config = Path(args.sites_config).resolve()
    snapshot_root = Path(args.snapshot_root).resolve()
    report_history_root = Path(args.report_history_root).resolve()
    report_latest_root = Path(args.report_latest_root).resolve()
    words_dir = Path(args.words_dir).resolve()
    chain_work_dir = Path(args.chain_work_dir).resolve()

    ensure_dir(snapshot_root)
    ensure_dir(report_history_root)
    ensure_dir(report_latest_root)
    ensure_dir(words_dir)

    cleanup_legacy_report_layout(report_history_root=report_history_root, report_latest_root=report_latest_root)

    sites = load_sites_config(sites_config)

    if args.site:
        if args.site not in sites:
            raise SystemExit(f"未找到 site: {args.site}")
        selected_ids = [args.site]
    else:
        selected_ids = sorted(set(list(sites.keys()) + [p.name for p in snapshot_root.iterdir() if p.is_dir()]))

    for site_id in selected_ids:
        site = sites.get(site_id)
        rebuild_single_site_reports(
            site=site,
            site_id=site_id,
            snapshot_root=snapshot_root,
            report_history_root=report_history_root,
            report_latest_root=report_latest_root,
        )

    merged_results: List[dict] = []
    merged_stamps: List[str] = []

    merged_site_ids = sorted(sites.keys())
    for site_id in merged_site_ids:
        site = sites.get(site_id)
        latest_snapshot, stamp = load_latest_snapshot_for_site(snapshot_root, site_id)
        if stamp:
            merged_stamps.append(stamp)
        merged_results.append(build_site_result_from_snapshot(site_id, site, latest_snapshot))

    merged_stamp = max(merged_stamps) if merged_stamps else now_stamp()
    publish_stamp = now_stamp()
    excel_history_path, excel_latest_path = write_merged_standard_word_table(
        stamp=merged_stamp,
        site_results=merged_results,
        report_history_root=report_history_root,
        report_latest_root=report_latest_root,
    )

    merged_standard_word_rows = build_merged_standard_word_rows(merged_results).get("rows") or []
    final_standard_word_rows = merged_standard_word_rows

    keyword_row_count = len(merged_standard_word_rows)
    if keyword_row_count > 0:
        chain_result = run_keyword_enrichment_chain(
            stamp=merged_stamp,
            seed_table_xlsx=excel_history_path,
            chain_work_dir=chain_work_dir,
            words_dir=words_dir,
            publish_stamp=publish_stamp,
        )
        final_words_xlsx_path = Path(chain_result["finalWordsXlsxPath"])
        candidate_rows = chain_result.get("rows")
        if isinstance(candidate_rows, list):
            final_standard_word_rows = candidate_rows
    else:
        final_words_xlsx_path = _publish_final_words_xlsx(
            final_xlsx=excel_history_path,
            words_dir=words_dir,
            stamp=publish_stamp,
        )

    shutil.copyfile(final_words_xlsx_path, excel_history_path)
    shutil.copyfile(final_words_xlsx_path, excel_latest_path)

    merged_history_path, merged_latest_path = write_merged_report(
        stamp=merged_stamp,
        site_results=merged_results,
        report_history_root=report_history_root,
        report_latest_root=report_latest_root,
        final_standard_word_rows=final_standard_word_rows,
    )
    print(f"[done] rebuilt merged history: {merged_history_path}")
    print(f"[done] rebuilt merged latest : {merged_latest_path}")
    print(f"[done] rebuilt excel history : {excel_history_path}")
    print(f"[done] rebuilt excel latest  : {excel_latest_path}")
    print(f"[done] words final           : {final_words_xlsx_path}")

    return 0


def validate_report(args: argparse.Namespace) -> int:
    report_path = Path(args.report).resolve()
    xlsx_path = Path(args.xlsx).resolve()

    if not report_path.exists():
        raise SystemExit(f"报告不存在: {report_path}")

    text = report_path.read_text(encoding="utf-8")

    missing = [sec for sec in REQUIRED_SECTIONS if sec not in text]
    missing += [line for line in REQUIRED_REMARKS if line not in text]

    if missing:
        print("[failed] 报告校验失败，缺少内容：")
        for item in missing:
            print(f"- {item}")
        return 1

    if not xlsx_path.exists():
        raise SystemExit(f"Excel 不存在: {xlsx_path}")

    _Workbook, _Font, _Alignment, _PatternFill, _get_column_letter, load_workbook = _require_openpyxl()
    workbook = load_workbook(xlsx_path, data_only=True)
    try:
        if "keywords" not in workbook.sheetnames:
            raise SystemExit("Excel 缺少 keywords sheet")
        sheet = workbook["keywords"]

        expected_headers = get_keywords_export_headers()
        actual_headers = [sheet.cell(row=1, column=index).value for index in range(1, sheet.max_column + 1)]
        if actual_headers != expected_headers:
            print("Excel 表头不匹配：")
            max_len = max(len(expected_headers), len(actual_headers))
            for index in range(1, max_len + 1):
                expected = expected_headers[index - 1] if index <= len(expected_headers) else None
                actual = actual_headers[index - 1] if index <= len(actual_headers) else None
                if expected != actual:
                    print(f"- 第 {index} 列 expected={expected!r} actual={actual!r}")
            raise SystemExit(1)

        keyword_col = expected_headers.index(get_header_for_field("keyword")) + 1
        domain_col = expected_headers.index(get_header_for_field("correspondingDomain")) + 1
        score_col = get_score_column_index_from_headers(expected_headers) + 1
        numeric_fields = [
            "simWindowVolume",
            "simKd",
            "simCpc",
            "semVolume",
            "semKd",
            "semCpc",
            "gefeiKD",
        ]
        numeric_indexes = [expected_headers.index(get_header_for_field(field)) + 1 for field in numeric_fields]

        def _is_numeric_cell(value) -> bool:
            return isinstance(value, (int, float)) and not isinstance(value, bool)

        for row_index in range(2, sheet.max_row + 1):
            keyword = sheet.cell(row=row_index, column=keyword_col).value
            domain = sheet.cell(row=row_index, column=domain_col).value
            score_value = sheet.cell(row=row_index, column=score_col).value
            if keyword in (None, ""):
                raise SystemExit(f"keyword 列为空（row={row_index}）")
            if domain in (None, ""):
                raise SystemExit(f"对应域名 列为空（row={row_index}）")

            if score_value not in (None, "") and not _is_numeric_cell(score_value):
                raise SystemExit(f"score 列应为数值类型而非文本（row={row_index}）")

            for col_index in numeric_indexes:
                value = sheet.cell(row=row_index, column=col_index).value
                if value in (None, ""):
                    continue
                if not _is_numeric_cell(value):
                    raise SystemExit(f"数字列应为数值类型而非文本（row={row_index}, col={col_index}）")
    finally:
        workbook.close()

    print(f"[ok] report validated: {report_path}")
    print(f"[ok] xlsx exists      : {xlsx_path}")
    print("[ok] xlsx headers / keyword / 对应域名 validated")
    return 0


def main() -> int:
    args = parse_args()
    if args.command == "run":
        return run_pipeline(args)
    if args.command == "rebuild-reports":
        return rebuild_reports(args)
    if args.command == "validate-report":
        return validate_report(args)
    raise SystemExit(f"未知命令: {args.command}")


if __name__ == "__main__":
    raise SystemExit(main())
