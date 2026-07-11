#!/usr/bin/env python3
"""批量查询哥飞 KD API，输出快照、Markdown 报告与标准词表。

输入可以是任意形式的关键词列表：

- 自由关键词：`--keyword`（可重复）/ `--keyword-file`（每行一个，支持 # 注释）/ `--keywords-csv`
- 标准词表：`--standard-word-table <path>`（`.json` 或 `.xlsx`，按 standard-word-analysis v1 规范解析）

无论哪种输入，`run` 都会：

1. 抽取去重关键词
2. 调用哥飞 KD API（复用 `shared_gefei_kd.fetch_gefei_kd_rows`）
3. 归档原始抓取结果到 `data/fetch-*.json`
4. 标准化写入快照到 `_internal/snapshots/snapshot-*.json`
5. 生成 Markdown 报告（history + latest）
6. 产出标准词表（JSON + XLSX），填入每个词的 `gefeiKD`

当输入本身就是标准词表时，仅回填 `gefeiKD`，其余字段保持原值不动。

Usage:
  python3 check-gefei-kd/_internal/scripts/check_gefei_kd.py run --keyword "image to text"
  python3 check-gefei-kd/_internal/scripts/check_gefei_kd.py run --keywords-csv "image to text,ocr online"
  python3 check-gefei-kd/_internal/scripts/check_gefei_kd.py run --standard-word-table path/to/table.json
  python3 check-gefei-kd/_internal/scripts/check_gefei_kd.py rebuild-reports
  python3 check-gefei-kd/_internal/scripts/check_gefei_kd.py validate-report
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from copy import deepcopy
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple
from urllib.parse import urlencode

# 仓库根目录：scripts -> _internal -> check-gefei-kd -> repo root
REPO_ROOT = Path(__file__).resolve().parents[3]
PROJECT_DIR = Path(__file__).resolve().parents[2]

# 复用仓库根目录下的共享哥飞 KD 查询实现
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))
import shared_gefei_kd  # noqa: E402

DEFAULT_API_BASE = shared_gefei_kd.DEFAULT_API_BASE
DEFAULT_ENDPOINT = shared_gefei_kd.DEFAULT_ENDPOINT
DEFAULT_API_KEY_ENV = shared_gefei_kd.DEFAULT_API_KEY_ENV
DEFAULT_API_KEY_FILE = PROJECT_DIR / "api_key.txt"

STANDARD_WORD_TABLE_VERSION = "v1"
STANDARD_WORD_TABLE_SPEC_PATH = (
    REPO_ROOT / "standard-word-analysis" / "spec" / f"standard-word-table.{STANDARD_WORD_TABLE_VERSION}.json"
)
SCORE_FORMULA = "simWindowVolume * simCpc / simKd"

SNAPSHOT_RE = re.compile(r"^snapshot-(\d{8}-\d{6})\.json$")
SPACE_RE = re.compile(r"\s+")

REQUEST_TIMEOUT_SECONDS = 60
TOP_PREVIEW_ROWS = 200

REQUIRED_SECTIONS = [
    "## 摘要",
    "## 请求参数",
    "## 关键词结果明细",
    "## 失败明细",
    "## 标准词表",
    "## 产物路径",
    "## 备注",
]

REQUIRED_REMARKS = [
    "API 默认使用 Header 鉴权（Authorization: Bearer <token>），可通过参数切换为 query token。",
    "MCP/API 频率限制：每分钟最多约 10 次请求，默认最小请求间隔为 6.2 秒。",
    "7 天内同词命中缓存会更快返回，但仍计入当日额度。",
    "当输入为标准词表时，仅回填 gefeiKD，其余字段保持原值不动。",
]


# --------------------------------------------------------------------------- #
# 参数解析
# --------------------------------------------------------------------------- #
def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="批量查询哥飞 KD API 并产出标准词表")
    subparsers = parser.add_subparsers(dest="command", required=True)

    run_parser = subparsers.add_parser("run", help="抓取 + 快照 + 报告 + 标准词表")
    run_parser.add_argument("--keyword", action="append", default=[], help="关键词，可重复")
    run_parser.add_argument(
        "--keyword-file",
        action="append",
        default=[],
        help="关键词文件（UTF-8，每行一个关键词，支持 # 注释）",
    )
    run_parser.add_argument(
        "--keywords-csv",
        action="append",
        default=[],
        help="逗号分隔关键词列表，可重复",
    )
    run_parser.add_argument(
        "--standard-word-table",
        action="append",
        default=[],
        help="标准词表文件（.json 或 .xlsx，按 standard-word-analysis v1 解析），可重复；仅回填 gefeiKD",
    )

    run_parser.add_argument("--api-base", default=DEFAULT_API_BASE)
    run_parser.add_argument("--endpoint", default=DEFAULT_ENDPOINT)
    run_parser.add_argument("--api-key", default="", help="直接传入 API Key（优先级最高）")
    run_parser.add_argument("--api-key-env", default=DEFAULT_API_KEY_ENV)
    run_parser.add_argument("--api-key-file", default=str(DEFAULT_API_KEY_FILE))
    run_parser.add_argument("--auth-mode", choices=["header", "query"], default="header")

    run_parser.add_argument("--gl", default="us")
    run_parser.add_argument("--hl", default="en")
    run_parser.add_argument("--force", type=int, choices=[0, 1], default=0)
    run_parser.add_argument("--response-format", default="json", choices=["json"])

    run_parser.add_argument("--min-interval-seconds", type=float, default=6.2)
    run_parser.add_argument("--max-retries", type=int, default=2)
    run_parser.add_argument("--timeout-seconds", type=int, default=REQUEST_TIMEOUT_SECONDS)
    run_parser.add_argument(
        "--max-keywords",
        type=int,
        default=0,
        help="最多分析前 N 个关键词（0 表示不限制）",
    )
    run_parser.add_argument("--dry-run", action="store_true")
    run_parser.add_argument(
        "--no-xlsx",
        action="store_true",
        help="跳过 XLSX 标准词表导出（当 openpyxl 不可用时仍可产出 JSON）",
    )

    run_parser.add_argument("--data-dir", default=str(PROJECT_DIR / "data"))
    run_parser.add_argument("--snapshot-dir", default=str(PROJECT_DIR / "_internal" / "snapshots"))
    run_parser.add_argument("--report-dir", default=str(PROJECT_DIR / "report" / "history"))
    run_parser.add_argument("--latest-report-path", default=str(PROJECT_DIR / "report" / "latest.md"))
    run_parser.add_argument(
        "--latest-standard-word-table-json",
        default=str(PROJECT_DIR / "report" / "latest.json"),
    )
    run_parser.add_argument(
        "--latest-standard-word-table-xlsx",
        default=str(PROJECT_DIR / "report" / "latest.xlsx"),
    )

    rebuild_parser = subparsers.add_parser("rebuild-reports", help="根据快照重建历史报告与标准词表")
    rebuild_parser.add_argument("--snapshot-dir", default=str(PROJECT_DIR / "_internal" / "snapshots"))
    rebuild_parser.add_argument("--report-dir", default=str(PROJECT_DIR / "report" / "history"))
    rebuild_parser.add_argument("--latest-report-path", default=str(PROJECT_DIR / "report" / "latest.md"))
    rebuild_parser.add_argument(
        "--latest-standard-word-table-json",
        default=str(PROJECT_DIR / "report" / "latest.json"),
    )
    rebuild_parser.add_argument(
        "--latest-standard-word-table-xlsx",
        default=str(PROJECT_DIR / "report" / "latest.xlsx"),
    )
    rebuild_parser.add_argument("--no-xlsx", action="store_true")

    validate_parser = subparsers.add_parser("validate-report", help="校验 Markdown 报告结构")
    validate_parser.add_argument("--report", default=str(PROJECT_DIR / "report" / "latest.md"))
    validate_parser.add_argument(
        "--xlsx",
        default=str(PROJECT_DIR / "report" / "latest.xlsx"),
        help="标准词表 Excel 路径（默认 report/latest.xlsx）",
    )

    return parser.parse_args()


# --------------------------------------------------------------------------- #
# 通用工具
# --------------------------------------------------------------------------- #
def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d-%H%M%S")


def load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def dump_json(path: Path, payload: dict) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def normalize_keyword_text(text: str) -> str:
    return SPACE_RE.sub(" ", str(text or "").strip().lower())


def mask_secret(secret: str) -> str:
    text = str(secret or "")
    if len(text) <= 10:
        return "*" * len(text)
    return f"{text[:6]}****{text[-4:]}"


def safe_float(value) -> float:
    try:
        if value is None or value == "":
            return 0.0
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def to_display_number(value, digits: int = 2) -> str:
    if value is None:
        return "-"
    number = safe_float(value)
    if abs(number - round(number)) < 1e-9:
        return str(int(round(number)))
    return f"{number:.{digits}f}"


def _get_nested(obj: dict, *keys: str):
    current = obj
    for key in keys:
        if not isinstance(current, dict):
            return None
        current = current.get(key)
    return current


def _value_to_short_text(value, max_len: int = 80) -> str:
    if value is None:
        return "-"
    if isinstance(value, (int, float, bool)):
        return str(value)
    if isinstance(value, str):
        text = value.strip()
    else:
        text = json.dumps(value, ensure_ascii=False)
    text = SPACE_RE.sub(" ", text).strip()
    if len(text) <= max_len:
        return text
    return text[: max_len - 1] + "…"


# --------------------------------------------------------------------------- #
# 标准词表规范
# --------------------------------------------------------------------------- #
def load_standard_word_table_spec() -> dict:
    if not STANDARD_WORD_TABLE_SPEC_PATH.exists():
        raise RuntimeError(
            f"标准词表规范缺失: {STANDARD_WORD_TABLE_SPEC_PATH}。"
            "请先补齐 standard-word-analysis/spec/standard-word-table.v1.json。"
        )
    spec = load_json(STANDARD_WORD_TABLE_SPEC_PATH)
    if not isinstance(spec, dict):
        raise RuntimeError("标准词表规范格式错误：顶层必须是对象")
    version = str(spec.get("version") or "").strip()
    if version != STANDARD_WORD_TABLE_VERSION:
        raise RuntimeError(f"标准词表版本不匹配：expected={STANDARD_WORD_TABLE_VERSION} actual={version or '-'}")
    columns = spec.get("excelColumns")
    if not isinstance(columns, list) or not columns:
        raise RuntimeError("标准词表规范缺少 excelColumns 数组")
    score_header = str(spec.get("scoreColumnHeader") or "").strip()
    if not score_header:
        raise RuntimeError("标准词表规范缺少 scoreColumnHeader")
    return {
        "version": version,
        "columns": columns,
        "scoreColumnHeader": score_header,
        "scoreField": str(spec.get("scoreField") or "score"),
        "scoreFormula": str(spec.get("scoreFormula") or SCORE_FORMULA),
    }


def spec_columns() -> List[dict]:
    return load_standard_word_table_spec()["columns"]


def spec_header_to_field() -> Dict[str, str]:
    return {str(col.get("header")): str(col.get("field")) for col in spec_columns()}


def get_header_for_field(field_name: str) -> str:
    for column in spec_columns():
        if str(column.get("field")) == field_name:
            return str(column.get("header") or "")
    raise RuntimeError(f"标准词表缺少字段: {field_name}")


def get_score_column_index_from_headers(headers: Sequence[str]) -> int:
    score_header = str(load_standard_word_table_spec().get("scoreColumnHeader") or "").strip()
    if not score_header:
        score_header = "score(simWindowVolume*cpc/kd)"
    try:
        return list(headers).index(score_header)
    except ValueError as exc:
        raise RuntimeError(f"标准词表缺少 score 列: {score_header}") from exc



def empty_standard_word_row(keyword: str) -> dict:
    """自由关键词输入时，构造一个仅含 keyword、其余字段为 null 的标准词表行。"""
    row = {str(col.get("field")): None for col in spec_columns()}
    row["keyword"] = str(keyword or "")
    # 对应域名 / group 默认空串，避免与 word-from-root 输出风格不一致
    row.setdefault("correspondingDomain", "")
    row["correspondingDomain"] = row.get("correspondingDomain") or ""
    row.setdefault("group", "")
    row["group"] = row.get("group") or ""
    return row


# --------------------------------------------------------------------------- #
# 输入解析
# --------------------------------------------------------------------------- #
def _read_keywords_from_file(path: Path) -> List[str]:
    if not path.exists():
        raise RuntimeError(f"关键词文件不存在: {path}")
    lines = path.read_text(encoding="utf-8").splitlines()
    keywords: List[str] = []
    for raw in lines:
        text = str(raw).strip()
        if not text or text.startswith("#"):
            continue
        keywords.append(text)
    return keywords


def _require_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "缺少依赖 openpyxl。请先执行 `pip3 install openpyxl` 再运行。"
        ) from exc
    return Workbook, Font, Alignment, PatternFill, get_column_letter, load_workbook


MULTILINE_EXPORT_FIELDS = {
    "group",
    "correspondingDomain",
}


def _format_multiline_export_text(field: str, value):
    if value is None:
        return value
    if field not in MULTILINE_EXPORT_FIELDS:
        return value
    text = str(value).strip()
    if not text:
        return text
    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    normalized = re.sub(r"\s*\|\s*", "\n", normalized)
    normalized = normalized.replace(" / ", "\n")
    lines = [line.strip() for line in normalized.split("\n") if line.strip()]
    return "\n".join(lines) if lines else ""


def _coerce_cell(value, field: str):
    """将 XLSX 单元格值还原为标准词表字段类型（数字列转 float/None）。"""
    if value is None or value == "":
        return None
    number_fields = {
        "score",
        "simWindowVolume",
        "simKd",
        "simCpc",
        "semVolume",
        "semKd",
        "semCpc",
        "gefeiKD",
    }
    if field in number_fields:
        try:
            f = float(value)
            return f
        except (TypeError, ValueError):
            return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value



def read_standard_word_table_file(path: Path) -> List[dict]:
    """读取标准词表文件（.json 或 .xlsx），返回按 spec 字段对齐的行列表。"""
    if not path.exists():
        raise RuntimeError(f"标准词表文件不存在: {path}")

    suffix = path.suffix.lower()
    header_to_field = spec_header_to_field()
    fields = [str(col.get("field")) for col in spec_columns()]

    if suffix == ".json":
        payload = load_json(path)
        raw_rows = []
        if isinstance(payload, dict):
            raw_rows = payload.get("rows") or []
        elif isinstance(payload, list):
            raw_rows = payload
        if not isinstance(raw_rows, list):
            raise RuntimeError(f"标准词表 JSON 缺少 rows 数组: {path}")

        rows: List[dict] = []
        for raw in raw_rows:
            if not isinstance(raw, dict):
                continue
            row = {field: raw.get(field) for field in fields}
            # 保留原始行中存在但不在 spec 内的额外字段（尽量不丢信息）
            for key, value in raw.items():
                if key not in row:
                    row[key] = value
            rows.append(row)
        return rows

    if suffix in (".xlsx", ".xlsm"):
        _, _, _, _, _, load_workbook = _require_openpyxl()
        wb = load_workbook(path, data_only=True)
        # 优先使用 keywords 表，其次第一个 sheet
        ws = wb["keywords"] if "keywords" in wb.sheetnames else wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []
        headers = [str(h).strip() if h is not None else "" for h in header_row]
        col_fields = [header_to_field.get(h, "") for h in headers]

        rows: List[dict] = []
        for record in rows_iter:
            if record is None or all(v is None or v == "" for v in record):
                continue
            row = {field: None for field in fields}
            for field, value in zip(col_fields, record):
                if not field:
                    continue
                row[field] = _coerce_cell(value, field)
            rows.append(row)
        return rows

    raise RuntimeError(f"不支持的标准词表格式: {path}（仅支持 .json / .xlsx）")


def collect_input(args: argparse.Namespace) -> dict:
    """收集输入，返回 {mode, keywords, standard_rows}。

    - mode: "free_keywords" / "standard_word_table"
    - keywords: 去重后的关键词条目列表 [{keyword, keywordNormalized}]
    - standard_rows: 标准词表行（自由输入时为仅 keyword 的骨架行；标准词表输入时为原值行）
    - input_paths: 输入文件路径列表（用于 meta 记录）
    """
    standard_rows: List[dict] = []
    input_paths: List[str] = []
    standard_input_used = False

    for table_arg in args.standard_word_table or []:
        path = Path(table_arg).resolve()
        input_paths.append(str(path))
        standard_input_used = True
        rows = read_standard_word_table_file(path)
        for row in rows:
            standard_rows.append(deepcopy(row))

    raw_keywords: List[str] = []
    for text in args.keyword or []:
        value = str(text or "").strip()
        if value:
            raw_keywords.append(value)
    for file_arg in args.keyword_file or []:
        raw_keywords.extend(_read_keywords_from_file(Path(file_arg).resolve()))
    for csv_arg in args.keywords_csv or []:
        text = str(csv_arg or "").strip()
        if not text:
            continue
        raw_keywords.extend([part.strip() for part in text.split(",") if part.strip()])

    if standard_input_used and raw_keywords:
        # 同时给了标准词表与自由关键词：把自由关键词作为骨架行追加进去
        for kw in raw_keywords:
            standard_rows.append(empty_standard_word_row(kw))

    if standard_input_used:
        mode = "standard_word_table"
        # 从标准词表行抽取 keyword
        keyword_items: List[dict] = []
        seen: set[str] = set()
        for row in standard_rows:
            keyword = str(row.get("keyword") or "").strip()
            normalized = normalize_keyword_text(keyword)
            if not normalized or normalized in seen:
                continue
            seen.add(normalized)
            keyword_items.append({"keyword": keyword, "keywordNormalized": normalized})
    else:
        mode = "free_keywords"
        keyword_items: List[dict] = []
        seen: set[str] = set()
        for keyword in raw_keywords:
            normalized = normalize_keyword_text(keyword)
            if not normalized or normalized in seen:
                continue
            seen.add(normalized)
            keyword_items.append({"keyword": keyword, "keywordNormalized": normalized})
            standard_rows.append(empty_standard_word_row(keyword))

    if not keyword_items:
        raise SystemExit(
            "未获取到有效关键词。请使用 --keyword / --keyword-file / --keywords-csv / --standard-word-table 提供关键词。"
        )

    keyword_count_before_limit = len(keyword_items)
    max_keywords = max(int(getattr(args, "max_keywords", 0) or 0), 0)
    if max_keywords > 0:
        keyword_items = keyword_items[:max_keywords]

    return {
        "mode": mode,
        "keywords": keyword_items,
        "standardRows": standard_rows,
        "inputPaths": input_paths,
        "maxKeywords": max_keywords,
        "keywordCountBeforeLimit": keyword_count_before_limit,
    }



# --------------------------------------------------------------------------- #
# 抓取（复用 shared_gefei_kd）
# --------------------------------------------------------------------------- #
def resolve_api_key(args: argparse.Namespace) -> Tuple[str, str]:
    direct = str(args.api_key or "").strip()
    if direct:
        return direct, "--api-key"
    env_name = str(args.api_key_env or "").strip()
    if env_name:
        env_val = str(os.getenv(env_name) or "").strip()
        if env_val:
            return env_val, f"env:{env_name}"
    file_path = Path(args.api_key_file).resolve()
    if not file_path.exists():
        raise RuntimeError(f"api_key 文件不存在: {file_path}")
    value = file_path.read_text(encoding="utf-8").strip()
    if not value:
        raise RuntimeError(f"api_key 文件为空: {file_path}")
    return value, str(file_path)


def fetch_all(args: argparse.Namespace, keywords: List[dict]) -> dict:
    api_key, _ = resolve_api_key(args)
    return shared_gefei_kd.fetch_gefei_kd_rows(
        keywords=[item["keyword"] for item in keywords],
        api_base=args.api_base,
        endpoint=args.endpoint,
        api_key=api_key,
        api_key_env=args.api_key_env,
        api_key_file=args.api_key_file,
        auth_mode=args.auth_mode,
        gl=args.gl,
        hl=args.hl,
        force=args.force,
        response_format=args.response_format,
        min_interval_seconds=args.min_interval_seconds,
        max_retries=args.max_retries,
        timeout_seconds=args.timeout_seconds,
    )


# --------------------------------------------------------------------------- #
# 标准化与标准词表
# --------------------------------------------------------------------------- #
def normalize_success_row(item: dict) -> dict:
    """从 shared_gefei_kd 的成功行构造快照行（score 与 gefeiKD 同源）。"""
    keyword = item.get("keyword")
    score = item.get("gefeiKD")
    return {
        "keyword": keyword,
        "score": score,
        "gefeiKD": score,
        "level": item.get("level"),
        "keywordType": item.get("keywordType"),
        "genericScore": item.get("genericScore"),
        "keywordVolume": item.get("keywordVolume"),
        "keywordTrend": item.get("keywordTrend"),
        "linkBudget": item.get("linkBudget"),
        "detailsCount": 0,
        "cached": item.get("cached"),
        "computedAt": item.get("computedAt"),
        "reasons": item.get("reasons") if isinstance(item.get("reasons"), list) else [],
    }


def summarize_rows(*, input_count: int, rows: List[dict], failures: List[dict]) -> dict:
    score_values = [safe_float(row.get("score")) for row in rows if row.get("score") is not None]
    cached_values = [row.get("cached") for row in rows if row.get("cached") is not None]

    level_dist: Dict[str, int] = {}
    type_dist: Dict[str, int] = {}
    for row in rows:
        level = str(row.get("level") or "未知").strip() or "未知"
        level_dist[level] = level_dist.get(level, 0) + 1
        ktype = str(row.get("keywordType") or "unknown").strip() or "unknown"
        type_dist[ktype] = type_dist.get(ktype, 0) + 1

    return {
        "inputCount": input_count,
        "successCount": len(rows),
        "failureCount": len(failures),
        "avgScore": round(sum(score_values) / len(score_values), 4) if score_values else None,
        "cachedHitRate": round(sum(1 for c in cached_values if c) / len(cached_values), 6) if cached_values else None,
        "levelDistribution": level_dist,
        "keywordTypeDistribution": type_dist,
    }


def sort_rows(rows: List[dict]) -> List[dict]:
    return sorted(
        rows,
        key=lambda row: (
            -(safe_float(row.get("score")) if row.get("score") is not None else -1),
            str(row.get("keyword") or ""),
        ),
    )


def build_standard_word_table(
    *,
    standard_rows: List[dict],
    score_by_keyword: Dict[str, Optional[float]],
    input_mode: str,
    input_paths: List[str],
    gefei_summary: dict,
    generated_at: str,
) -> dict:
    """回填 gefeiKD 到标准词表行。

    - 标准词表输入：保留原值，仅覆盖 gefeiKD
    - 自由关键词输入：骨架行仅 keyword，填入 gefeiKD（其余保持 null/空）
    """
    fields = [str(col.get("field")) for col in spec_columns()]
    out_rows: List[dict] = []
    for row in standard_rows:
        out_row = dict(row) if isinstance(row, dict) else {}
        # 确保所有 spec 字段都存在（缺失补 null，但不覆盖已有值）
        for field in fields:
            if field not in out_row:
                out_row[field] = None
        keyword = str(out_row.get("keyword") or "").strip()
        normalized = normalize_keyword_text(keyword)
        out_row["gefeiKD"] = score_by_keyword.get(normalized)
        out_rows.append(out_row)

    return {
        "version": STANDARD_WORD_TABLE_VERSION,
        "meta": {
            "generatedAt": generated_at,
            "standardWordTableVersion": STANDARD_WORD_TABLE_VERSION,
            "scoreFormula": SCORE_FORMULA,
            "source": "check-gefei-kd",
            "inputMode": input_mode,
            "inputPaths": input_paths,
            "gefeiKdSummary": gefei_summary or {},
            "rowCount": len(out_rows),
        },
        "rows": out_rows,
    }


# --------------------------------------------------------------------------- #
# Markdown 报告
# --------------------------------------------------------------------------- #
def _fmt_md_cell(value) -> str:
    if value is None:
        text = "-"
    else:
        text = str(value).strip()
    if not text:
        text = "-"

    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    normalized = re.sub(r"\s*\|\s*", "\n", normalized)
    normalized = normalized.replace(" / ", "\n")
    lines = [line.strip() for line in normalized.split("\n") if line.strip()]
    if not lines:
        lines = ["-"]

    escaped: List[str] = []
    for line in lines:
        line = re.sub(r"([/?&#=_-])", r"\1<wbr>", line)
        escaped.append(line.replace("|", "\\|"))
    return "<br>".join(escaped)


def _md_table(headers: Sequence[str], rows: Sequence[Sequence[str]]) -> List[str]:
    if not rows:
        return ["（无）"]
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(_fmt_md_cell(cell) for cell in row) + " |")
    return lines


def _format_computed_at(value) -> str:
    if value is None:
        return "-"
    if isinstance(value, (int, float)):
        if value <= 0:
            return str(value)
        try:
            return datetime.fromtimestamp(float(value)).isoformat(timespec="seconds")
        except (OverflowError, OSError, ValueError):
            return str(value)
    return _value_to_short_text(value, max_len=30)


def render_report(snapshot: dict) -> str:
    meta = snapshot.get("meta") or {}
    request = meta.get("request") or {}
    summary = meta.get("summary") or {}
    output = meta.get("output") or {}
    input_meta = meta.get("input") or {}

    rows = snapshot.get("rows") or []
    failures = snapshot.get("failures") or []
    swt = snapshot.get("standardWordTable") or {}
    swt_rows = swt.get("rows") or []

    result_rows: List[List[str]] = []
    for row in rows[:TOP_PREVIEW_ROWS]:
        trend_domain = _get_nested(row, "keywordTrend", "domain")
        trend_ratio = _get_nested(row, "keywordTrend", "ratio")
        trend_text = f"{_value_to_short_text(trend_domain, 20)} / {_value_to_short_text(trend_ratio, 10)}"
        target_dr = _get_nested(row, "linkBudget", "targetDr")
        link_budget_text = _value_to_short_text(target_dr, 12)
        reasons = row.get("reasons") or []
        reason_text = "；".join(_value_to_short_text(item, 40) for item in reasons[:3]) if isinstance(reasons, list) else "-"
        result_rows.append(
            [
                _value_to_short_text(row.get("keyword"), 40),
                to_display_number(row.get("score"), digits=0),
                _value_to_short_text(row.get("level"), 10),
                _value_to_short_text(row.get("keywordType"), 12),
                to_display_number(row.get("genericScore"), digits=0),
                to_display_number(row.get("keywordVolume"), digits=0),
                trend_text,
                link_budget_text,
                _value_to_short_text(row.get("cached"), 5),
                _format_computed_at(row.get("computedAt")),
                _value_to_short_text(reason_text, 100),
            ]
        )

    failure_rows: List[List[str]] = []
    for row in failures:
        failure_rows.append(
            [
                _value_to_short_text(row.get("keyword"), 40),
                str(row.get("httpStatus") or "-"),
                _value_to_short_text(row.get("errorCode"), 12),
                _value_to_short_text(row.get("error"), 80),
                str(row.get("attempts") or "-"),
            ]
        )

    swt_preview_rows: List[List[str]] = []
    for row in swt_rows[:TOP_PREVIEW_ROWS]:
        swt_preview_rows.append(
            [
                _value_to_short_text(row.get("keyword"), 40),
                _value_to_short_text(row.get("correspondingDomain"), 24),
                _value_to_short_text(row.get("group"), 24),
                to_display_number(row.get("gefeiKD")),
            ]
        )

    lines: List[str] = []
    lines.append(f"# 哥飞KD批量查询报告（{meta.get('stamp', '-')}）")
    lines.append("")

    lines.append("## 摘要")
    lines.append("")
    lines.append(f"- 输入模式：{input_meta.get('mode', '-')}")
    lines.append(f"- 输入关键词数：{summary.get('inputCount', 0)}")
    lines.append(f"- 成功关键词数：{summary.get('successCount', 0)}")
    lines.append(f"- 失败关键词数：{summary.get('failureCount', 0)}")
    lines.append(f"- 平均 score：{to_display_number(summary.get('avgScore'))}")
    cached_rate = summary.get("cachedHitRate")
    lines.append(f"- cached 命中率：{to_display_number(cached_rate * 100 if cached_rate is not None else None)}%")
    level_dist = summary.get("levelDistribution") or {}
    if isinstance(level_dist, dict) and level_dist:
        dist_text = ", ".join(f"{k}:{v}" for k, v in sorted(level_dist.items(), key=lambda x: x[0]))
        lines.append(f"- level 分布：{dist_text}")
    lines.append("")

    lines.append("## 请求参数")
    lines.append("")
    lines.append(f"- API：{request.get('apiUrl', '-')}")
    lines.append(f"- 鉴权方式：{request.get('authMode', '-')}（key 来源：{request.get('apiKeySource', '-')}）")
    lines.append(f"- gl={request.get('gl', '-')} hl={request.get('hl', '-')} force={request.get('force', '-')} format={request.get('responseFormat', '-')}")
    lines.append(f"- minIntervalSeconds={request.get('minIntervalSeconds', '-')} maxRetries={request.get('maxRetries', '-')} timeoutSeconds={request.get('timeoutSeconds', '-')}")
    lines.append("")

    lines.append("## 关键词结果明细")
    lines.append("")
    lines.extend(
        _md_table(
            [
                "keyword",
                "score",
                "level",
                "keywordType",
                "genericScore",
                "keywordVolume",
                "keywordTrend(domain/ratio)",
                "linkBudget(targetDr)",
                "cached",
                "computedAt",
                "reasons",
            ],
            result_rows,
        )
    )
    lines.append("")

    lines.append("## 失败明细")
    lines.append("")
    lines.extend(_md_table(["keyword", "status", "errorCode", "error", "attempts"], failure_rows))
    lines.append("")

    global_error = meta.get("globalError")
    if isinstance(global_error, dict) and global_error:
        lines.append("- 全局错误：")
        lines.append(f"  - status={global_error.get('httpStatus', '-')}, code={global_error.get('errorCode', '-')}, error={global_error.get('error', '-')}")
        lines.append("")

    lines.append("## 标准词表")
    lines.append("")
    lines.append(f"- 版本：{swt.get('version', STANDARD_WORD_TABLE_VERSION)}")
    lines.append(f"- 行数：{len(swt_rows)}")
    lines.append(f"- JSON：{output.get('standardWordTableJsonPath', '-')}")
    lines.append(f"- XLSX：{output.get('standardWordTableXlsxPath', '-')}")
    lines.append("")
    lines.extend(
        _md_table(
            ["keyword", "对应域名", "group", "gefeiKD"],
            swt_preview_rows,
        )
    )
    lines.append("")

    lines.append("## 产物路径")
    lines.append("")
    lines.append(f"- Markdown 报告：{output.get('reportHistoryPath', '-')}")
    lines.append(f"- 标准词表 JSON：{output.get('standardWordTableJsonPath', '-')}")
    lines.append(f"- 标准词表 XLSX：{output.get('standardWordTableXlsxPath', '-')}")
    lines.append("")

    lines.append("## 备注")
    lines.append("")
    for remark in REQUIRED_REMARKS:
        lines.append(f"- {remark}")
    lines.append("")

    return "\n".join(lines)


# --------------------------------------------------------------------------- #
# XLSX 标准词表导出
# --------------------------------------------------------------------------- #
DEFAULT_COLUMN_PADDING = 2
DEFAULT_COLUMN_MIN_WIDTH = 12
DEFAULT_COLUMN_MAX_WIDTH = 64

SIM_COLUMN_FILL_COLOR = "FFEAF4FF"
SEM_COLUMN_FILL_COLOR = "FFF4EAFF"


def _transform_export_value(column_def: dict, value, row: Optional[dict] = None):
    transform = column_def.get("transform")
    if transform == "source_presence_label":
        mapping = {
            "both": "both（SIM+SEM）",
            "sim_only": "sim_only（仅 SIM）",
            "sem_only": "sem_only（仅 SEM）",
        }
        transformed = mapping.get(str(value or ""), value)
    else:
        transformed = value
    field = str(column_def.get("field") or "")
    return _format_multiline_export_text(field, transformed)


def _compute_column_width(values: Sequence[str], width_profile: Optional[dict] = None) -> float:
    profile = width_profile or {}
    fixed = profile.get("fixed")
    if fixed is not None:
        return float(fixed)

    max_length = 0
    for raw in values:
        text = "" if raw is None else str(raw)
        parts = text.splitlines() or [text]
        longest_line = max((len(part) for part in parts), default=0)
        if longest_line > max_length:
            max_length = longest_line

    width = max_length + DEFAULT_COLUMN_PADDING
    min_width = int(profile.get("min", DEFAULT_COLUMN_MIN_WIDTH))
    max_width = int(profile.get("max", DEFAULT_COLUMN_MAX_WIDTH))
    return float(max(min_width, min(width, max_width)))


def write_standard_word_table_xlsx(standard_word_table: dict, output_path: Path) -> None:
    Workbook, Font, Alignment, PatternFill, get_column_letter, _ = _require_openpyxl()

    meta = standard_word_table.get("meta") or {}
    rows = standard_word_table.get("rows") or []
    columns = spec_columns()
    headers = [str(col.get("header")) for col in columns]

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    keywords_sheet = workbook.create_sheet("keywords")

    summary_rows = [
        ("generatedAt", meta.get("generatedAt", "")),
        ("standardWordTableVersion", meta.get("standardWordTableVersion", STANDARD_WORD_TABLE_VERSION)),
        ("scoreFormula", meta.get("scoreFormula", SCORE_FORMULA)),
        ("source", meta.get("source", "check-gefei-kd")),
        ("inputMode", meta.get("inputMode", "")),
        ("rowCount", meta.get("rowCount", len(rows))),
        ("gefeiKdSummary", json.dumps(meta.get("gefeiKdSummary") or {}, ensure_ascii=False)),
    ]
    for idx, (name, value) in enumerate(summary_rows, start=1):
        summary_sheet.cell(row=idx, column=1, value=name)
        summary_sheet.cell(row=idx, column=2, value=value)
    summary_sheet.freeze_panes = "A2"

    keywords_sheet.append(headers)
    for cell in keywords_sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    field_to_index = {str(column.get("field") or ""): index for index, column in enumerate(columns, start=1)}
    sim_fill = PatternFill(fill_type="solid", fgColor=SIM_COLUMN_FILL_COLOR)
    sem_fill = PatternFill(fill_type="solid", fgColor=SEM_COLUMN_FILL_COLOR)
    sim_field_indexes = [
        field_to_index[field]
        for field in ("simWindowVolume", "simKd", "simCpc")
        if field in field_to_index
    ]
    sem_field_indexes = [
        field_to_index[field]
        for field in ("semVolume", "semKd", "semCpc")
        if field in field_to_index
    ]

    for row in rows:
        export_row = []
        for column in columns:
            field = str(column.get("field"))
            raw_value = row.get(field)
            export_row.append(_transform_export_value(column, raw_value, row))
        keywords_sheet.append(export_row)

    for row_cells in keywords_sheet.iter_rows(min_row=2, max_row=keywords_sheet.max_row, min_col=1, max_col=len(headers)):
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row_index in range(2, keywords_sheet.max_row + 1):
        for col_index in sim_field_indexes:
            keywords_sheet.cell(row=row_index, column=col_index).fill = sim_fill
        for col_index in sem_field_indexes:
            keywords_sheet.cell(row=row_index, column=col_index).fill = sem_fill

    keywords_sheet.freeze_panes = "A2"
    keywords_sheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(headers))}{max(keywords_sheet.max_row, 1)}"
    )

    for index, column in enumerate(columns, start=1):
        number_format = column.get("number_format")
        if not number_format:
            continue
        col_letter = get_column_letter(index)
        for cell in keywords_sheet[col_letter][1:]:
            cell.number_format = number_format

    for column_index, column in enumerate(columns, start=1):
        column_letter = get_column_letter(column_index)
        values = []
        for row_index in range(1, keywords_sheet.max_row + 1):
            value = keywords_sheet.cell(row=row_index, column=column_index).value
            values.append("" if value is None else str(value))
        keywords_sheet.column_dimensions[column_letter].width = _compute_column_width(
            values, column.get("width_profile")
        )

    for column_cells in summary_sheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        summary_sheet.column_dimensions[column_cells[0].column_letter].width = _compute_column_width(
            values, {"min": 14, "max": 48}
        )

    ensure_dir(output_path.parent)
    workbook.save(output_path)


# --------------------------------------------------------------------------- #
# 快照与产物
# --------------------------------------------------------------------------- #
def build_snapshot(
    *,
    stamp: str,
    args: argparse.Namespace,
    input_info: dict,
    rows: List[dict],
    failures: List[dict],
    summary: dict,
    api_url: str,
    api_key_source: str,
    global_error: Optional[dict],
    standard_word_table: dict,
    report_history_path: Path,
    standard_word_table_json_path: Path,
    standard_word_table_xlsx_path: Path,
) -> dict:
    return {
        "meta": {
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "stamp": stamp,
            "input": {
                "mode": input_info["mode"],
                "paths": input_info["inputPaths"],
                "standardWordTableVersion": STANDARD_WORD_TABLE_VERSION,
                "keywordCount": len(input_info["keywords"]),
                "keywordCountBeforeLimit": int(input_info.get("keywordCountBeforeLimit") or len(input_info["keywords"])),
                "maxKeywords": int(input_info.get("maxKeywords") or 0),
            },
            "target": {
                "keywords": [item["keyword"] for item in input_info["keywords"]],
                "keywordCount": len(input_info["keywords"]),
            },
            "request": {
                "apiBase": args.api_base,
                "endpoint": args.endpoint,
                "apiUrl": api_url,
                "gl": args.gl,
                "hl": args.hl,
                "force": int(args.force),
                "responseFormat": args.response_format,
                "authMode": args.auth_mode,
                "apiKeySource": api_key_source,
                "minIntervalSeconds": float(args.min_interval_seconds),
                "maxRetries": int(args.max_retries),
                "timeoutSeconds": int(args.timeout_seconds),
            },
            "summary": summary,
            "globalError": global_error,
            "output": {
                "reportHistoryPath": str(report_history_path),
                "standardWordTableJsonPath": str(standard_word_table_json_path),
                "standardWordTableXlsxPath": str(standard_word_table_xlsx_path),
            },
        },
        "rows": rows,
        "failures": failures,
        "standardWordTable": standard_word_table,
    }


def write_artifacts(
    *,
    stamp: str,
    snapshot: dict,
    fetch_archive: dict,
    standard_word_table: dict,
    data_dir: Path,
    snapshot_dir: Path,
    report_dir: Path,
    latest_report_path: Path,
    latest_standard_word_table_json: Path,
    latest_standard_word_table_xlsx: Path,
    no_xlsx: bool,
) -> dict:
    ensure_dir(data_dir)
    ensure_dir(snapshot_dir)
    ensure_dir(report_dir)
    ensure_dir(latest_report_path.parent)
    ensure_dir(latest_standard_word_table_json.parent)
    ensure_dir(latest_standard_word_table_xlsx.parent)

    fetch_archive_path = data_dir / f"fetch-{stamp}.json"
    snapshot_path = snapshot_dir / f"snapshot-{stamp}.json"
    report_history_path = report_dir / f"report-{stamp}.md"
    swt_json_history_path = report_dir / f"keyword-table-{stamp}.json"
    swt_xlsx_history_path = report_dir / f"keyword-table-{stamp}.xlsx"

    snapshot_meta = snapshot.setdefault("meta", {})
    output_meta = snapshot_meta.setdefault("output", {})
    output_meta["reportHistoryPath"] = str(report_history_path)
    output_meta["standardWordTableJsonPath"] = str(swt_json_history_path)
    output_meta["standardWordTableXlsxPath"] = str(swt_xlsx_history_path)
    snapshot_meta["latestReportPath"] = str(latest_report_path)
    snapshot_meta["latestStandardWordTableJsonPath"] = str(latest_standard_word_table_json)
    snapshot_meta["latestStandardWordTableXlsxPath"] = str(latest_standard_word_table_xlsx)

    report_text = render_report(snapshot)

    dump_json(fetch_archive_path, fetch_archive)
    dump_json(snapshot_path, snapshot)
    dump_json(swt_json_history_path, standard_word_table)
    report_history_path.write_text(report_text, encoding="utf-8")
    latest_report_path.write_text(report_text, encoding="utf-8")

    xlsx_written = False
    if not no_xlsx:
        write_standard_word_table_xlsx(standard_word_table, swt_xlsx_history_path)
        try:
            from shutil import copyfile

            copyfile(swt_xlsx_history_path, latest_standard_word_table_xlsx)
            xlsx_written = True
        except Exception:
            xlsx_written = False

    # latest JSON
    dump_json(latest_standard_word_table_json, standard_word_table)

    return {
        "fetchArchivePath": fetch_archive_path,
        "snapshotPath": snapshot_path,
        "reportHistoryPath": report_history_path,
        "standardWordTableJsonPath": swt_json_history_path,
        "standardWordTableXlsxPath": swt_xlsx_history_path if not no_xlsx else None,
        "xlsxWritten": xlsx_written,
    }


def rebuild_history_artifact(snapshot: dict, report_dir: Path, no_xlsx: bool) -> Tuple[Path, Optional[Path], str]:
    ensure_dir(report_dir)
    meta = snapshot.get("meta") or {}
    stamp = str(meta.get("stamp") or "")
    if not stamp:
        raise RuntimeError("snapshot 缺少 meta.stamp")

    report_history_path = report_dir / f"report-{stamp}.md"
    swt_json_history_path = report_dir / f"keyword-table-{stamp}.json"
    swt_xlsx_history_path = report_dir / f"keyword-table-{stamp}.xlsx"

    snapshot_meta = snapshot.setdefault("meta", {})
    output_meta = snapshot_meta.setdefault("output", {})
    output_meta["reportHistoryPath"] = str(report_history_path)
    output_meta["standardWordTableJsonPath"] = str(swt_json_history_path)
    output_meta["standardWordTableXlsxPath"] = str(swt_xlsx_history_path)

    report_text = render_report(snapshot)
    report_history_path.write_text(report_text, encoding="utf-8")

    standard_word_table = snapshot.get("standardWordTable")
    if isinstance(standard_word_table, dict):
        dump_json(swt_json_history_path, standard_word_table)
        if not no_xlsx:
            try:
                write_standard_word_table_xlsx(standard_word_table, swt_xlsx_history_path)
            except RuntimeError:
                # openpyxl 缺失时跳过 XLSX，仅保留 JSON
                return report_history_path, None, report_text
            return report_history_path, swt_xlsx_history_path, report_text
        return report_history_path, None, report_text

    return report_history_path, None, report_text


# --------------------------------------------------------------------------- #
# 主流程
# --------------------------------------------------------------------------- #
def run_pipeline(args: argparse.Namespace) -> int:
    data_dir = Path(args.data_dir).resolve()
    snapshot_dir = Path(args.snapshot_dir).resolve()
    report_dir = Path(args.report_dir).resolve()
    latest_report_path = Path(args.latest_report_path).resolve()
    latest_swt_json = Path(args.latest_standard_word_table_json).resolve()
    latest_swt_xlsx = Path(args.latest_standard_word_table_xlsx).resolve()

    api_url = args.api_base.rstrip("/") + args.endpoint

    input_info = collect_input(args)
    api_key, api_key_source = resolve_api_key(args)

    if args.dry_run:
        print("[dry-run] 配置检查通过，不发起网络请求。")
        print(f"[dry-run] input-mode: {input_info['mode']}")
        print(f"[dry-run] keywords(before-limit): {input_info.get('keywordCountBeforeLimit', len(input_info['keywords']))}")
        print(f"[dry-run] keywords(after-limit): {len(input_info['keywords'])}")
        print(f"[dry-run] max-keywords: {input_info.get('maxKeywords', 0)}")
        print(f"[dry-run] api: {api_url}")
        print(f"[dry-run] auth-mode: {args.auth_mode}")
        print(f"[dry-run] api-key-source: {api_key_source}")
        print(f"[dry-run] api-key(masked): {mask_secret(api_key)}")
        print(f"[dry-run] gl={args.gl} hl={args.hl} force={args.force} format={args.response_format}")
        print(f"[dry-run] min-interval={args.min_interval_seconds}s max-retries={args.max_retries}")
        return 0

    fetch_result = fetch_all(args, input_info["keywords"])
    score_by_keyword = fetch_result.get("scoreByKeyword") or {}
    raw_rows = fetch_result.get("rows") or []
    failures_raw = fetch_result.get("failures") or []

    # shared_gefei_kd 会把失败行同时追加到 rows 与 failures；用 failures 集合区分成功/失败
    failure_norms = {
        normalize_keyword_text(str(f.get("keyword") or "")) for f in failures_raw
    }

    success_rows: List[dict] = []
    failures: List[dict] = []
    global_error: Optional[dict] = None

    for item in raw_rows:
        normalized = normalize_keyword_text(str(item.get("keyword") or ""))
        if normalized in failure_norms:
            failures.append(
                {
                    "keyword": item.get("keyword"),
                    "keywordNormalized": item.get("keywordNormalized"),
                    "attempts": len(item.get("attempts") or []),
                    "error": str(item.get("error") or "未知错误"),
                    "errorCode": item.get("errorCode"),
                    "httpStatus": item.get("httpStatus"),
                    "fatalGlobal": bool(item.get("fatalGlobal")),
                }
            )
            if item.get("fatalGlobal") and global_error is None:
                global_error = {
                    "keyword": item.get("keyword"),
                    "httpStatus": item.get("httpStatus"),
                    "errorCode": item.get("errorCode"),
                    "error": str(item.get("error") or "未知错误"),
                }
        else:
            success_rows.append(normalize_success_row(item))

    sorted_rows = sort_rows(success_rows)
    summary = summarize_rows(input_count=len(input_info["keywords"]), rows=sorted_rows, failures=failures)

    stamp = now_stamp()
    generated_at = datetime.now().isoformat(timespec="seconds")
    standard_word_table = build_standard_word_table(
        standard_rows=input_info["standardRows"],
        score_by_keyword=score_by_keyword,
        input_mode=input_info["mode"],
        input_paths=input_info["inputPaths"],
        gefei_summary=fetch_result.get("summary") or {},
        generated_at=generated_at,
    )

    report_history_path = report_dir / f"report-{stamp}.md"
    swt_json_history_path = report_dir / f"keyword-table-{stamp}.json"
    swt_xlsx_history_path = report_dir / f"keyword-table-{stamp}.xlsx"

    snapshot = build_snapshot(
        stamp=stamp,
        args=args,
        input_info=input_info,
        rows=sorted_rows,
        failures=failures,
        summary=summary,
        api_url=api_url,
        api_key_source=api_key_source,
        global_error=global_error,
        standard_word_table=standard_word_table,
        report_history_path=report_history_path,
        standard_word_table_json_path=swt_json_history_path,
        standard_word_table_xlsx_path=swt_xlsx_history_path,
    )

    fetch_archive = {
        "meta": {
            "generatedAt": generated_at,
            "stamp": stamp,
            "input": snapshot["meta"]["input"],
            "summary": summary,
            "request": {
                "apiUrl": api_url,
                "gl": args.gl,
                "hl": args.hl,
                "force": int(args.force),
                "responseFormat": args.response_format,
                "authMode": args.auth_mode,
                "apiKeySource": api_key_source,
                "minIntervalSeconds": float(args.min_interval_seconds),
                "maxRetries": int(args.max_retries),
                "timeoutSeconds": int(args.timeout_seconds),
            },
        },
        "requests": raw_rows,
    }

    artifact_paths = write_artifacts(
        stamp=stamp,
        snapshot=snapshot,
        fetch_archive=fetch_archive,
        standard_word_table=standard_word_table,
        data_dir=data_dir,
        snapshot_dir=snapshot_dir,
        report_dir=report_dir,
        latest_report_path=latest_report_path,
        latest_standard_word_table_json=latest_swt_json,
        latest_standard_word_table_xlsx=latest_swt_xlsx,
        no_xlsx=args.no_xlsx,
    )

    print(f"[done] fetch archive        : {artifact_paths['fetchArchivePath']}")
    print(f"[done] snapshot             : {artifact_paths['snapshotPath']}")
    print(f"[done] report history       : {artifact_paths['reportHistoryPath']}")
    print(f"[done] report latest        : {latest_report_path}")
    print(f"[done] standard word table(json): {artifact_paths['standardWordTableJsonPath']}")
    if artifact_paths.get("standardWordTableXlsxPath"):
        print(f"[done] standard word table(xlsx): {artifact_paths['standardWordTableXlsxPath']}")
    else:
        print("[done] standard word table(xlsx): skipped (--no-xlsx 或 openpyxl 不可用)")
    print(f"[done] standard word table(latest json): {latest_swt_json}")
    print(f"[done] standard word table(latest xlsx): {latest_swt_xlsx}")
    print(f"[done] input keywords: {summary['inputCount']}")
    print(f"[done] success count : {summary['successCount']}")
    print(f"[done] failure count : {summary['failureCount']}")

    if global_error:
        print(
            f"[warn] 全局错误：status={global_error.get('httpStatus')} code={global_error.get('errorCode')} error={global_error.get('error')}"
        )
        return 2

    return 0


def rebuild_reports(args: argparse.Namespace) -> int:
    snapshot_dir = Path(args.snapshot_dir).resolve()
    report_dir = Path(args.report_dir).resolve()
    latest_report_path = Path(args.latest_report_path).resolve()
    latest_swt_json = Path(args.latest_standard_word_table_json).resolve()
    latest_swt_xlsx = Path(args.latest_standard_word_table_xlsx).resolve()

    ensure_dir(snapshot_dir)
    ensure_dir(report_dir)
    ensure_dir(latest_report_path.parent)
    ensure_dir(latest_swt_json.parent)
    ensure_dir(latest_swt_xlsx.parent)

    files = list_snapshot_files(snapshot_dir)
    if not files:
        raise SystemExit(f"未找到快照文件: {snapshot_dir}")

    latest_report_text = ""
    latest_swt: Optional[dict] = None
    for path in files:
        snapshot = load_json(path)
        report_history_path, xlsx_path, report_text = rebuild_history_artifact(snapshot, report_dir, args.no_xlsx)
        latest_report_text = report_text
        if isinstance(snapshot.get("standardWordTable"), dict):
            latest_swt = snapshot["standardWordTable"]
        print(f"[rebuild] report: {report_history_path}")
        if xlsx_path:
            print(f"[rebuild] standard word table(xlsx): {xlsx_path}")

    latest_report_path.write_text(latest_report_text, encoding="utf-8")
    print(f"[done] latest report: {latest_report_path}")

    if latest_swt is not None:
        dump_json(latest_swt_json, latest_swt)
        print(f"[done] latest standard word table(json): {latest_swt_json}")
        if not args.no_xlsx:
            try:
                write_standard_word_table_xlsx(latest_swt, latest_swt_xlsx)
                print(f"[done] latest standard word table(xlsx): {latest_swt_xlsx}")
            except RuntimeError:
                print("[done] latest standard word table(xlsx): skipped（openpyxl 不可用）")

    return 0


def validate_report(args: argparse.Namespace) -> int:
    report_path = Path(args.report).resolve()
    xlsx_path = Path(args.xlsx).resolve()
    if not report_path.exists():
        raise SystemExit(f"报告不存在: {report_path}")

    text = report_path.read_text(encoding="utf-8")
    missing_sections = [section for section in REQUIRED_SECTIONS if section not in text]
    missing_remarks = [remark for remark in REQUIRED_REMARKS if remark not in text]

    if missing_sections or missing_remarks:
        if missing_sections:
            print("缺少 section:")
            for item in missing_sections:
                print(f"- {item}")
        if missing_remarks:
            print("缺少 remark:")
            for item in missing_remarks:
                print(f"- {item}")
        raise SystemExit(1)

    if not xlsx_path.exists():
        raise SystemExit(f"Excel 不存在: {xlsx_path}")

    _Workbook, _Font, _Alignment, _PatternFill, _get_column_letter, load_workbook = _require_openpyxl()
    workbook = load_workbook(xlsx_path, data_only=True)
    try:
        if "keywords" not in workbook.sheetnames:
            raise SystemExit("Excel 缺少 keywords sheet")
        sheet = workbook["keywords"]

        expected_headers = [str(col.get("header") or "") for col in spec_columns()]
        actual_headers = [sheet.cell(row=1, column=index).value for index in range(1, sheet.max_column + 1)]
        if actual_headers != expected_headers:
            print("Excel 表头不匹配：")
            max_len = max(len(expected_headers), len(actual_headers))
            for index in range(1, max_len + 1):
                expected = expected_headers[index - 1] if index <= len(expected_headers) else None
                actual = actual_headers[index - 1] if index <= len(actual_headers) else None
                if expected != actual:
                    print(f"- 第 {index} 列 expected={expected!r} actual={actual!r}")
            raise SystemExit(1)

        keyword_col = expected_headers.index(get_header_for_field("keyword")) + 1
        score_col = get_score_column_index_from_headers(expected_headers) + 1
        numeric_fields = [
            "simWindowVolume",
            "simKd",
            "simCpc",
            "semVolume",
            "semKd",
            "semCpc",
            "gefeiKD",
        ]
        numeric_indexes = [expected_headers.index(get_header_for_field(field)) + 1 for field in numeric_fields]

        def _is_numeric_cell(value) -> bool:
            return isinstance(value, (int, float)) and not isinstance(value, bool)

        for row_index in range(2, sheet.max_row + 1):
            keyword = sheet.cell(row=row_index, column=keyword_col).value
            if keyword in (None, ""):
                raise SystemExit(f"keyword 列为空（row={row_index}）")

            score_value = sheet.cell(row=row_index, column=score_col).value
            if score_value not in (None, "") and not _is_numeric_cell(score_value):
                raise SystemExit(f"score 列应为数值类型而非文本（row={row_index}）")

            for col_index in numeric_indexes:
                value = sheet.cell(row=row_index, column=col_index).value
                if value in (None, ""):
                    continue
                if not _is_numeric_cell(value):
                    raise SystemExit(f"数字列应为数值类型而非文本（row={row_index}, col={col_index}）")
    finally:
        workbook.close()

    print(f"[ok] report validated: {report_path}")
    print(f"[ok] xlsx validated  : {xlsx_path}")
    return 0


def list_snapshot_files(snapshot_dir: Path) -> List[Path]:
    if not snapshot_dir.exists():
        return []
    files = []
    for path in snapshot_dir.iterdir():
        if path.is_file() and SNAPSHOT_RE.match(path.name):
            files.append(path)
    files.sort(key=lambda p: p.name)
    return files


def main() -> int:
    args = parse_args()
    if args.command == "run":
        return run_pipeline(args)
    if args.command == "rebuild-reports":
        return rebuild_reports(args)
    if args.command == "validate-report":
        return validate_report(args)
    raise SystemExit(f"未知命令: {args.command}")


if __name__ == "__main__":
    raise SystemExit(main())
